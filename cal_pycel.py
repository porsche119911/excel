from openpyxl import load_workbook, Workbook
from pycel import ExcelCompiler


def process_workbook(source_filename, result_filename):
    # 加载源Excel文件
    source_wb = load_workbook(source_filename)

    # 创建ExcelCompiler对象
    excel_compiler = ExcelCompiler(filename=source_filename)

    # 创建新的工作簿来保存结果
    result_wb = Workbook()

    # 删除结果工作簿中的默认工作表
    result_wb.remove(result_wb.active)

    # 遍历源工作簿中的所有工作表
    for sheet_name in source_wb.sheetnames:
        source_ws = source_wb[sheet_name]

        # 在结果工作簿中创建相同名称的工作表
        result_ws = result_wb.create_sheet(title=sheet_name)

        # 遍历当前工作表中的单元格
        ranges=[(1,92,1,11),(143,154,1,9),(108,124,1,11)]
        ranges=[(1,451,1,11)]
        for range in ranges:
            for row in source_ws.iter_rows(min_row=range[0],max_row=range[1],min_col=range[2],max_col=range[3]):

                for cell in row:
                    # 获取单元格地址
                    address = f"'{sheet_name}'!{cell.coordinate}"
                    formula = source_ws[cell.coordinate].value

                    # 使用pycel计算单元格值
                    try:
                        calculated_value = excel_compiler.evaluate(address)
                        print(calculated_value,address)
                    except:
                        print('failed,',formula,calculated_value,address)
                        calculated_value = cell.value  # 如果计算失败，使用原始值

                    # 将计算结果写入新工作表的相同位置
                    result_ws[cell.coordinate] = calculated_value

    # 保存结果到新的Excel文件
    result_wb.save(result_filename)
    print(f"计算完成，结果已保存到 {result_filename}")


# 使用示例
source_file = 'cal_教材检测_0000_2024_09_14_15_29_27.xlsx'
result_file = f'result_{source_file}'
process_workbook(source_file, result_file)