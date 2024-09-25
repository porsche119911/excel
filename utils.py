from openpyxl import load_workbook
import openpyxl
from openpyxl.utils import get_column_letter

def compare_excel_files(file1_path, file2_path):
    wb1 = openpyxl.load_workbook(file1_path, data_only=True)
    wb2 = openpyxl.load_workbook(file2_path, data_only=True)

    differences = []

    for sheet_name in wb1.sheetnames:
        if sheet_name in wb2.sheetnames:
            sheet1 = wb1[sheet_name]
            sheet2 = wb2[sheet_name]

            for row in range(1, max(sheet1.max_row, sheet2.max_row) + 1):
                for col in range(1, max(sheet1.max_column, sheet2.max_column) + 1):
                    cell1 = sheet1.cell(row=row, column=col)
                    cell2 = sheet2.cell(row=row, column=col)

                    if cell1.value != cell2.value:
                        col_letter = get_column_letter(col)
                        differences.append(f"工作表: {sheet_name}, 单元格: {col_letter}{row}, "
                                           f"文件1值: {cell1.value}, 文件2值: {cell2.value}")

    return differences




def extract_formula(excel_path, sheet_names, start_coordinate = None, end_coordinate = None, None_permitted = False):
    wb = load_workbook(excel_path, data_only=False)
    sheet_formulas = {}
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        sheet_formulas[sheet_name] = {}
        if (start_coordinate is None) or (end_coordinate is None):
            # 遍历走所有的内容
            for row in ws.columns:
                for cell in row:
                    if not (cell.value is None) or None_permitted:
                        sheet_formulas[sheet_name][cell.coordinate]=cell.value
        else:
            cell_range = ws[start_coordinate:end_coordinate]
            for cell_tuple in cell_range:
                for cell in cell_tuple:
                    sheet_formulas[sheet_name][cell.coordinate] = cell.value

    return sheet_formulas




if __name__ == '__main__':

    list_names = []
    sheetIndexs = {'2': '附表2流动资金估算', '3': '附表3借款还本付息表', '4': '附表4投资使用计划与资金筹措表',
                   '5': '附表5固定资产折旧表', '6': '附表6无形资产及其他资产摊销表', '7': '附表7总成本估算表',
                   '8': '附表8营业收入、营业税金及附加表', '9': '附表9资金来源与运用表1', '10': '附表10损益表',
                   '11': '附表11项目全部投资现金流量表', '12': '附表11-2项目资本金现金流量表', '13': '附表12资产负债','14': '汇总表'
                   }
    for i in range(14,15):
        ind = str(i)
        sheet_names = sheetIndexs[ind]
        list_names.append(sheet_names)
        listing = extract_formula("副本2024-(融资后724）(1) - 副本.xlsx", list_names)[sheet_names].items()
        ft = {}
        for ft_key, ft_value in listing:
                ft[ft_key] = ft_value
        print(ft)
    # from openpyxl.utils import FORMULAE
    # print('NPV' in FORMULAE)
    #
    # from openpyxl.utils import FORMULAE
    #
    # from openpyxl.formula.translate import Translator
    # print(Translator("= (SUM($D$14:C14) + SUM($D$13: D13))*'基础数据4-表2、3、5、6、7、8、10'!$C$20",'D14').translate_formula('E14'))

    # from openpyxl import load_workbook

    # 打开源工作簿
    # source_wb = load_workbook('副本2024-(融资后724）(1).xlsx')
    # counter = -1
    # ts = {}
    # for i in range(2,14):
    #     ind = str(i)
    #     ts[ind] = []

    # # 复制所有工作表
    # for sheet_name in source_wb.sheetnames:
    #     counter += 1
    #     ind = str(counter)
    #     ts[ind] = sheet_name
    # print(ts)


    # # 使用示例
    # file1_path = '副本自用提取模型.xlsx'
    # file2_path = '示例_0000_2024_08_05_14_11_50.xlsx'
    #
    # differences = compare_excel_files(file1_path, file2_path)
    #
    # if differences:
    #     print("发现以下差异:")
    #     for diff in differences:
    #         print(diff)
    # else:
    #     print("两个文件完全一致")

