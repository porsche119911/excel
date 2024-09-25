# coding:utf-8
from openpyxl.utils import range_boundaries
from openpyxl.styles import Alignment
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string, coordinate_to_tuple
import ast
from openpyxl.formula.translate import Translator
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl import load_workbook, Workbook
from pycel import ExcelCompiler
import json
import copy
import re


def process_workbook(source_filename, result_filename,sheetmapping):
    # 加载源Excel文件
    source_wb = load_workbook(source_filename)

    # 创建ExcelCompiler对象
    excel_compiler = ExcelCompiler(filename=source_filename)

    # 创建新的工作簿来保存结果
    result_wb = Workbook()

    # 删除结果工作簿中的默认工作表
    result_wb.remove(result_wb.active)

    sheetmaps = list(sheetmapping.values())
    # print(sheetmaps)
    # 设定计算格式,sheetmaps索引等于表数-1
    percentage = [(sheetmaps[0]+2,sheetmaps[0]+13,8,8), (sheetmaps[2]+3,sheetmaps[2]+3,3,3),
                  (sheetmaps[4]+3,sheetmaps[4]+7,5,5),(sheetmaps[6] + 16, sheetmaps[6] + 16, 4, 100),
                  (sheetmaps[10] + 24, sheetmaps[10] + 24, 5, 6),(sheetmaps[11] + 24, sheetmaps[11] + 24, 5, 6)
                  ,(sheetmaps[12] + 23, sheetmaps[12] + 23, 3, 100), (sheetmaps[13] + 11, sheetmaps[13] + 11, 3, 100)]
    percentage_int = [(sheetmaps[6] + 3, sheetmaps[6] + 3, 4, 100),
                      (sheetmaps[7] + 3, sheetmaps[7] + 3, 4, 100),
                      (sheetmaps[8] + 3, sheetmaps[8] + 3, 4, 100),
                      (sheetmaps[9] + 3, sheetmaps[9] + 3, 4, 100),
                      (sheetmaps[10] + 3, sheetmaps[10] + 3, 3, 100),
                      (sheetmaps[11] + 3, sheetmaps[11] + 3, 3, 100),
                      (sheetmaps[12] + 3, sheetmaps[12] + 3, 2, 100),(sheetmaps[14]+2,sheetmaps[14]+16,4,4)
                      ]

    percentage_none = [
        (sheetmaps[15] + 30, sheetmaps[15] + 31, 4, 4),
        (sheetmaps[15] + 34, sheetmaps[15] + 34, 4, 4),
        (sheetmaps[15] + 39, sheetmaps[15] + 39, 4, 4),
        (sheetmaps[15] + 45, sheetmaps[15] + 45, 4, 4),
        (sheetmaps[15] + 50, sheetmaps[15] + 50, 4, 4),
        (sheetmaps[15] + 54, sheetmaps[15] + 54, 4, 4)
    ]

    cal_dot = [(6,6,3,100),
                (sheetmaps[2] + 15, sheetmaps[2] + 16, 5, 100),
               (sheetmaps[10] + 22, sheetmaps[10] + 23, 5, 6),
               (sheetmaps[11] + 22, sheetmaps[11] + 23, 5, 6)
               ,(sheetmaps[12] + 24, sheetmaps[12] + 25, 5, 100),
               (sheetmaps[13] + 9, sheetmaps[13] + 10, 3, 100),
                (sheetmaps[14] + 2, sheetmaps[14] + 16, 2, 2),
               (sheetmaps[15] + 36, sheetmaps[15] + 37, 4, 4),
               (sheetmaps[15] + 41, sheetmaps[15] + 42, 4, 4),
               (sheetmaps[15] + 47, sheetmaps[15] + 48, 4, 4),
               (sheetmaps[15] + 52, sheetmaps[15] + 53, 4, 4),
               (sheetmaps[15] + 56, sheetmaps[15] + 57, 4, 4)
               ]

    def isType(coordination,arr):
        k = 0
        for bound in arr:
            if bound[0]<=coordination[0]<=bound[1] and bound[2]<=coordination[1]<=bound[3]:
                k=1
        return k


    # 遍历源工作簿中的所有工作表
    for sheet_name in source_wb.sheetnames:
        source_ws = source_wb[sheet_name]

        # 在结果工作簿中创建相同名称的工作表
        result_ws = result_wb.create_sheet(title=sheet_name)

        results = {}

        # 遍历当前工作表中的单元格
        for row in source_ws.iter_rows():
            for cell in row:
                # 获取单元格地址
                    address = f"'{sheet_name}'!{cell.coordinate}"
                    coor = coordinate_to_tuple(cell.coordinate)
                    # 使用pycel计算单元格值
                    try:
                        calculated_value = excel_compiler.evaluate(address)
                    except:
                        calculated_value = cell.value  # 如果计算失败，使用原始值

                    # 将计算结果写入新工作表的相同位置
                    if isinstance(calculated_value, str):
                        result_ws[cell.coordinate] = calculated_value
                        results[cell.coordinate] = calculated_value
                    elif isinstance(calculated_value, (int,float)):
                        # 处理普通、百分比、计算小数、百分比计算小数等不同格式
                        if isType(coor,percentage):
                            result_ws[cell.coordinate] = str(round(calculated_value*100, 1))+"%"
                            results[cell.coordinate] = str(round(calculated_value*100, 1))+"%"
                        elif isType(coor,percentage_int):
                            result_ws[cell.coordinate] = str(round(calculated_value*100))+"%"
                            results[cell.coordinate] = str(round(calculated_value*100))+"%"
                        elif isType(coor,cal_dot):
                            result_ws[cell.coordinate] = round(calculated_value,1)
                            results[cell.coordinate] = str(round(calculated_value,1))
                        elif isType(coor,percentage_none):
                            result_ws[cell.coordinate] = round(calculated_value*100, 1)
                            results[cell.coordinate] = str(round(calculated_value*100, 1))
                        else:
                            result_ws[cell.coordinate] = round(calculated_value)
                            results[cell.coordinate] = str(round(calculated_value))
                    # else:
                        # print("特别",address,calculated_value,type(calculated_value))

    # 保存结果到新的Excel文件
    result_wb.save(result_filename)
    return results


def number_to_letter(n):
    # 转换1-26的数字为A-Z
    if 1 <= n <= 26:
        return chr(n + 64)
    else:
        return 'Number out of range'

def df2xspreadsheetjson(df) -> str:
    '''
    df对象转为 x-spreadsheet格式的json字符串
    :param df: 从数据库得到的dataframe
    :return: str
    '''
    cols = []
    # 重命名标题行，主要考虑标题行空等情况
    for col in df.columns:
        if col == "":
            cols.append("N/A")  # 标题行空，一般不可能
        elif col is None:
            cols.append("NULL")  # 表是空的
        else:
            cols.append(col)  # 复制过去
    # print(cols)
    df.columns = cols

    if df.shape[0] < 1 or df.shape[1] < 1:
        return '{}'

    metrics = cols
    df = df[metrics]  # 取我们需要的字段
    # 直接拼接字符串
    info = ''
    # 先拼接标题行
    # info += '\"0\":{\"cells\":{'
    # for i in range(len(metrics)):
    #     if i != len(metrics)-1:
    #         info += '\"' + str(i) + '\":' + '{\"text\":\"' + str(metrics[i]) + "\"},"
    #     else:
    #         info += '\"' + str(i) + '\":' + '{\"text\":\"' + str(metrics[i]) + "\"}"
    # info += '}},'

    for index, row in df.iterrows():
        info += '\"'+str(index) + '\":'
        for j in range(len(metrics)):
            if j == 0:
                info += "{\"cells\":{"
            # print(row[col])
            if j != len(metrics)-1:
                info += '\"' + str(j)+'\":' + '{\"text\":\"' + str(row[metrics[j]]) + "\"},"
            else:
                info += '\"' + str(j) + '\":' + '{\"text\":\"' + str(row[metrics[j]]) + "\"}"
        info += '}},'

    rows = '{' + info + '\"' +"len" + '\"' + ':' + str(df.shape[0]) + '}'  # 加标题行长度，最好再加上一行+1+1
    cols = 'cols\":{\"len\":' + str(df.shape[1]+1) + '}}'
    return rows


# 将特定excel表的内容转换为spreadsheet的json字符串形式（包含为数据添加merge）
def renderexcel(excel_path,sheetmapping,xspreadsheets):  #需要加一个参数
    sheetnums = 16
    output_path = "cal_"+excel_path

    results = process_workbook(excel_path,output_path,sheetmapping)
    # print(sheetmapping)
    print(results)
    # 根据sheetmapping将results插入xspreadsheets
    sheetmaps = list(sheetmapping.values())
    for coor, value in results.items():
        row, col = coordinate_to_tuple(coor)
        sheet = 0
        while sheet <= sheetnums - 1 and row >= sheetmaps[sheet] :
            sheet += 1
        # 特殊处理表5空一行的问题
        if (sheet == 5) and (row - sheetmaps[sheet-1]>=7):
            row = row + 1
        if sheet != 0:
            # if (sheet == 15):
            #     print('debug')

            if row - sheetmaps[sheet-1]<=xspreadsheets[sheet]['rows']["len"]:
                # print(sheet)
                s_row = str(row - sheetmaps[sheet-1])
                s_col = str(col - 1)
                # print(s_row)
                # print(s_col)

                try:

                    xspreadsheets[sheet]['rows'][s_row]["cells"][s_col]["text"] = value
                except KeyError:

                    copy_dict = next(iter(xspreadsheets[sheet]['rows'][s_row]["cells"].values()))
                    xspreadsheets[sheet]['rows'][s_row]["cells"][s_col] = copy.deepcopy(copy_dict)
                    xspreadsheets[sheet]['rows'][s_row]["cells"][s_col]["text"] = value
                # 调整百分数、整数和小数保留位数
    index = 0
    while results[f'{get_column_letter(index+3)}6'] != '1' and index< int(results[f'C4']) - int(results[f'C5']):
        index += 1
    da = index + 1
    totalRevenue = []
    totalCost = []
    y = float(results[f'{get_column_letter(da+2)}{sheetmaps[13]+9}'])*(int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+6}']) - int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+7}']))
    breakEvenPoint = {'x': float(results[f'{get_column_letter(da+2)}{sheetmaps[13]+9}']), 'y': y}
    totalprofit = []
    for i in range(11):
        pro = i/10 * int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+3}'])
        rev = pro*(int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+6}']) - int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+7}']))
        cost = int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+4}'])+pro*int(results[f'{get_column_letter(da+2)}{sheetmaps[13]+5}'])
        profit = rev - cost
        totalRevenue.append({'x': pro, 'y': rev})
        totalCost.append({'x': pro, 'y': cost})
        totalprofit.append({'x': pro, 'y': profit})
    xspreadsheets[17] = {
    'totalRevenue': totalRevenue,
    'totalCost': totalCost,
    'breakEvenPoint': breakEvenPoint,
        'totalprofit': totalprofit
    }
    constructionInvestment = []
    constructionInvestment.append(int(results[f'C{sheetmaps[14]+6}']))
    constructionInvestment.append(int(results[f'C{sheetmaps[14] + 5}']))
    constructionInvestment.append(int(results[f'C{sheetmaps[14] + 2}']))
    constructionInvestment.append(int(results[f'C{sheetmaps[14] + 4}']))
    constructionInvestment.append(int(results[f'C{sheetmaps[14] + 3}']))

    operatingCost = []
    operatingCost.append(int(results[f'C{sheetmaps[14]+11}']))
    operatingCost.append(int(results[f'C{sheetmaps[14] + 10}']))
    operatingCost.append(int(results[f'C{sheetmaps[14] + 7}']))
    operatingCost.append(int(results[f'C{sheetmaps[14] + 9}']))
    operatingCost.append(int(results[f'C{sheetmaps[14] + 8}']))

    salesRevenue = []
    salesRevenue.append(int(results[f'C{sheetmaps[14]+16}']))
    salesRevenue.append(int(results[f'C{sheetmaps[14] + 15}']))
    salesRevenue.append(int(results[f'C{sheetmaps[14] + 12}']))
    salesRevenue.append(int(results[f'C{sheetmaps[14] + 14}']))
    salesRevenue.append(int(results[f'C{sheetmaps[14] + 13}']))
    xspreadsheets[18] = {
    'constructionInvestment': constructionInvestment,
    'operatingCost': operatingCost,
    'salesRevenue': salesRevenue
  }
    return xspreadsheets

def mergeexcel(start_cell, end_cell,worksheet):
    min_col, min_row, max_col, max_row = range_boundaries(
        start_cell + ':' + end_cell)  # 获取合并范围的边界（openpyxl 3.0.0+ 提供了range_boundaries函数）
    worksheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)  # 合并单元格
    return 0

class Fdata:
    def __init__(self, fdata):
        self.fund_data = fdata
        # print(self.fund_data)

        self.invenstment = {
            'tudi' : 0,  # 默认土地开办费是第一个其他费用
            'wx': 0, # 默认第三个是无形费用
            'qt': 0,
        }

        for loop in range(1,int(self.fund_data['othercostCounter'])+1):
            if self.fund_data[f'additionalOtherCostType_{loop}'] == "intangible":
                self.invenstment['wx'] += float(self.fund_data[f'additionalOtherCost_{loop}'])
            if self.fund_data[f'additionalOtherCostType_{loop}'] == "other":
                self.invenstment['qt'] += float(self.fund_data[f'additionalOtherCost_{loop}'])
            if self.fund_data[f'additionalOtherCostType_{loop}'] == "tudi":
                self.invenstment['tudi'] += float(self.fund_data[f'additionalOtherCost_{loop}'])

        #计算无形资产和其他资产总和 if additionalCostType_i == "other": self.total_qt += additionCost_i  else: self.total_wx += additionCost_i

        # 计算达产期
        if self.fund_data['input2'] != '' and self.fund_data['input1'] != '' and self.fund_data['input1'] > self.fund_data['input2']:
            index_d = int(self.fund_data['input2']) + 1
            while self.fund_data[f'productionLoad{str(index_d)}'] != '100':
                index_d += 1
            self.fund_data['input3'] = index_d


        # 转为整形
        int_lists = ['input1', 'input2']
        for i in int_lists:
            if self.fund_data[i] != '':
                self.fund_data[i] = int(self.fund_data[i])

        float_list = []
        # 对于与时间相关的表格，除了时间还有几列
        self.sheetAddCols = {'2':4, '3':4,'4': 3,
                             '5':5, '6':4,
                             '7':3, '8':3,
                             '9':3, '10':3,
                             '11':3,'12':3,
                             '13':2, '14':2}

        # 这几列分别列标题是什么
        self.sheetAddColsNames = {'2': ['序号', '项目', '最低周转天数', '最低周转次数'],
                                  '3': ['序号', '项目','利率','合计'],
                                  '4': ['序号', '项目', '合计'],
                                  '5': ['序号', '项目', '原值', '折旧年限', '残值率'],
                                  '6': ['序号', '项目', '原值', '摊销年限'],
                                  '7': ['序号', '项目', '合计'],
                                  '8': ['序号', '项目', '合计'],
                                  '9': ['序号', '项目', '合计'],
                                  '10': ['序号', '项目', '合计'],
                                  '11': ['序号', '项目', '合计'],
                                  '12': ['序号', '项目', '合计'],
                                  '13': ['序号', '项目'],
                                  '14': ['序号', '项目']}

        # 第二列的名称是什么
        self.sheetDefaultArrows = {'2': ['流动资产', '应收帐款', '存   货', '原材料', '燃料动力', '在产品', '产成品', '现 金', '流动负债',
                                         '应付帐款', '流动资金', '流动资金本年增加额', '流动资金借款（本年）', '流动资金利息'],
                                  '3': [ '借款及还本付息', '年初借款本息和', '当期借款', '当期应计利息', '当期还本付息', '其中：还本',
                                         '付息', '期末借款余额', '偿还借款本金资金来源', '利润', '折旧', '摊销', '利息备付率', '偿债备付率'],
                                  '4': ['资金总额','建设投资','建设期利息','流动资金','资金筹措','项目资本金','用于建设投资','用于流动资金',
                                         '债务资金','用于建设投资','用于建设期利息','用于流动资金'],
                                  '5': [ '建筑工程', '原值', '折旧费', '净值', '设备工程', '原值', '折旧费', '净值', '',
                                         '土地征用费', '原值', '折旧费', '净值', '', '固定资产合计', '原值', '折旧费',
                                         '净值', '', '建筑工程包含预备费和建设期利息', '设备工程包含安装工程费'],
                                  '6': [ '无形资产', '摊销费', '净值', '其他资产', '摊销费', '净值', '无形资产及其他资产合计', '摊销合计', '净值合计'],
                                  '7': ['生产负荷（%）', '外购原材料', '外购燃料动力', '工资及福利费', '修理费', '折旧费', '摊销费',
                                        '利息支出', '其他费用', '总成本费用', '固定成本', '可变成本', '经营成本', '盈亏平衡点（%）'],
                                  '8': ['生产负荷（%）', '营业收入', '营业税金及附加', '增值税（17%）', '城市维护建设税（4%）', '教育费附加（2%）'],
                                  '9': ['生产负荷（%）', '资金来源', '利润总额', '折旧费', '摊销费', '长期借款', '流动资金借款', '自有资金',
                                        '回收固定资产余值', '回收流动资金', '资金应用', '建设投资', '建设期利息', '流动资金', '所得税',
                                        '长期借款本金偿还', '盈余资金', '累计盈余资金'],
                                  '10': ['生产负荷（%）', '营业收入', '营业税金及附加', '总成本费用', '利润总额（1-2-3）',
                                         '所得税', '税后利润（4-5）', '可供分配利润', '盈余公积金', '盈余公益金', '未分配利润(7-7.1-7.2)', '累计未分配利润'],
                                  '11': ['生产负荷（%）', '现金流入', '营业收入', '回收固定资产余值', '回收流动资金', '现金流出', '建设投资',
                                         '流动资金', '经营成本', '营业税金及附加', '所得税', '净现金流量', '累计现金流量',
                                         '所得税前净现金流量', '所得税前累计净现金流量'],
                                  '12': ['生产负荷（%）', '现金流入', '营业收入', '回收固定资产余值', '回收流动资金', '现金流出',
                                         '项目资本金', '借款本金偿还', '借款利息支付', '经营成本', '营业税金及附加', '所得税',
                                         '净现金流量', '累计现金流量','', '净现金流量', '累计现金流量'],
                                  '13': ['生产负荷（%）', '资产', '流动资产总额', '应收帐款', '存货', '现金', '累计盈余资金',
                                         '固定资产净值', '无形及递延资产净值', '在建工程', '负债及所有者权益', '流动负债总额',
                                         '应付帐款', '流动资金借款', '长期借款', '所有者权益', '累计资本金', '累计盈余公积金',
                                         '累计盈余公益金', '累计未分配利润', '资产负债率', '流动比率', '速动比率'],
                                  '14': ['年设计生产能力', '固定成本','单位产品变动成本','单位产品售价','单位产品销售税金及附加','','盈亏平衡产量','盈亏平衡销售价格','盈亏平衡生产能力利用率']
                                   }

        # 第一列的序号是什么
        self.sheetDefaultArrowNum = {'2': [ '1', '1.1', '1.2', '1.2.1', '1.2.2', '1.2.3', '1.2.4', '1.3', '2', '2.1', '3', '4', '5', '6'],
                                  '3': [ '1', '1.1', '1.2', '1.3', '1.4', '1.4.1', '1.4.2', '1.5', '2', '2.1', '2.2', '2.3', '3', '4'],
                                  '4': ['1','1.1','1.2','1.3','2','2.1','2.1.1','2.1.2','2.2','2.2.1','2.2.2','2.2.3'],
                                  '5': ['1', '1.1', '1.2', '1.3', '2', '2.1', '2.2', '2.3', '', '3', '3.1', '3.2', '3.3', '', '4', '4.1', '4.2', '4.3','', '备注', ''],
                                  '6': [ '1', '1.1', '1.2', '2', '2.1', '2.2', '3', '3.1', '3.2'],
                                  '7': ['','1', '2', '3', '4', '5', '6', '7', '8', '9', '9.1', '9.2', '9.3', '10'],
                                  '8': [ '','1', '2', '2.1', '2.2', '2.3'],
                                  '9': [ '','1', '1.1', '1.2', '1.3', '1.4', '1.5', '1.6', '1.7', '1.8', '2', '2.1', '2.2', '2.3', '2.4', '2.6', '3', '4'],
                                  '10': ['', '1', '2', '3', '4', '5', '6', '7', '7.1', '7.2', '7.3', '8'],
                                  '11': ['', '1', '1.1', '1.2', '1.3', '2', '2.1', '2.2', '2.3', '2.4', '2.5', '3', '4', '5', '6'],
                                  '12': ['', '1', '1.1', '1.2', '1.3', '2', '2.1', '2.2', '2.3', '2.4', '2.5', '2.6', '3', '4', '所得税前', '7', '8'],
                                  '13': ['', '1', '1.1', '1.1.1', '1.1.2', '1.1.3', '1.1.4', '1.2', '1.3', '1.4', '2', '2.1', '2.1.1', '2.1.2', '2.2', '2.3', '2.3.1', '2.3.2', '2.3.3', '2.3.4', '计算指标','',''],
                                  '14': ['','1','2','3','4','','5','6','7']}

        # 对应序号的表头是什么
        self.sheetNames = {'2': '附表2：流动资金估算', '3': '附表3：借款还本付息表', '4': '附表4：投资使用计划与资金筹措表',
                            '5': '附表5：固定资产折旧估算表', '6': '附表6：无形资产及其他资产摊销表', '7': '附表7：总成本费用估算表',
                            '8': '附表8：营业收入、营业税金及附加表', '9': '附表9：资金来源与运用表', '10': '附表10：损益表',
                            '11': '附表11-1：现金流量表（全部投资)', '12': '附表11-2：现金流量表（自有资金)', '13': '附表12：资产负债表',
                           '14': '盈亏平衡分析数据'
                           }
        # 对应序号的表名是什么
        self.sheetIndexs = {'1':'附表1建设投资估算表','2': '附表2流动资金估算', '3': '附表3借款还本付息表', '4': '附表4投资使用计划与资金筹措表',
                            '5': '附表5固定资产折旧表', '6': '附表6无形资产及其他资产摊销表', '7': '附表7总成本估算表',
                            '8': '附表8营业收入、营业税金及附加表', '9': '附表9资金来源与运用表', '10': '附表10损益表',
                            '11': '附表11项目全部投资现金流量表', '12': '附表11~2项目资本金现金流量表', '13': '附表12资产负债',
                            '14': '盈亏平衡分析数据','15': '敏感性分析数据', '16': '汇总表'
                           }

        self.yy_col_index = {
            '3': self.sheetAddCols['3'] + self.fund_data['input2'] + 1,
            '4': self.sheetAddCols['4'] + self.fund_data['input2'] + 1,
            '9': self.sheetAddCols['9'] + self.fund_data['input2'] + 1,
            '13': self.sheetAddCols['13'] + self.fund_data['input2'] + 1,
            '11' : self.sheetAddCols['11'] + self.fund_data['input2'] + 2,
            '12' : self.sheetAddCols['12'] + self.fund_data['input2'] + 2,
            '14': self.sheetAddCols['14'] + 1
        }

        # 设定每个表处理数据的方式（公式）,先生成一个固定部分的字典，其它由代码生成，如果由疑难点可加入手动设定(最后生成固定部分，因此优先级最高)
        self.sheetFormula = {'2': {'C5': "='基础数据4-表2、3、5、6、7、8、10'!C3",
   'C7': "='基础数据4-表2、3、5、6、7、8、10'!C4",
   'C8': "='基础数据4-表2、3、5、6、7、8、10'!C5",
   'C9': "='基础数据4-表2、3、5、6、7、8、10'!C6",
   'C10': "='基础数据4-表2、3、5、6、7、8、10'!C7",
   'C11': "='基础数据4-表2、3、5、6、7、8、10'!C8",
   'C13': "='基础数据4-表2、3、5、6、7、8、10'!C9",
   'D5': '=360/C5',
   'D7': '=360/C7',
   'D8': '=360/C8',
   'D9': '=360/C9',
   'D10': '=360/C10',
   'D11': '=360/C11',
   'D13': '=360/C13',
    'Basic' : {'1' : { 'E4': '=E5+E6+E11',
   'E5': '=附表7总成本估算表!D16/$D$5',
   'E6': '=E7+E8+E9+E10',
   'E7': '=(附表7总成本估算表!D5)/$D$7',
   'E8': '=附表7总成本估算表!D6/$D$8',
   'E9': '=(附表7总成本估算表!D5+附表7总成本估算表!D6+附表7总成本估算表!D7+附表7总成本估算表!D8)/$D$9',
   'E10': '=附表7总成本估算表!D16/$D$10',
   'E11': '=(附表7总成本估算表!D7+附表7总成本估算表!D12)/$D$11',
   'E12': '=E13',
   'E13': '=(附表7总成本估算表!D5+附表7总成本估算表!D6)/$D$13',
   'E14': '=E4-E12',
   'E15': '=E14-D14',
   'E16': "=E15*'基础数据4-表2、3、5、6、7、8、10'!$C$11",
   'E17': "=E14*'基础数据4-表2、3、5、6、7、8、10'!$C$10"}
               }
                                    },

                             '3': {
                                 'C4': "='基础数据4-表2、3、5、6、7、8、10'!C20",
                                 'D9': f"=SUM(G9:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}9)",
                                 'D10': f"=SUM(G10:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}10)",
                                 'D12': '=D13+D14+D15',
                                 'D13': f"=SUM(G13:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}13)",
                                 'D14': f"=SUM(G14:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}14)",
                                 'D15': f"=SUM(G15:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}15)",
                                 f'{get_column_letter(self.yy_col_index["3"])}5': f'={get_column_letter(self.yy_col_index["3"] - 1)}11',

                                 f'{get_column_letter(self.yy_col_index["3"])}11': f'={get_column_letter(self.yy_col_index["3"])}5-{get_column_letter(self.yy_col_index["3"])}9',
                                 'Basic': {
                                     '1':{
                                     'E5': '=D5+D6+D7',
                                     'E6': '=附表4投资使用计划与资金筹措表!D13',
                                     'E7': '=附表4投资使用计划与资金筹措表!D6',
                                     'E11': '=D11+E6+E7',
                                     'E12': '=E13+E14+E15',
                                     },
                                     '2': {
                                         'G8': '=G9+G10',
                                         'G9': f"=IF(F11>${get_column_letter(self.yy_col_index['3'])}$5/'基础数据4-表2、3、5、6、7、8、10'!$B$23,${get_column_letter(self.yy_col_index['3'])}$5/'基础数据4-表2、3、5、6、7、8、10'!$B$23,F11)",
                                         'G10': "=F11*'基础数据4-表2、3、5、6、7、8、10'!$C$20",
                                         'G11': '=F11-G9',
                                         'G12': '=G13+G14+G15',
                                         f'{get_column_letter(self.yy_col_index["3"])}13': '=附表10损益表!D8',
                                         f'{get_column_letter(self.yy_col_index["3"])}14': '=附表5固定资产折旧表!F20',
                                         f'{get_column_letter(self.yy_col_index["3"])}15': '=附表6无形资产及其他资产摊销表!E11',
                                         f'{get_column_letter(self.yy_col_index["3"])}16': '=(附表10损益表!D8+附表7总成本估算表!D11)/附表7总成本估算表!D11',
                                         f'{get_column_letter(self.yy_col_index["3"])}17': f'=(附表10损益表!D10+附表7总成本估算表!D9+附表7总成本估算表!D10+附表7总成本估算表!D11)/({get_column_letter(self.yy_col_index["3"])}9+附表7总成本估算表!D11)',
                                     }
                                 },
                             },
                             '4': {
                                 'C4': f'=SUM(D4:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}4)',
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}15)',
                                 'D14': "=D13*'基础数据4-表2、3、5、6、7、8、10'!$C$20",
                                 'Basic': {
                                     '1':{
                                         'D4': '=D5+D6+D7',
                                         'D6': '=D14',
                                         'D8': '=D9+D12',
                                         'D9': '=D5-D13',
                                         'D10': '=D5-D13',
                                         'D12': '=D13+D14+D15',
                                         'D13': "=D5*'基础数据3-投资使用计划与资金筹措'!B8",
                                         'D14': "= (SUM($D$14:C14) + SUM($D$13: D13))*'基础数据4-表2、3、5、6、7、8、10'!$C$20",

                                     },
                                     '2':{
                                         'F4': '=F5+F6+F7',
                                         f'{get_column_letter(self.yy_col_index["4"])}7': '=附表2流动资金估算!E15',  # 要改为第一个运营期列，让他对应公式：第一列流动资金
                                         'F8': '=F9+F12',
                                         'F9': '=F10+F11',
                                         'F11': '=F7-F15',
                                         'F12': '=F13+F14+F15',
                                         f'{get_column_letter(self.yy_col_index["4"])}15': '=附表2流动资金估算!E16',  # 要改为第一个运营期列，让他对应公式：第一列流动资金
                                     }
                                 },
                             },
                             '5': {
                                 'C9': '=附表1建设投资估算表!D3+附表1建设投资估算表!E3',
                                 'C19': '=C5+C9+C14',
                                 'D4': "='基础数据4-表2、3、5、6、7、8、10'!C32",
                                 'D8': "='基础数据4-表2、3、5、6、7、8、10'!C33",
                                 'D13': "='基础数据4-表2、3、5、6、7、8、10'!C35",
                                 'E4': "='基础数据4-表2、3、5、6、7、8、10'!D32",
                                 'E8': "='基础数据4-表2、3、5、6、7、8、10'!D33",
                                 'F7': '=C5-F6',
                                 'F11': '=C9-F10',
                                 'F16': '=C14-F15',
                                 'Basic': {
                                     '1':{
                                         'F6': '=$C$5*(1-$E$4)/$D$4',
                                         'F7': '=E7-F6',
                                         'F10': '=$C$9*(1-$E$8)/$D$8',
                                         'F11': '=E11-F10',
                                         'F15': '=$C$14/$D$13',
                                         'F16': '=E16-F15',
                                         'F20': '=F6+F10+F15',
                                         'F21': '=F7+F16+F11',
                                     }
                                 },
                             },
                             '6': {
                                 'C10': '=C4+C7',
                                 'D4': "='基础数据4-表2、3、5、6、7、8、10'!C46",
                                 'D7': "='基础数据4-表2、3、5、6、7、8、10'!C47",
                                 'E6': '=C4-E5',
                                 'E9': '=C7-E8',
                                 'Basic': {
                                     '1':{
                                         'E5': '=IF(E3-$E$3+1>$D$4,0,$C$4/$D$4)',
                                         'E6': '=D6-E5',
                                         'E8': '==IF(E3-$E$3+1>$D$7,0,$C$7/$D$7)',
                                         'E9': '=D9-E8',
                                         'E11': '=E5+E8',
                                         'E12': '=E6+E9',
                                     }
                                 },
                             },
                             '7': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}15)',
                                 'C16': '=C5+C6+C7+C8+C12',
                                 'Basic': {
                                     '1':{
                                   }
                                 },
                             },
                             '8': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}9)',
                                 'Basic': {
                                     '1':{

                                     }

                                 },
                             },
                             '9': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}15)',
                                 'C16': f'=SUM(D16:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}16)',
                                 'C17': f'=SUM(D17:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}17)',
                                 'C18': f'=SUM(D18:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}18)',
                                 'C19': f'=SUM(D19:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}19)',
                                 'C20': f'=SUM(D20:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}20)',
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['9'])}12": f"=附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}21",
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['9'])}13": f"=附表2流动资金估算!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['2'])}14",

                                 'Basic': {
                                     '1': {
                                         'D5': '=SUM(D6:D13)',
                                         'D9': '=附表4投资使用计划与资金筹措表!D12',
                                         'D11': '=附表4投资使用计划与资金筹措表!D9',
                                         'D14': '=SUM(D15:D19)',
                                         'D15': '=附表4投资使用计划与资金筹措表!D5',
                                         'D16': '=附表4投资使用计划与资金筹措表!D6',
                                         'D20': '=D5-D14',
                                         'D21': '=D20',
                                     },
                                     '2': {
                                         f'{get_column_letter(self.yy_col_index["9"])}4': "='基础数据1-项目基础信息'!C9",
                                         'F5': '=SUM(F6:F13)',
                                         f'{get_column_letter(self.yy_col_index["9"])}6': '=附表10损益表!D8',
                                         f'{get_column_letter(self.yy_col_index["9"])}7': '=附表5固定资产折旧表!F20',
                                         f'{get_column_letter(self.yy_col_index["9"])}8': '=附表6无形资产及其他资产摊销表!E11',
                                         f'{get_column_letter(self.yy_col_index["9"])}10': '=附表2流动资金估算!E16',
                                         f'{get_column_letter(self.yy_col_index["9"])}11': '=附表4投资使用计划与资金筹措表!F9',
                                         'F14': '=SUM(F15:F19)',
                                         f'{get_column_letter(self.yy_col_index["9"])}17': '=附表2流动资金估算!E15',
                                         f'{get_column_letter(self.yy_col_index["9"])}18': '=附表10损益表!D9',
                                         'F19': '=附表3借款还本付息表!G9',
                                         'F20': '=F5-F14',
                                         'F21': '=E21+F20',
                                     }
                                 },
                             },
                             '10': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}13)',
                                 'C14': '=C11-C12-C13',
                                 'Basic': {
                                     '1':{
                                         'D4': "='基础数据1-项目基础信息'!C9",
                                         'D5': "='附表8营业收入、营业税金及附加表'!D5",
                                         'D6': "='附表8营业收入、营业税金及附加表'!D6",
                                         'D7': '=附表7总成本估算表!D13',
                                         'D8': '=D5-D6-D7',
                                         'D9': "=D8*'基础数据4-表2、3、5、6、7、8、10'!$C$74",
                                         'D10': '=D8-D9',
                                         'D11': '=D10',
                                         'D12': "=D11*'基础数据4-表2、3、5、6、7、8、10'!$C$75",
                                         'D13': "=D11*'基础数据4-表2、3、5、6、7、8、10'!$C$76",
                                         'D14': '=D11-D12-D13',
                                         'D15': '=SUM($D$14:D14)',
                                     }
                                 },
                             },
                             '11': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}15)',
                                 'C17': f'=SUM(D17:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"])}17)',
                                 f"{get_column_letter(self.fund_data['input2'] + self.sheetAddCols['11'] + 1)}10":'',
                                 f"{get_column_letter(self.fund_data['input2'] + self.sheetAddCols['11'] + 1)}11": f"=附表2流动资金估算!E15",
                                 'D22':"指标计算",
                                 'D23': "静态回收期（年）",
                                 'D24': "动态回收期（年）",
                                 'D25': "内部收益率",
                                 'D26': "净现值（万元）",
                                 'E22': "税前",
                                 'F22': "税后",
                                 'G23': "（从建设期算起）",
                                 'G24': "（从建设期算起）",
                                 "E25": f"=IRR(D17:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}17)",
                                 "E26": f"=D17+NPV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},E17:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}17)",
                                 "F25": f"=IRR(D15:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}15)",
                                 "F26": f"=D15+NPV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},E15:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}15)",
                                 # E23,F23,E24,F24涉及数组计算公式放到fillsheet里
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}7": f"=附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}21",
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}8": f"=附表2流动资金估算!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['2'])}14",

                                 'Basic': {
                                     '1':{
                                        'D9': "= D10 + D11 + D12 + D13 + D14",
                                        'D10': "=附表4投资使用计划与资金筹措表!D4",
                                        'D15': "=D5-D9",
                                        'D16': "=C16+D15",
                                        'D17': "=D15+D14",
                                        'D18': "=C18+D17",
                                     },
                                     '2':
                                         {
                                             f"{get_column_letter(self.yy_col_index['11'])}4": "='基础数据1-项目基础信息'!C9",
                                             'G5': '=G6+G7+G8',
                                             f'{get_column_letter(self.yy_col_index["11"])}6': '=附表8营业收入、营业税金及附加表!D5',
                                             'G9': '=G10+G11+G12+G13+G14',
                                             f'{get_column_letter(self.yy_col_index["11"])}11': '=附表2流动资金估算!F15',
                                             f'{get_column_letter(self.yy_col_index["11"])}12': '=附表7总成本估算表!D16',
                                             f'{get_column_letter(self.yy_col_index["11"])}13': '=附表8营业收入、营业税金及附加表!D6',
                                             f'{get_column_letter(self.yy_col_index["11"])}14': '=附表10损益表!D9',
                                             'G15': "=G5-G9",
                                            'G16': "=F16+G15",
                                            'G17': "=G15+G14",
                                            'G18': "=F18+G17",
                                         }
                                 },
                             },
                             '12': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}10)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}15)',
                                 'C16': f'=SUM(D16:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}16)',
                                 'C19': f'=SUM(E19:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"])}19)',
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}7": f"=附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}21",
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}8": f"=附表2流动资金估算!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['2'])}14",
                                 'D22': "指标计算",
                                 'D23': "静态回收期（年）",
                                 'D24': "动态回收期（年）",
                                 'D25': "内部收益率",
                                 'D26': "净现值（万元）",
                                 'E22': "税前",
                                 'F22': "税后",
                                 'G23': "（从建设期算起）",
                                 'G24': "（从建设期算起）",
                                 "E25": f"=IRR(D19:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}19)",
                                 "E26": f"=D19+NPV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},E19:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}19)",
                                 "F25": f"=IRR(D16:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}16)",
                                 "F26": f"=D16+NPV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},E16:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}16)",
                                 'Basic': {
                                     '1':{
                                         'D9': '=SUM(D10:D15)',
                                         'D10': '=附表4投资使用计划与资金筹措表!D9',
                                         'D16': '=D5-D9',
                                         'D17': '=C17+D16',
                                         'D19': '=D16+D15',
                                         'D20': '=C20+D19',
                                     },
                                     '2':
                                         {
                                         f"{get_column_letter(self.yy_col_index['12'])}4": "='基础数据1-项目基础信息'!C9",
                                         'G5': '=SUM(G6:G8)',
                                         f"{get_column_letter(self.yy_col_index['12'])}6": "='附表8营业收入、营业税金及附加表'!D5",
                                         f'G9': '=SUM(G10:G15)',
                                         f'G10': '=附表4投资使用计划与资金筹措表!F9',
                                         f'G11': '=附表3借款还本付息表!G9',
                                         f'G12': '=附表3借款还本付息表!G10',
                                         f"{get_column_letter(self.yy_col_index['12'])}13": '=附表7总成本估算表!D16',
                                         f"{get_column_letter(self.yy_col_index['12'])}14": "='附表8营业收入、营业税金及附加表'!D6",
                                         f"{get_column_letter(self.yy_col_index['12'])}15": '=附表10损益表!D9',
                                         'G16': '=G5-G9',
                                         'G17': '=F17+G16',
                                         'G19': '=G16+G15',
                                         'G20': '=F20+G19',
                                         }
                                 },
                             },
                             '13': {
                                 'Basic': {
                                     '1':{
                                         'C5': '=C6+C11+C12+C13',
                                         'C6': '=C7+C8+C9+C10',
                                         'C13': '=SUM(附表4投资使用计划与资金筹措表!$D$4:附表4投资使用计划与资金筹措表!D4)',
                                         'C14': '=C15+C18+C19',
                                         'C15': '=C16+C17',
                                         'C18': '=附表3借款还本付息表!E11',
                                         'C19': '=SUM(C20:C23)',
                                         'C20': '=SUM(附表4投资使用计划与资金筹措表!$D$9:附表4投资使用计划与资金筹措表!D9)',
                                     },
                                     '2':{
                                         f'{get_column_letter(self.yy_col_index["13"])}4': "='基础数据1-项目基础信息'!C9",
                                         'E5': '=E6+E11+E12+E13-附表9资金来源与运用表!F12-附表9资金来源与运用表!F13',
                                         'E6': '=E7+E8+E9+E10',
                                         f'{get_column_letter(self.yy_col_index["13"])}7': '=附表2流动资金估算!E5',
                                         f'{get_column_letter(self.yy_col_index["13"])}8': '=附表2流动资金估算!E6',
                                         f'{get_column_letter(self.yy_col_index["13"])}9': '=附表2流动资金估算!E11',
                                         'E10': '=附表9资金来源与运用表!F21',
                                         f'{get_column_letter(self.yy_col_index["13"])}11': '=附表5固定资产折旧表!F21',
                                         f'{get_column_letter(self.yy_col_index["13"])}12': '=附表6无形资产及其他资产摊销表!E12',
                                         'E14': '=E15+E18+E19',
                                         'E15': '=E16+E17',
                                         f'{get_column_letter(self.yy_col_index["13"])}16': '=附表2流动资金估算!E13',
                                         f'{get_column_letter(self.yy_col_index["13"])}17': "=附表2流动资金估算!E14*'基础数据4-表2、3、5、6、7、8、10'!$C$11",
                                         'E18': '=附表3借款还本付息表!G11',
                                         'E19': '=SUM(E20:E23)',
                                         'E20': '=SUM(附表4投资使用计划与资金筹措表!$D$9:附表4投资使用计划与资金筹措表!F9)',
                                         f'{get_column_letter(self.yy_col_index["13"])}21': '=SUM(附表10损益表!D12:附表10损益表!D12)',
                                         f'{get_column_letter(self.yy_col_index["13"])}22': '=SUM(附表10损益表!D13:附表10损益表!D13)',
                                         f'{get_column_letter(self.yy_col_index["13"])}23': '=附表10损益表!D15',
                                         'E24': '=(E15+E18)/E5*100%',
                                         'E25': '=E6/E15',
                                         'E26': '=(E6-E8)/E15',
                                     }
                                 },
                             }

                             }
        self.calsheetFormula = {'2': {'C5': "=保留!C22",
                                   'C7': "=保留!C23",
                                   'C8': "=保留!C24",
                                   'C9': "=保留!C25",
                                   'C10': "=保留!C26",
                                   'C11': "=保留!C27",
                                   'C13': "=保留!C28",
                                   'D5': '=360/C5',
                                   'D7': '=360/C7',
                                   'D8': '=360/C8',
                                   'D9': '=360/C9',
                                   'D10': '=360/C10',
                                   'D11': '=360/C11',
                                   'D13': '=360/C13',
                                   'Basic': {'1': {'E4': '=E5+E6+E11',
                                                   'E5': '=附表7总成本估算表!D16/$D$5',
                                                   'E6': '=E7+E8+E9+E10',
                                                   'E7': '=(附表7总成本估算表!D5)/$D$7',
                                                   'E8': '=附表7总成本估算表!D6/$D$8',
                                                   'E9': '=(附表7总成本估算表!D5+附表7总成本估算表!D6+附表7总成本估算表!D7+附表7总成本估算表!D8)/$D$9',
                                                   'E10': '=附表7总成本估算表!D16/$D$10',
                                                   'E11': '=(附表7总成本估算表!D7+附表7总成本估算表!D12)/$D$11',
                                                   'E12': '=E13',
                                                   'E13': '=(附表7总成本估算表!D5+附表7总成本估算表!D6)/$D$13',
                                                   'E14': '=E4-E12',
                                                   'E15': '=E14-D14',
                                                   'E16': "=E15*保留!$C$30",
                                                   'E17': "=E14*保留!$C$30*保留!$C$29"}
                                             }
                                   },

                             '3': {
                                 'C4': "=保留!C33",
                                 'D9': f"=SUM(G9:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}9)",
                                 'D10': f"=SUM(G10:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}10)",
                                 'D12': '=D13+D14+D15',
                                 'D13': f"=SUM(G13:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}13)",
                                 'D14': f"=SUM(G14:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}14)",
                                 'D15': f"=SUM(G15:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['3'])}15)",
                                 f'{get_column_letter(self.yy_col_index["3"])}5': f'={get_column_letter(self.yy_col_index["3"] - 1)}11',

                                 f'{get_column_letter(self.yy_col_index["3"])}11': f'={get_column_letter(self.yy_col_index["3"])}5-{get_column_letter(self.yy_col_index["3"])}9',
                                 'Basic': {
                                     '1': {
                                         'E5': '=D5+D6+D7',
                                         'E6': '=附表4投资使用计划与资金筹措表!D13',
                                         'E7': '=附表4投资使用计划与资金筹措表!D6',
                                         'E11': '=D11+E6+E7',
                                         'E12': '=E13+E14+E15',
                                     },
                                     '2': {
                                         'G8': '=G9+G10',
                                         'G9': f"=IF(F11>${get_column_letter(self.yy_col_index['3'])}$5/保留!$C$35,${get_column_letter(self.yy_col_index['3'])}$5/保留!$C$35,F11)",
                                         'G10': "=F11*保留!$C$33",
                                         'G11': '=F11-G9',
                                         'G12': '=G13+G14+G15',
                                         f'{get_column_letter(self.yy_col_index["3"])}13': '=附表10损益表!D8',
                                         f'{get_column_letter(self.yy_col_index["3"])}14': '=附表5固定资产折旧表!F20',
                                         f'{get_column_letter(self.yy_col_index["3"])}15': '=附表6无形资产及其他资产摊销表!E11',
                                         f'{get_column_letter(self.yy_col_index["3"])}16': '=(附表10损益表!D8+附表7总成本估算表!D11)/附表7总成本估算表!D11',
                                         f'{get_column_letter(self.yy_col_index["3"])}17': f'=(附表10损益表!D10+附表7总成本估算表!D9+附表7总成本估算表!D10+附表7总成本估算表!D11)/({get_column_letter(self.yy_col_index["3"])}9+附表7总成本估算表!D11)',
                                     }
                                 },
                             },
                             '4': {
                                 'C4': f'=SUM(D4:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}4)',
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["4"])}15)',
                                 'D14': "=D13*保留!$C$33",
                                 'Basic': {
                                     '1': {
                                         'D4': '=D5+D6+D7',
                                         'D6': '=D14',
                                         'D8': '=D9+D12',
                                         'D9': '=D5-D13',
                                         'D10': '=D5-D13',
                                         'D12': '=D13+D14+D15',
                                         'D13': "=D5*保留!B19",
                                         'D14': "= (SUM($D$14:C14) + SUM($D$13: D13))*保留!$C$33",
                                     },
                                     '2': {
                                         'F4': '=F5+F6+F7',
                                         f'{get_column_letter(self.yy_col_index["4"])}7': '=附表2流动资金估算!E15',
                                         # 要改为第一个运营期列，让他对应公式：第一列流动资金
                                         'F8': '=F9+F12',
                                         'F9': '=F10+F11',
                                         'F11': '=F7-F15',
                                         'F12': '=F13+F14+F15',
                                         f'{get_column_letter(self.yy_col_index["4"])}15': '=附表2流动资金估算!E16',
                                         # 要改为第一个运营期列，让他对应公式：第一列流动资金
                                     }
                                 },
                             },
                             '5': {
                                 'C9': '=附表1建设投资估算表!D3+附表1建设投资估算表!E3',
                                 'C19': '=C5+C9+C14',
                                 'D4': "=保留!C37",
                                 'D8': "=保留!C38",
                                 'D13': "=保留!C39",
                                 'E4': "=保留!C40",
                                 'E8': "=保留!C41",
                                 'F7': '=C5-F6',
                                 'F11': '=C9-F10',
                                 'F16': '=C14-F15',
                                 'F29': '=C27-F28',
                                 'Basic': {
                                     '1': {
                                         'F28': '=$C$27*(1-$E$4)/$D$4',
                                         'F29': '=E29-F28',
                                         'F6': '=$C$5*(1-$E$4)/$D$4',
                                         'F7': '=E7-F6',
                                         'F10': '=$C$9*(1-$E$8)/$D$8',
                                         'F11': '=E11-F10',
                                         'F15': '=$C$14/$D$13',
                                         'F16': '=E16-F15',
                                         'F20': '=F6+F10+F15',
                                         'F21': '=F7+F16+F11',
                                     }
                                 },
                             },
                             '6': {
                                 'C10': '=C4+C7',
                                 'D4': "=保留!C42",
                                 'D7': "=保留!C43",
                                 'E6': '=C4-E5',
                                 'E9': '=C7-E8',
                                 'Basic': {
                                     '1': {
                                         'E5': '=IF(E3-$E$3+1>$D$4,0,$C$4/$D$4)',
                                         'E6': '=D6-E5',
                                         'E8': '==IF(E3-$E$3+1>$D$7,0,$C$7/$D$7)',
                                         'E9': '=D9-E8',
                                         'E11': '=E5+E8',
                                         'E12': '=E6+E9',
                                     }
                                 },
                             },
                             '7': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["7"])}15)',
                                 'C16': '=C5+C6+C7+C8+C12',
                                 'Basic': {
                                     '1': {
                                     }
                                 },
                             },
                             '8': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["8"])}9)',
                                 'Basic': {
                                     '1': {

                                     }

                                 },
                             },
                             '9': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}15)',
                                 'C16': f'=SUM(D16:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}16)',
                                 'C17': f'=SUM(D17:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}17)',
                                 'C18': f'=SUM(D18:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}18)',
                                 'C19': f'=SUM(D19:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}19)',
                                 'C20': f'=SUM(D20:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["9"])}20)',
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['9'])}12": f"=附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}21",
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['9'])}13": f"=附表2流动资金估算!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['2'])}14",

                                 'Basic': {
                                     '1': {
                                         'D5': '=SUM(D6:D13)',
                                         'D9': '=附表4投资使用计划与资金筹措表!D12',
                                         'D11': '=附表4投资使用计划与资金筹措表!D9',
                                         'D14': '=SUM(D15:D19)',
                                         'D15': '=附表4投资使用计划与资金筹措表!D5',
                                         'D16': '=附表4投资使用计划与资金筹措表!D6',
                                         'D20': '=D5-D14',
                                         'D21': '=D20',
                                     },
                                     '2': {
                                         f'{get_column_letter(self.yy_col_index["9"])}4': "=保留!C6",
                                         'F5': '=SUM(F6:F13)',
                                         f'{get_column_letter(self.yy_col_index["9"])}6': '=附表10损益表!D8',
                                         f'{get_column_letter(self.yy_col_index["9"])}7': '=附表5固定资产折旧表!F20',
                                         f'{get_column_letter(self.yy_col_index["9"])}8': '=附表6无形资产及其他资产摊销表!E11',
                                         f'{get_column_letter(self.yy_col_index["9"])}10': '=附表2流动资金估算!E16',
                                         f'{get_column_letter(self.yy_col_index["9"])}11': '=附表4投资使用计划与资金筹措表!F9',
                                         'F14': '=SUM(F15:F19)',
                                         f'{get_column_letter(self.yy_col_index["9"])}17': '=附表2流动资金估算!E15',
                                         f'{get_column_letter(self.yy_col_index["9"])}18': '=附表10损益表!D9',
                                         'F19': '=附表3借款还本付息表!G9',
                                         'F20': '=F5-F14',
                                         'F21': '=E21+F20',
                                     }
                                 },
                             },
                             '10': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["10"])}13)',
                                 'C14': '=C11-C12-C13',
                                 'Basic': {
                                     '1': {
                                         'D4': "=保留!C6",
                                         'D5': "='附表8营业收入、营业税金及附加表'!D5",
                                         'D6': "='附表8营业收入、营业税金及附加表'!D6",
                                         'D7': '=附表7总成本估算表!D13',
                                         'D8': '=D5-D6-D7',
                                         'D9': "=D8*保留!$C$57",
                                         'D10': '=D8-D9',
                                         'D11': '=D10',
                                         'D12': "=D11*保留!$C$58",
                                         'D13': "=D11*保留!$C$59",
                                         'D14': '=D11-D12-D13',
                                         'D15': '=SUM($D$14:D14)',
                                     }
                                 },
                             },
                             '11': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}10)',
                                 'C11': f'=SUM(D11:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}11)',
                                 'C12': f'=SUM(D12:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}12)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}15)',
                                 'C17': f'=SUM(D17:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["11"]+1)}17)',
                                 f"{get_column_letter(self.fund_data['input2'] + self.sheetAddCols['11'] + 1)}10": '',
                                 f"{get_column_letter(self.fund_data['input2'] + self.sheetAddCols['11'] + 1)}11": f"=附表2流动资金估算!E15",
                                 'D22': "指标计算",
                                 'D23': "静态回收期（年）",
                                 'D24': "动态回收期（年）",
                                 'D25': "内部收益率",
                                 'D26': "净现值（万元）",
                                 'E22': "税前",
                                 'F22': "税后",
                                 'G23': "（从建设期算起）",
                                 'G24': "（从建设期算起）",
                                 "E25": f"=IRR(D17:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}17)",
                                 "E26": f"=D17+NPV(保留!$C$7,E17:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}17)",
                                 "F25": f"=IRR(D15:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}15)",
                                 "F26": f"=D15+NPV(保留!$C$7,E15:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}15)",
                                 # E23,F23,E24,F24涉及数组计算公式放到fillsheet里
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}7": f"=附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}21-附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}7+附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}29",
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}8": f"=附表2流动资金估算!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['2'])}14",

                                 'Basic': {
                                     '1': {
                                         'D9': "= D10 + D11 + D12 + D13 + D14",
                                         'D10': "=附表4投资使用计划与资金筹措表!D4",
                                         'D15': "=D5-D9",
                                         'D16': "=C16+D15",
                                         'D17': "=D15+D14",
                                         'D18': "=C18+D17",
                                     },
                                     '2':
                                         {
                                             f"{get_column_letter(self.yy_col_index['11'])}4": "=保留!C6",
                                             'G5': '=G6+G7+G8',
                                             f'{get_column_letter(self.yy_col_index["11"])}6': '=附表8营业收入、营业税金及附加表!D5',
                                             'G9': '=G10+G11+G12+G13+G14',
                                             f'{get_column_letter(self.yy_col_index["11"])}11': '=附表2流动资金估算!F15',
                                             f'{get_column_letter(self.yy_col_index["11"])}12': '=附表7总成本估算表!D16',
                                             f'{get_column_letter(self.yy_col_index["11"])}13': '=附表8营业收入、营业税金及附加表!D6',
                                             f'{get_column_letter(self.yy_col_index["11"])}14': '=附表10损益表!D9',
                                             'G15': "=G5-G9",
                                             'G16': "=F16+G15",
                                             'G17': "=G15+G14",
                                             'G18': "=F18+G17",
                                         }
                                 },
                             },
                             '12': {
                                 'C5': f'=SUM(D5:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}5)',
                                 'C6': f'=SUM(D6:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}6)',
                                 'C7': f'=SUM(D7:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}7)',
                                 'C8': f'=SUM(D8:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}8)',
                                 'C9': f'=SUM(D9:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}9)',
                                 'C10': f'=SUM(D10:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}10)',
                                 'C13': f'=SUM(D13:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}13)',
                                 'C14': f'=SUM(D14:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}14)',
                                 'C15': f'=SUM(D15:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}15)',
                                 'C16': f'=SUM(D16:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}16)',
                                 'C19': f'=SUM(E19:{get_column_letter(self.fund_data["input1"] + self.sheetAddCols["12"]+1)}19)',
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}7": f"=附表5固定资产折旧表!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['5'])}21",
                                 f"{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}8": f"=附表2流动资金估算!{get_column_letter(self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols['2'])}14",
                                 'D22': "指标计算",
                                 'D23': "静态回收期（年）",
                                 'D24': "动态回收期（年）",
                                 'D25': "内部收益率",
                                 'D26': "净现值（万元）",
                                 'E22': "税前",
                                 'F22': "税后",
                                 'G23': "（从建设期算起）",
                                 'G24': "（从建设期算起）",
                                 "E25": f"=IRR(D19:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}19)",
                                 "E26": f"=D19+NPV(保留!$C$7,E19:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}19)",
                                 "F25": f"=IRR(D16:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['12'] + 1)}16)",
                                 "F26": f"=D16+NPV(保留!$C$7,E16:{get_column_letter(self.fund_data['input1'] + self.sheetAddCols['11'] + 1)}16)",
                                 'Basic': {
                                     '1': {
                                         'D9': '=SUM(D10:D15)',
                                         'D10': '=附表4投资使用计划与资金筹措表!D9',
                                         'D16': '=D5-D9',
                                         'D17': '=C17+D16',
                                         'D19': '=D16+D15',
                                         'D20': '=C20+D19',
                                     },
                                     '2':
                                         {
                                             f"{get_column_letter(self.yy_col_index['12'])}4": "=保留!C6",
                                             'G5': '=SUM(G6:G8)',
                                             f"{get_column_letter(self.yy_col_index['12'])}6": "='附表8营业收入、营业税金及附加表'!D5",
                                             f'G9': '=SUM(G10:G15)',
                                             f'G10': '=附表4投资使用计划与资金筹措表!F9',
                                             f'G11': '=附表3借款还本付息表!G9',
                                             f'G12': '=附表3借款还本付息表!G10',
                                             f"{get_column_letter(self.yy_col_index['12'])}13": '=附表7总成本估算表!D16',
                                             f"{get_column_letter(self.yy_col_index['12'])}14": "='附表8营业收入、营业税金及附加表'!D6",
                                             f"{get_column_letter(self.yy_col_index['12'])}15": '=附表10损益表!D9',
                                             'G16': '=G5-G9',
                                             'G17': '=F17+G16',
                                             'G19': '=G16+G15',
                                             'G20': '=F20+G19',
                                         }
                                 },
                             },
                             '13': {
                                 'Basic': {
                                     '1': {
                                         'C5': '=C6+C11+C12+C13',
                                         'C6': '=C7+C8+C9+C10',
                                         'C13': '=SUM(附表4投资使用计划与资金筹措表!$D$4:附表4投资使用计划与资金筹措表!D4)',
                                         'C14': '=C15+C18+C19',
                                         'C15': '=C16+C17',
                                         'C18': '=附表3借款还本付息表!E11',
                                         'C19': '=SUM(C20:C23)',
                                         'C20': '=SUM(附表4投资使用计划与资金筹措表!$D$9:附表4投资使用计划与资金筹措表!D9)',
                                     },
                                     '2': {
                                         f'{get_column_letter(self.yy_col_index["13"])}4': "=保留!C6",
                                         'E5': '=E6+E11+E12+E13-附表9资金来源与运用表!F12-附表9资金来源与运用表!F13',
                                         'E6': '=E7+E8+E9+E10',
                                         f'{get_column_letter(self.yy_col_index["13"])}7': '=附表2流动资金估算!E5',
                                         f'{get_column_letter(self.yy_col_index["13"])}8': '=附表2流动资金估算!E6',
                                         f'{get_column_letter(self.yy_col_index["13"])}9': '=附表2流动资金估算!E11',
                                         'E10': '=附表9资金来源与运用表!F21',
                                         f'{get_column_letter(self.yy_col_index["13"])}11': '=附表5固定资产折旧表!F21',
                                         f'{get_column_letter(self.yy_col_index["13"])}12': '=附表6无形资产及其他资产摊销表!E12',
                                         'E14': '=E15+E18+E19',
                                         'E15': '=E16+E17',
                                         f'{get_column_letter(self.yy_col_index["13"])}16': '=附表2流动资金估算!E13',
                                         f'{get_column_letter(self.yy_col_index["13"])}17': "=附表2流动资金估算!E14*保留!$C$30",
                                         'E18': '=附表3借款还本付息表!G11',
                                         'E19': '=SUM(E20:E23)',
                                         'E20': '=SUM(附表4投资使用计划与资金筹措表!$D$9:附表4投资使用计划与资金筹措表!F9)',
                                         f'{get_column_letter(self.yy_col_index["13"])}21': '=SUM(附表10损益表!D12:附表10损益表!D12)',
                                         f'{get_column_letter(self.yy_col_index["13"])}22': '=SUM(附表10损益表!D13:附表10损益表!D13)',
                                         f'{get_column_letter(self.yy_col_index["13"])}23': '=附表10损益表!D15',
                                         'E24': '=(E15+E18)/E5*100%',
                                         'E25': '=E6/E15',
                                         'E26': '=(E6-E8)/E15',
                                     }
                                 },
                             },
                            '14':{
                                'Basic': {
                                    '1': {
                                        'C4': "=保留!C6*保留!$C$52",
                                        'C5': '=附表7总成本估算表!D14',
                                        'C6': '=附表7总成本估算表!D15/C4',
                                        'C7': '=保留!$C$53',
                                        'C8': '=附表8营业收入、营业税金及附加表!D6/C4',
                                        'C10': '=C5/(C7-C6-C8)',
                                        'C11': '=C5/C4+C6+C8',
                                        'C12': '=C10/C4'
                                    }
                                }
                            }

                             }

        self.datasheetNames = ['基础数据1-项目基础信息','基础数据2-建设投资估算 ','基础数据3-投资使用计划与资金筹措','基础数据4-表2、3、5、6、7、8、10']

    def makesheet(self,sheetindex,workbook,start_row=1):
        # 建表思路：先根据时间+k（k与表相关，可以列一个数组堆进去）确定列数，然后分行渲染，第一行是表头，第二行是列名（含横向合并），第三行是年份（含纵向合并），从第四行开始是
        # 正文了，通过一个字典存储这些表结构，如果涉及到添加数据，那么就更新这个字典，否则就按照原字典分行渲染，最后将输入数据导入其中
        wb = workbook
        if start_row==1:
            ws = wb.create_sheet(self.sheetIndexs[sheetindex])
        else:
            ws = wb.active
        # 表头
        ws[f'A{start_row}'] = self.sheetNames[sheetindex]
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndex = self.fund_data['input1'] + self.sheetAddCols[sheetindex]
            if sheetindex in ['11','12']:
                colIndex += 1
            colIndexLetter = number_to_letter(colIndex) #找到单位对应的那一列列名
            ws[f'{colIndexLetter}{start_row}'] = "单位：万元"
        else:
            colIndex = self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols[sheetindex]
            colIndexLetter = number_to_letter(colIndex) #找到单位对应的那一列列名
            ws[f'{colIndexLetter}{start_row}'] = "单位：万元"

        # 表第一行
        iterindex = 0
        for row in ws.iter_rows(min_row=start_row+1, max_col=self.sheetAddCols[sheetindex], max_row=start_row+1):
            for cell in row:
                cell.value = self.sheetAddColsNames[sheetindex][iterindex]
                iterindex += 1

        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndex = self.sheetAddCols[sheetindex] + 1
            colIndexLetter = number_to_letter(colIndex)
            ws[f'{colIndexLetter}{start_row+1}'] = '建设期'
            conIndexLetter = colIndexLetter
        else:
            colIndex = self.sheetAddCols[sheetindex] + 1
            colIndexLetter = number_to_letter(colIndex)

        # 表第二行
        lifeTime = int(self.fund_data['input1'])
        if sheetindex in ['2','5','6','7','8','10','14']:
            start = self.fund_data['input2'] + 1
        elif sheetindex in ['11','12']:
            start = 0
        else:
            start = 1
        for i in range(start, lifeTime +1):
            ws[f'{colIndexLetter}{start_row+2}'] = i
            colIndex = colIndex + 1
            colIndexLetter = number_to_letter(colIndex)

            # 计算投产期对应的列名
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndex = self.fund_data['input2'] + self.sheetAddCols[sheetindex] + 1
            if sheetindex in ['11','12']:
                colIndex += 1
            colIndexLetter = number_to_letter(colIndex)  # 找到投产期对应的那一列列名
            ws[f'{colIndexLetter}{start_row+1}'] = '投产期'
            touIndexLetter = colIndexLetter
        else:
            colIndex = self.sheetAddCols[sheetindex] + 1
            colIndexLetter = number_to_letter(colIndex)  # 找到投产期对应的那一列列名
            ws[f'{colIndexLetter}{start_row+1}'] = '投产期'
            touIndexLetter = colIndexLetter

            # 计算达产期对应的列名
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndex = self.fund_data['input3'] + self.sheetAddCols[sheetindex]
            if sheetindex in ['11','12']:
                colIndex += 1
            colIndexLetter = number_to_letter(colIndex)  # 找到达产期对应的那一列列名
            ws[f'{colIndexLetter}{start_row+1}'] = '达产期'
            daIndexLetter = colIndexLetter
        else:
            colIndex = self.fund_data['input3'] -self.fund_data['input2'] + self.sheetAddCols[sheetindex]
            colIndexLetter = number_to_letter(colIndex)  # 找到达产期对应的那一列列名
            ws[f'{colIndexLetter}{start_row+1}'] = '达产期'
            daIndexLetter = colIndexLetter



        # 表第一列和第二列
        rowIndex = start_row+3
        for i in range(len(self.sheetDefaultArrows[sheetindex])):
            ws[f'A{rowIndex}'] = self.sheetDefaultArrowNum[sheetindex][i]
            ws[f'B{rowIndex}'] = self.sheetDefaultArrows[sheetindex][i]
            rowIndex = rowIndex + 1

        # 遍历工作表中的所有行并设定所有对齐
        for row in ws:
            # 遍历每一行中的所有单元格
            for cell in row:
                # 清除单元格的值
                cell.alignment = Alignment(horizontal='center', vertical='center')

        #  合并表头
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndex = self.fund_data['input1'] + self.sheetAddCols[sheetindex] - 1
            if sheetindex in ['11','12']:
                colIndex += 1
            colIndexLetter = number_to_letter(colIndex)
        else:
            colIndex = self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols[sheetindex] - 1
            colIndexLetter = number_to_letter(colIndex)
        mergeexcel(f'A{start_row}', f'{colIndexLetter}{start_row}', ws)
        #  合并第一行和第二行
        for colIndex in range(1,self.sheetAddCols[sheetindex]+1):
            colIndexLetter = number_to_letter(colIndex)
            mergeexcel(f'{colIndexLetter}{start_row+1}', f'{colIndexLetter}{start_row+2}', ws)
        #  合并第一行
        #建设期
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndexLetter2 = get_column_letter(column_index_from_string(touIndexLetter) - 1)
            mergeexcel(f'{conIndexLetter}{start_row+1}', f'{colIndexLetter2}{start_row+1}', ws)
        #投产期
        colIndexLetter2 = get_column_letter(column_index_from_string(daIndexLetter) - 1)
        mergeexcel(f'{touIndexLetter}{start_row+1}', f'{colIndexLetter2}{start_row+1}', ws)
        # 达产期
        colIndex2 = self.fund_data['input1'] + self.sheetAddCols[sheetindex] if sheetindex not in ['2', '5', '6', '7', '8', '10','14'] else self.fund_data['input1'] + self.sheetAddCols[sheetindex] - self.fund_data['input2']
        if sheetindex in ['11', '12']:
            colIndex2 += 1
        colIndexLetter2 = number_to_letter(colIndex2)
        mergeexcel(f'{daIndexLetter}{start_row+1}', f'{colIndexLetter2}{start_row+1}', ws)
        return start_row + 3 + len(self.sheetDefaultArrows[sheetindex])

    def fillsheet(self,sheetindex,workbook,start_row=1):
        # 建表思路：先根据时间+k（k与表相关，可以列一个数组堆进去）确定列数，然后分行渲染，第一行是表头，第二行是列名（含横向合并），第三行是年份（含纵向合并），从第四行开始是
        # 正文了，通过一个字典存储这些表结构，如果涉及到添加数据，那么就更新这个字典，否则就按照原字典分行渲染，最后将输入数据导入其中
        wb = workbook
        if start_row==1:
            ws = wb[self.sheetIndexs[sheetindex]]
        else:
            ws = wb.active
        # print(self.invenstment)
        # 计算达产期在表7的列名以帮助更新公式
        da_letter = get_column_letter(
            self.fund_data['input3'] + self.sheetAddCols['7'] - self.fund_data['input2'])  # (达产期-建设期+多余的添加列),表7和8公用

        sheetFormula_rev = {
            '4':{
                'Basic':{
                    'D5': f"=附表1建设投资估算表!{self.invenstment['total_index']}*'基础数据3-投资使用计划与资金筹措'!B4",
                }
            },
            '5':{
                'C5': f"=附表1建设投资估算表!C3+附表1建设投资估算表!{self.invenstment['preservedindex']}+附表4投资使用计划与资金筹措表!C6",
                'C14': self.invenstment['tudi'],  # 直接注入原值，可考虑按excel公式输入，略麻烦
            },
            '6':{
                'C4': self.invenstment['wx'],
                'C7': self.invenstment['qt'],
            },
            '7':{
                f'{da_letter}5': "='基础数据4-表2、3、5、6、7、8、10'!$D$52*'基础数据4-表2、3、5、6、7、8、10'!$D$53",
                f'{da_letter}6': "='基础数据4-表2、3、5、6、7、8、10'!$D$55*'基础数据4-表2、3、5、6、7、8、10'!$D$56",
                f'{da_letter}7': "='基础数据4-表2、3、5、6、7、8、10'!$D$58*'基础数据4-表2、3、5、6、7、8、10'!$D$59",
                'Basic':{
                    'D4': "='基础数据1-项目基础信息'!C9",
                    'D5': f'=${da_letter}$5*D4',
                    'D6': f'=${da_letter}$6*D4',
                    'D7': f'=${da_letter}$7*D4',
                    'D8': "=附表5固定资产折旧表!F20*'基础数据4-表2、3、5、6、7、8、10'!$D$60",
                    'D9': '=附表5固定资产折旧表!F20',
                    'D10': '=附表6无形资产及其他资产摊销表!F11',
                    'D11': f'=附表3借款还本付息表!{get_column_letter(self.yy_col_index["3"])}10+附表2流动资金估算!E17',
                    'D12': "=(D5+D6+D7+D8)*'基础数据4-表2、3、5、6、7、8、10'!$D$61",
                    'D13': '=D5+D6+D7+D8+D9+D10+D11+D12',
                    'D14': '=D8+D9+D10+D11',
                    'D15': '=D13-D14',
                    'D16': '=D5+D6+D7+D8+D12',
                    'D17': "=D14/('附表8营业收入、营业税金及附加表'!D5-'附表8营业收入、营业税金及附加表'!D6-D15)",

                }
            },
            '8':{
                f'{da_letter}5': "='基础数据4-表2、3、5、6、7、8、10'!C66*'基础数据4-表2、3、5、6、7、8、10'!C67",
                'Basic':{
                    'D4': "='基础数据1-项目基础信息'!C9",
                    'D5': f'=${da_letter}$5*D4',
                    'D6': '=D7+D8+D9',
                    'D7': "=(D5-附表7总成本估算表!D5-附表7总成本估算表!D6)*'基础数据4-表2、3、5、6、7、8、10'!$C$68/(1+'基础数据4-表2、3、5、6、7、8、10'!$C$68)",
                    'D8': "=D7*'基础数据4-表2、3、5、6、7、8、10'!$C$69",
                    'D9': "=D7*'基础数据4-表2、3、5、6、7、8、10'!$C$70",
                }
            }
        }

        #  首次运行时添加，避免重复和删除冲突
        if sheetindex == '4':
            self.sheetFormula['4']['Basic']['1'].update({'D5': f"=附表1建设投资估算表!{self.invenstment['total_index']}*'基础数据3-投资使用计划与资金筹措'!B4"})
            self.sheetFormula['5']['C5'] = f"=附表1建设投资估算表!C3+附表1建设投资估算表!{self.invenstment['preservedindex']}+附表4投资使用计划与资金筹措表!C6"
            self.sheetFormula['5']['C14'] = float(self.invenstment['tudi'])  # 直接注入原值，可考虑按excel公式输入，略麻烦
            self.sheetFormula['6']['C4'] = self.invenstment['wx']
            self.sheetFormula['6']['C7'] = self.invenstment['qt']
            self.sheetFormula['7']['Basic']['1'].update(sheetFormula_rev['7']['Basic'])
            self.sheetFormula['7'][f'{da_letter}5'] = "='基础数据4-表2、3、5、6、7、8、10'!$D$52*'基础数据4-表2、3、5、6、7、8、10'!$D$53"
            self.sheetFormula['7'][f'{da_letter}6'] = "='基础数据4-表2、3、5、6、7、8、10'!$D$55*'基础数据4-表2、3、5、6、7、8、10'!$D$56"
            self.sheetFormula['7'][f'{da_letter}7'] = "='基础数据4-表2、3、5、6、7、8、10'!$D$58*'基础数据4-表2、3、5、6、7、8、10'!$D$59"
            self.sheetFormula['8']['Basic']['1'].update(sheetFormula_rev['8']['Basic'])
            self.sheetFormula['8'][f'{da_letter}5'] =  "='基础数据4-表2、3、5、6、7、8、10'!C66*'基础数据4-表2、3、5、6、7、8、10'!C67"

        origin_index = {
            '3': get_column_letter(self.yy_col_index['3']),
            '4': get_column_letter(self.yy_col_index['4']),
            '9': get_column_letter(self.yy_col_index['9']),
            '13': get_column_letter(self.yy_col_index['13']),
            '11': get_column_letter(self.yy_col_index['11']),
            '12': get_column_letter(self.yy_col_index['12']),
        }
        # print(origin_index)


        # 向sheetFormula字典中添加内容
        # 处理时间区域公式
        for basic_index, content in self.sheetFormula[sheetindex]['Basic'].items():
            # 录入Basic内容
            for cell, formula in content.items():
                # 进行公式转换
                index = 0
                if sheetindex in ['2', '5', '6', '7', '8', '10']:
                    ws[cell] = formula
                    for translation_times in range(self.fund_data['input1'] - self.fund_data['input2']-1):
                        index += 1
                        target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                        ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)
                if sheetindex in ['3','4','13','9']:
                    if basic_index == "1":  # 建设期公式
                        ws[cell] = formula
                        for translation_times in range(self.fund_data['input2'] - 1):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                            ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)
                    if basic_index == '2':
                        origin = origin_index[sheetindex] + cell[1:]
                        new_base = Translator(formula, origin=cell).translate_formula(origin)
                        ws[origin] = new_base
                        if origin != cell:
                            print("found",sheetindex, origin, formula)
                            print(cell,new_base)
                        for translation_times in range(self.fund_data['input1'] - self.fund_data['input2'] - 1):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(origin[0]) + index) + origin[1:]
                            ws[target_cor] = Translator(new_base, origin=origin).translate_formula(target_cor)
                if sheetindex in ['11','12']:
                    if basic_index == "1":  # 建设期公式
                        ws[cell] = formula
                        for translation_times in range(self.fund_data['input2']):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                            ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)
                    if basic_index == '2':
                        origin = origin_index[sheetindex] + cell[1:]
                        new_base = Translator(formula, origin=cell).translate_formula(origin)
                        ws[origin] = new_base
                        if origin != cell:
                            print("found",sheetindex, origin, formula)
                            print(cell,new_base)
                        for translation_times in range(self.fund_data['input1'] - self.fund_data['input2'] - 1):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(origin[0]) + index) + origin[1:]
                            ws[target_cor] = Translator(new_base, origin=origin).translate_formula(target_cor)

            # 表2，5，6，7，8，10只有一个Basic
            # 表3大致分为建设期和偿还期，其中建设期第一年，偿还期第一年的第五行有一点不同，可手动调整
            # 表9大致分为建设期和运营期，其中运营期最后一年的9、12、13行有一点不同，可手动调整
            # 表4、13可分为两个部分
            # 表11，12可分为建设期、建设期最后一年，运营期，运营期最后一年




        # 删除sheetFormula字典中的基础信息
        self.sheetFormula[sheetindex].pop('Basic')


        # 遍历字典中的单元格位置
        for cell_position, new_value in self.sheetFormula[sheetindex].items():
            # print(cell)
            # print(new_value)
            # 获取单元格
            cell = ws[cell_position]
            # 更新单元格的值
            cell.value = new_value

        # 处理表11和12中的数组公式
        ass_letter = get_column_letter(self.fund_data['input1']+1+3+1)
        array_create={
            "11":{
                     'D28': f"=PV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},D3,,-D15)",
                     'D29': "=C29+D28",
                     'D30': f"=PV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},D$3,,-D17)",
                     'D31': "=C31+D30",
                    f'{ass_letter}29': "=IF(AND(E29<0,F29>0),E3-E29/F28,0)",
                    f'{ass_letter}31': "=IF(AND(E31<0,F31>0),E3-E31/F30,0)",
                    f'{ass_letter}16': "=IF(AND(E16<0,F16>0),E3-E16/F15,0)",
                    f'{ass_letter}18': "=IF(AND(E18<0,F18>0),E3-E18/F17,0)"
            },
            "12":{
                'D28': f"=PV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},D3,,-D16)",
                'D29': "=C29+D28",
                'D30': f"=PV('基础数据1-项目基础信息'!$C${9 + self.fund_data['input1'] - self.fund_data['input2']},D$3,,-D19)",
                'D31': "=C31+D30",
                f'{ass_letter}29': "=IF(AND(E29<0,F29>0),E3-E29/F28,0)",
                f'{ass_letter}31': "=IF(AND(E31<0,F31>0),E3-E31/F30,0)",
                f'{ass_letter}17': "=IF(AND(E17<0,F17>0),E3-E17/F16,0)",
                f'{ass_letter}19': "=IF(AND(E19<0,F19>0),E3-E19/F18,0)"
            },
            }


        if sheetindex in ['11','12']:
            for cell, formula in array_create[sheetindex].items():
                ws[cell] = formula
                index = 0
                for translation_times in range(self.fund_data['input1']):
                    index += 1
                    target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                    ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)

            # # 设置数组公式
            # if sheetindex == "11":
            #     ws["E23"] = ArrayFormula("E23","=MATCH(TRUE,E18:J18>0,0)-1+(-INDEX(E18:J18,1,MATCH(TRUE,E18:J18>0,0)-1)/INDEX(E17:J17,1,MATCH(TRUE,E18:J18>0,0)))")
            #     ws["E24"] = ArrayFormula("E24",
            #                              "=MATCH(TRUE,E31:J31>0,0)-1+(-INDEX(E31:J31,1,MATCH(TRUE,E31:J31>0,0)-1)/INDEX(E30:J30,1,MATCH(TRUE,E31:J31>0,0)))")
            #     ws["F23"] = ArrayFormula("F23",
            #                              "=MATCH(TRUE,E16:J16>0,0)-1+(-INDEX(E16:J16,1,MATCH(TRUE,E16:J16>0,0)-1)/INDEX(E15:J15,1,MATCH(TRUE,E16:J16>0,0)))")
            #     ws["F24"] = ArrayFormula("F24",
            #                              "=MATCH(TRUE,E29:J29>0,0)-1+(-INDEX(E29:J29,1,MATCH(TRUE,E29:J29>0,0)-1)/INDEX(E28:J28,1,MATCH(TRUE,E29:J29>0,0)))")
            # else:
            #     ws["E23"] = ArrayFormula("E23","=MATCH(TRUE,E20:J20>0,0)-1+(-INDEX(E20:J20,1,MATCH(TRUE,E20:J20>0,0)-1)/INDEX(E19:J19,1,MATCH(TRUE,E20:J20>0,0)))")
            #     ws["E24"] = ArrayFormula("E24",
            #                              "=MATCH(TRUE,E31:J31>0,0)-1+(-INDEX(E31:J31,1,MATCH(TRUE,E31:J31>0,0)-1)/INDEX(E30:J30,1,MATCH(TRUE,E31:J31>0,0)))")
            #     ws["F23"] = ArrayFormula("F23",
            #                              "=MATCH(TRUE,E17:J17>0,0)-1+(-INDEX(E17:J17,1,MATCH(TRUE,E17:J17>0,0)-1)/INDEX(E16:J16,1,MATCH(TRUE,E17:J17>0,0)))")
            #     ws["F24"] = ArrayFormula("F24",
            #                              "=MATCH(TRUE,E29:J29>0,0)-1+(-INDEX(E29:J29,1,MATCH(TRUE,E29:J29>0,0)-1)/INDEX(E28:J28,1,MATCH(TRUE,E29:J29>0,0)))")
        # 设置数组公式
        if sheetindex == "11":
            ws["E23"] = f"=SUM({ass_letter}18:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}18)"
            ws["E24"] = f"=SUM({ass_letter}31:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}31)"
            ws["F23"] = f"=SUM({ass_letter}16:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}16)"
            ws["F24"] = f"=SUM({ass_letter}29:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}29)"
        else:
            ws[
                "E23"] = f"=SUM({ass_letter}19:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}19)"
            ws[
                "E24"] = f"=SUM({ass_letter}31:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}31)"
            ws[
                "F23"] = f"=SUM({ass_letter}17:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}17)"
            ws[
                "F24"] = f"=SUM({ass_letter}29:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}29)"

        # E23,F23,E24,F24涉及数组计算公式放到fillsheet里


        # 遍历工作表中的所有单元格并设定对齐
        for row in ws:
            # 遍历每一行中的所有单元格
            for cell in row:
                # 设定对齐
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def fillsheet_cal(self,sheetindex,workbook,start_row=1):
        # 建表思路：先根据时间+k（k与表相关，可以列一个数组堆进去）确定列数，然后分行渲染，第一行是表头，第二行是列名（含横向合并），第三行是年份（含纵向合并），从第四行开始是
        # 正文了，通过一个字典存储这些表结构，如果涉及到添加数据，那么就更新这个字典，否则就按照原字典分行渲染，最后将输入数据导入其中
        wb = workbook
        if start_row==1:
            ws = wb[self.sheetIndexs[sheetindex]]
        else:
            ws = wb.active
        # print(self.invenstment)
        # 计算达产期在表7的列名以帮助更新公式
        da_letter = get_column_letter(
            self.fund_data['input3'] + self.sheetAddCols['7'] - self.fund_data['input2'])  # (达产期-建设期+多余的添加列),表7和8公用

        sheetFormula_rev = {
            '4':{
                'Basic':{
                    'D5': f"=保留!{self.invenstment['total_index']}*保留!B16",
                }
            },
            '5':{
                'C5': f"=附表1建设投资估算表!C3+保留!{self.invenstment['preservedindex']}+附表4投资使用计划与资金筹措表!C6",
                'C14': self.invenstment['tudi'],  # 直接注入原值，可考虑按excel公式输入，略麻烦
                'C27': f"=附表1建设投资估算表!C3+保留!{self.invenstment['preservedindex']}"

            },
            '6':{
                'C4': self.invenstment['wx'],
                'C7': self.invenstment['qt'],
            },
            '7':{
                f'{da_letter}5': "=保留!$C$44*保留!$C$45",
                f'{da_letter}6': "=保留!$C$46*保留!$C$47",
                f'{da_letter}7': "=保留!$C$48*保留!$C$49",
                'Basic':{
                    'D4': "=保留!C6",
                    'D5': f'=${da_letter}$5*D4',
                    'D6': f'=${da_letter}$6*D4',
                    'D7': f'=${da_letter}$7*D4',
                    'D8': "=附表5固定资产折旧表!F20*保留!$C$50",
                    'D9': '=附表5固定资产折旧表!F20',
                    'D10': '=附表6无形资产及其他资产摊销表!E11',
                    'D11': f'=附表3借款还本付息表!{get_column_letter(self.yy_col_index["3"])}10+附表2流动资金估算!E17',
                    'D12': "=(D5+D6+D7+D8)*保留!$C$51",
                    'D13': '=D5+D6+D7+D8+D9+D10+D11+D12',
                    'D14': '=D8+D9+D10+D11',
                    'D15': '=D13-D14',
                    'D16': '=D5+D6+D7+D8+D12',
                    'D17': "=D14/('附表8营业收入、营业税金及附加表'!D5-'附表8营业收入、营业税金及附加表'!D6-D15)",

                }
            },
            '8':{
                f'{da_letter}5': "=保留!C52*保留!C53",
                'Basic':{
                    'D4': "=保留!C6",
                    'D5': f'=${da_letter}$5*D4',
                    'D6': '=D7+D8+D9',
                    'D7': "=(D5-附表7总成本估算表!D5-附表7总成本估算表!D6)*保留!$C$54/(1+保留!$C$54)",
                    'D8': "=D7*保留!$C$55",
                    'D9': "=D7*保留!$C$56",
                }
            }
        }


        def process_formula(formula, sheet_mapping):
            sheet_mapping['保留'] = 1
            def replace_reference(match):
                full_match = match.group(0)
                prefix = match.group(1) or ''  # 可能是 =, +, -, * , /,(
                sheet = match.group(2)
                dollar_column = match.group(3) or ''  # 列的 $ 符号
                column = match.group(4)
                dollar_row = match.group(5) or ''  # 行的 $ 符号
                row = match.group(6)

                # 移除 sheet 名称周围的引号（如果有）
                sheet = sheet.strip("'")

                # print(
                #     f"匹配: full_match='{full_match}',prefix ='{prefix}', sheet='{sheet}',dollar_column ={dollar_column}, column='{column}', row='{dollar_row}{row}'")

                if sheet in sheet_mapping:
                    new_row = int(row) + sheet_mapping[sheet] - 1
                    new_ref = f"{dollar_column}{column}{dollar_row}{new_row}"
                    # print(f"替换: {full_match} -> {new_ref}")
                    return f"{prefix}{new_ref}"
                else:
                    # print(f"未替换: {full_match} (sheet 不在映射中)")
                    return full_match

            # 更新正则表达式以正确匹配公式开头的等号、sheet 名称、列和行，并保留所有 $ 符号
            pattern = r"(=|,|\+|-|\*|/|\(|:)('?[^'a-zA-Z\+!-(]+'?)!(\$?)([A-Z]+)(\$?)(\d+)"

            # print(f"处理公式: {formula}")
            # 替换所有匹配的引用
            new_formula = re.sub(pattern, replace_reference, formula)
            # print(f"处理后的公式: {new_formula}")
            # print()
            return new_formula

        def process_multiple_formulas(formulas, sheet_mapping):
            def process_item(item):
                if isinstance(item, str):
                    return process_formula(item, sheet_mapping)
                elif isinstance(item, dict):
                    return {k: process_item(v) for k, v in item.items()}
                else:
                    return item

            return process_item(formulas)

        def relocate_formula(formula, sheet_mapping,ind):
            def replace_reference(match):
                full_match = match.group(0)
                prefix = match.group(1) or ''  # 可能是 =, +, -, * , /,(
                dollar_column = match.group(2) or ''  # 列的 $ 符号
                column = match.group(3)
                dollar_row = match.group(4) or ''  # 行的 $ 符号
                row = match.group(5)

                # print(
                #     f"匹配: full_match='{full_match}',prefix ='{prefix}', dollar_column ={dollar_column}, column='{column}', row='{dollar_row}{row}'")

                return prefix+dollar_column+column+dollar_row+str(int(row)+sheet_mapping[self.sheetIndexs[ind]]-1)

            # 更新正则表达式以正确匹配公式开头的等号、sheet 名称、列和行，并保留所有 $ 符号
            pattern = r"([^!$]{1})(\$?)([A-Z])(\$?)(\d+)"

            # print(f"定位公式: {formula}")
            # 替换所有匹配的引用
            new_formula = re.sub(pattern, replace_reference, formula)
            # print(f"重定位公式: {new_formula}")
            # print()
            return new_formula

        def relocate_multiple_formulas(formulas, sheet_mapping,ind):
            def process_item(item):
                if isinstance(item, str):
                    return relocate_formula(item, sheet_mapping,ind)
                elif isinstance(item, dict):
                    ans = {}
                    for k, v in item.items():
                        if k != 'Basic':
                            temp = process_item("~"+k)
                            ans[temp[1:]] = process_item(v)
                        else:
                            ans[k] = process_item(v)
                    return ans
                else:
                    return item

            return process_item(formulas)

        # if sheetindex in sheetFormula_rev.keys():
        #     processed_sheetFormula_rev={}
        #     newsheetdicts = relocate_multiple_formulas(sheetFormula_rev[sheetindex], start_row,sheetindex)
        #     processed_sheetFormula_rev[sheetindex] = process_multiple_formulas(newsheetdicts, start_row)

        #  首次运行时添加，避免重复和删除冲突
        processed_sheetFormula={}
        if sheetindex == '2':
            self.calsheetFormula['4']['Basic']['1'].update({'D5': f"=保留!{self.invenstment['total_index']}*保留!B16"})
            self.calsheetFormula['5']['C5'] = f"=附表1建设投资估算表!C3+保留!{self.invenstment['preservedindex']}+附表4投资使用计划与资金筹措表!C6"
            self.calsheetFormula['5']['C14'] = float(self.invenstment['tudi'])  # 直接注入原值，可考虑按excel公式输入，略麻烦
            self.calsheetFormula['5'][
                'C27'] = f"=附表1建设投资估算表!C3+保留!{self.invenstment['preservedindex']}"
            self.calsheetFormula['6']['C4'] = self.invenstment['wx']
            self.calsheetFormula['6']['C7'] = self.invenstment['qt']
            self.calsheetFormula['7']['Basic']['1'].update(sheetFormula_rev['7']['Basic'])
            self.calsheetFormula['7'][f'{da_letter}5'] = "=保留!$C$44*保留!$C$45"
            self.calsheetFormula['7'][f'{da_letter}6'] = "=保留!$C$46*保留!$C$47"
            self.calsheetFormula['7'][f'{da_letter}7'] = "=保留!$C$48*保留!$C$49"
            self.calsheetFormula['8']['Basic']['1'].update(sheetFormula_rev['8']['Basic'])
            self.calsheetFormula['8'][f'{da_letter}5'] = "=保留!C52*保留!C53"
        newsheetdicts = relocate_multiple_formulas(self.calsheetFormula[sheetindex], start_row, sheetindex)
        processed_sheetFormula[sheetindex] = process_multiple_formulas(newsheetdicts, start_row)

        # print(processed_sheetFormula)

        origin_index = {
            '3': get_column_letter(self.yy_col_index['3']),
            '4': get_column_letter(self.yy_col_index['4']),
            '9': get_column_letter(self.yy_col_index['9']),
            '13': get_column_letter(self.yy_col_index['13']),
            '11': get_column_letter(self.yy_col_index['11']),
            '12': get_column_letter(self.yy_col_index['12']),
        }
        # print(origin_index)
        startrow = start_row[self.sheetIndexs[sheetindex]]
        # 涉及折旧和还款方式的选项
        if sheetindex == "3":
            # 处理还款方式,默认等额本金不用修改
            if self.fund_data['repayMethod'] == "ave_cap_int":
                # 等额本息替换三个公式
                keys = list(processed_sheetFormula[sheetindex]['Basic']['2'].keys())
                processed_sheetFormula[sheetindex]['Basic']['2'][keys[0]] = f"=IF(G{startrow+2}-$C$5>$C$35,0,PMT($C$33,$C$35,-$G${startrow+4}))"
                processed_sheetFormula[sheetindex]['Basic']['2'][keys[1]] = f"=IF(G{startrow+2}-$C$5>$C$35,0,PPMT($C$33,G{startrow+2}-$C$5,$C$35,-$G${startrow+4}))"
                processed_sheetFormula[sheetindex]['Basic']['2'][keys[2]] = f"=G{startrow+7} - G{startrow+8}"
        elif sheetindex == "5":
            # 处理折旧方式,默认平均年限法不用修改
            if self.fund_data['depreciationMethod'] == "DDB":
                keys = list(processed_sheetFormula[sheetindex]['Basic']['1'].keys())
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[2]] = f"=IF(F{startrow+2}-$C$5-$D${startrow+3}<-1,DDB($C${startrow+4},$C${startrow+4}*$E${startrow+3},$D${startrow+3},F{startrow+2}-$C$5,2),(E{startrow+6}-$C${startrow+4}*$E${startrow+3})/($D${startrow+3}-F{startrow+2}+1+$C$5))"
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[0]] = f"=IF(F{startrow+2}-$C$5-$D${startrow+3}<-1,DDB($C${startrow+26},$C${startrow+26}*$E${startrow+3},$D${startrow+3},F{startrow+2}-$C$5,2),(E{startrow+6}-$C${startrow+26}*$E${startrow+3})/($D${startrow+3}-F{startrow+2}+1+$C$5))"
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[4]] = f"=IF(F{startrow+2}-$C$5-$D${startrow+7}<-1,DDB($C${startrow+8},$C${startrow+8}*$E${startrow+7},$D${startrow+7},F{startrow+2}-$C$5,2),(E{startrow+10}-$C${startrow+8}*$E${startrow+7})/($D${startrow+7}-F{startrow+2}+1+$C$5))"
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[6]] = f"=IF(F{startrow+2}-$C$5-$D${startrow+12}<-1,DDB($C${startrow+13},0,$D${startrow+12},F{startrow+2}-$C$5,2),E{startrow+15}/($D${startrow+12}-F{startrow+2}+1+$C$5))"
            elif self.fund_data['depreciationMethod'] == "SYD":
                keys = list(processed_sheetFormula[sheetindex]['Basic']['1'].keys())
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[
                    2]] = f"=SYD($C${startrow+4},$C${startrow+4}*$E${startrow+3},$D${startrow+3},F{startrow+2}-$C$5)"
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[
                    0]] = f"=SYD($C${startrow+26},$C${startrow+26}*$E${startrow+3},$D${startrow+3},F{startrow+2}-$C$5)"
                processed_sheetFormula[sheetindex]['Basic']['1'][keys[
                    4]] = f"=SYD($C${startrow+8},$C${startrow+8}*$E${startrow+7},$D${startrow+7},F{startrow+2}-$C$5)"
                processed_sheetFormula[sheetindex]['Basic']['1'][
                    keys[6]] = f"=SYD($C${startrow+13},0,$D${startrow+12},F{startrow+2}-$C$5)"

        # 向sheetFormula字典中添加内容
        # 处理时间区域公式
        for basic_index, content in processed_sheetFormula[sheetindex]['Basic'].items():
            # 录入Basic内容
            for cell, formula in content.items():
                # 进行公式转换
                index = 0
                if sheetindex in ['2', '5', '6', '7', '8', '10','14']:
                    ws[cell] = formula
                    for translation_times in range(self.fund_data['input1'] - self.fund_data['input2']-1):
                        index += 1
                        target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                        ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)
                if sheetindex in ['3','4','13','9']:
                    if basic_index == "1":  # 建设期公式
                        ws[cell] = formula
                        for translation_times in range(self.fund_data['input2'] - 1):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                            ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)
                    if basic_index == '2':
                        origin = origin_index[sheetindex] + cell[1:]
                        new_base = Translator(formula, origin=cell).translate_formula(origin)
                        ws[origin] = new_base
                        if origin != cell:
                            print("found",sheetindex, origin, formula)
                            print(cell,new_base)
                        for translation_times in range(self.fund_data['input1'] - self.fund_data['input2'] - 1):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(origin[0]) + index) + origin[1:]
                            ws[target_cor] = Translator(new_base, origin=origin).translate_formula(target_cor)
                if sheetindex in ['11','12']:
                    if basic_index == "1":  # 建设期公式
                        ws[cell] = formula
                        for translation_times in range(self.fund_data['input2']):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                            ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)
                    if basic_index == '2':
                        origin = origin_index[sheetindex] + cell[1:]
                        new_base = Translator(formula, origin=cell).translate_formula(origin)
                        ws[origin] = new_base
                        if origin != cell:
                            print("found",sheetindex, origin, formula)
                            print(cell,new_base)
                        for translation_times in range(self.fund_data['input1'] - self.fund_data['input2'] - 1):
                            index += 1
                            target_cor = get_column_letter(column_index_from_string(origin[0]) + index) + origin[1:]
                            ws[target_cor] = Translator(new_base, origin=origin).translate_formula(target_cor)

            # 表2，5，6，7，8，10只有一个Basic
            # 表3大致分为建设期和偿还期，其中建设期第一年，偿还期第一年的第五行有一点不同，可手动调整
            # 表9大致分为建设期和运营期，其中运营期最后一年的9、12、13行有一点不同，可手动调整
            # 表4、13可分为两个部分
            # 表11，12可分为建设期、建设期最后一年，运营期，运营期最后一年




        # 删除sheetFormula字典中的基础信息
        processed_sheetFormula[sheetindex].pop('Basic')


        # 遍历字典中的单元格位置
        for cell_position, new_value in processed_sheetFormula[sheetindex].items():
            # print(cell)
            # print(new_value)
            # 获取单元格
            cell = ws[cell_position]
            # 更新单元格的值
            cell.value = new_value
        # 处理表11和12中的数组公式
        ass_letter = get_column_letter(self.fund_data['input1'] + 1 + 3 + 1)
        array_create={
            "11":{
                     'D28': f"=PV(保留!$C$7,D3,,-D15)",
                     'D29': "=C29+D28",
                     'D30': f"=PV(保留!$C$7,D$3,,-D17)",
                     'D31': "=C31+D30",
                f'{ass_letter}29': "=IF(AND(E29<0,F29>0),E3-E29/F28,0)",
                f'{ass_letter}31': "=IF(AND(E31<0,F31>0),E3-E31/F30,0)",
                f'{ass_letter}16': "=IF(AND(E16<0,F16>0),E3-E16/F15,0)",
                f'{ass_letter}18': "=IF(AND(E18<0,F18>0),E3-E18/F17,0)"
            },
            "12":{
                'D28': f"=PV(保留!$C$7,D3,,-D16)",
                'D29': "=C29+D28",
                'D30': f"=PV(保留!$C$7,D$3,,-D19)",
                'D31': "=C31+D30",
                f'{ass_letter}29': "=IF(AND(E29<0,F29>0),E3-E29/F28,0)",
                f'{ass_letter}31': "=IF(AND(E31<0,F31>0),E3-E31/F30,0)",
                f'{ass_letter}17': "=IF(AND(E17<0,F17>0),E3-E17/F16,0)",
                f'{ass_letter}20': "=IF(AND(E20<0,F20>0),E3-E20/F19,0)"
            },
            }
        array_create["11"] = relocate_multiple_formulas(array_create["11"],start_row,"11")
        array_create["11"] = process_multiple_formulas(array_create["11"], start_row)
        array_create["12"] = relocate_multiple_formulas(array_create["12"], start_row,"12")
        array_create["12"] = process_multiple_formulas(array_create["12"], start_row)
        if sheetindex in ['11','12']:
            for cell, formula in array_create[sheetindex].items():
                ws[cell] = formula
                index = 0
                times = self.fund_data['input1'] if cell[0] == "D" else self.fund_data['input1'] - 2
                for translation_times in range(times):
                    index += 1
                    target_cor = get_column_letter(column_index_from_string(cell[0]) + index) + cell[1:]
                    ws[target_cor] = Translator(formula, origin=cell).translate_formula(target_cor)

            # 设置数组公式
            array_formula = {
            "11":{
                     "E23": f"=SUM({ass_letter}18:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}18)",
                     "E24": f"=SUM({ass_letter}31:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}31)",
                     "F23":  f"=SUM({ass_letter}16:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}16)",
                     "F24": f"=SUM({ass_letter}29:{get_column_letter(column_index_from_string(ass_letter)+self.fund_data['input1']-1)}29)"},
            "12":{
                "E23": f"=SUM({ass_letter}20:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}20)",
                "E24":  f"=SUM({ass_letter}31:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}31)",
                "F23": f"=SUM({ass_letter}17:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}17)",
                "F24": f"=SUM({ass_letter}29:{get_column_letter(column_index_from_string(ass_letter) + self.fund_data['input1'] - 1)}29)"},
            }
            array_formula["11"] = relocate_multiple_formulas(array_formula["11"], start_row, "11")
            array_formula["11"] = process_multiple_formulas(array_formula["11"], start_row)
            array_formula["12"] = relocate_multiple_formulas(array_formula["12"], start_row, "12")
            array_formula["12"] = process_multiple_formulas(array_formula["12"], start_row)
            if sheetindex in array_formula.keys():
                for address,formula in array_formula[sheetindex].items():
                    ws[address] = formula



        # 遍历工作表中的所有单元格并设定对齐
        for row in ws:
            # 遍历每一行中的所有单元格
            for cell in row:
                # 设定对齐
                cell.alignment = Alignment(horizontal='center', vertical='center')


    def toexcel(self,module_path="D:/Program Data/flask-html/基础数据.xlsx"):
        # 把fdata(df格式)写进excel文件里

        # 复制基础数据模板
        source_wb = load_workbook(module_path)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in source_wb.sheetnames:
            source_sheet = source_wb[sheet_name]
            target_sheet = workbook.create_sheet(title=sheet_name)

            for row in source_sheet.rows:
                for cell in row:
                    target_sheet[cell.coordinate].value = cell.value

        # 保存新工作簿
        uid = "0000"
        current_time = datetime.datetime.now()
        create_time = current_time.strftime("%Y_%m_%d_%H_%M_%S")
        projectname = self.fund_data['projectname']
        workbookname = projectname + "_" + uid + "_" + create_time
        # workbook.save(f'{workbookname}.xlsx')

        # 将fund_data中的数据写入基础数据中
        for name in self.datasheetNames:
            # 尝试通过名称访问工作表
            try:
                worksheet = workbook[name]
                # print('found')
            except KeyError:
                # 如果工作表不存在，则创建一个名为'表1'的新工作表
                print("unfound worksheet",name)
            # 基础数据1的写入
            if name == self.datasheetNames[0]:
                worksheet['C2'] = self.fund_data['projectname'] if self.fund_data['projectname'] != '' else "缺失"
                worksheet['C3'] = self.fund_data['projectplace'] if self.fund_data['projectplace'] != '' else "缺失"
                worksheet['C4'] = self.fund_data['projectind'] if self.fund_data['projectind'] != '' else "缺失"
                worksheet['C5'] = self.fund_data['input1'] if self.fund_data['input1'] != '' else "缺失"
                worksheet['C6'] = self.fund_data['input2'] if self.fund_data['input2'] != '' else "缺失"
                worksheet['C7'] = self.fund_data['input1'] - self.fund_data['input2'] if self.fund_data['input2'] != '' and self.fund_data['input1'] != '' and self.fund_data['input1'] - self.fund_data['input2']>0 else "缺失"
                index = 8
                colindex = 3
                if self.fund_data['input2'] != '' and self.fund_data['input1'] != '' and self.fund_data['input1'] - self.fund_data['input2']>0:
                    for loop in range(self.fund_data['input2'] + 1,self.fund_data['input1'] + 1):
                        index += 1
                        worksheet[f'A{index}'] = ''
                        worksheet[f'B{index}'] = f'第{loop}年'
                        worksheet[f'C{index}'] = float(self.fund_data[f'productionLoad{loop}']) / 100
                        worksheet[f'{get_column_letter(colindex)}9'] = float(self.fund_data[f'productionLoad{loop}']) / 100
                        colindex += 1
                index += 1
                worksheet[f'A{index}'] = 8
                worksheet[f'B{index}'] = '基准收益率i0'
                worksheet[f'C{index}'] = float(self.fund_data['benchmarkyield']) / 100 if self.fund_data['benchmarkyield'] != '' else "缺失"
                index += 1
                worksheet[f'A{index}'] = 9
                worksheet[f'B{index}'] = '基准静态回收期'
                worksheet[f'C{index}'] = float(self.fund_data['benchmark_s_paybackperiod']) if self.fund_data['benchmark_s_paybackperiod'] != '' else "缺失"
                index += 1
                worksheet[f'A{index}'] = 10
                worksheet[f'B{index}'] = '基准动态回收期'
                worksheet[f'C{index}'] = float(self.fund_data['benchmark_d_paybackperiod']) if self.fund_data['benchmark_d_paybackperiod'] != '' else "缺失"
            elif name == self.datasheetNames[1]:
                worksheet['C4'] = float(self.fund_data['install_rate']) / 100 if self.fund_data['install_rate'] != '' else "缺失"
                worksheet['C5'] = float(self.fund_data['trans_rate']) / 100 if self.fund_data['trans_rate'] != '' else "缺失"
                worksheet['C6'] = float(self.fund_data['pre_rate']) / 100 if self.fund_data['pre_rate'] != '' else "缺失"
                worksheet['C7'] = float(self.fund_data['preup_rate']) / 100 if self.fund_data['preup_rate'] != '' else "缺失"

                insert_index = 12  # 初始的工程费用行数为3

                for i in range(int(self.fund_data['projectcostCounter'])):
                    insert_index += 1
                    worksheet.insert_rows(insert_index)
                    # 新建项目赋值
                    worksheet[f'A{insert_index}'] = f'{i+1}'
                    worksheet[f'B{insert_index}'] = self.fund_data[f'additionalCostName_{i+1}']
                    worksheet[f'C{insert_index}'] = f'=F{insert_index}*G{insert_index} / 10000'
                    worksheet[f'D{insert_index}'] = float(self.fund_data[f'additionalCostEquipment_{i + 1}'])
                    worksheet[f'E{insert_index}'] = f"=D{insert_index}*C4*(1-C5)"
                    worksheet[f'F{insert_index}'] = self.fund_data[f'additionalCostAmount_{i + 1}']
                    worksheet[f'G{insert_index}'] = self.fund_data[f'additionalCostPrice_{i + 1}']

                insert_index += 4  # 切换到其他费用基础数据的位置

                for i in range(int(self.fund_data['othercostCounter'])):
                    insert_index += 1
                    worksheet.insert_rows(insert_index)
                    # 新建项目赋值
                    worksheet[f'A{insert_index}'] = f'{i+1}'
                    # print(self.fund_data)
                    worksheet[f'B{insert_index}'] = self.fund_data[f'additionalotherCostName_{i+1}']
                    worksheet[f'C{insert_index}'] = float(self.fund_data[f'additionalOtherCost_{i + 1}'])
            elif name == self.datasheetNames[2]:
                inv_plan = ast.literal_eval(self.fund_data['investmentPlan'])
                loan_plan = ast.literal_eval(self.fund_data['LoanPlan'])
                inv_colindex = 'B'
                for year,values in inv_plan.items():
                    worksheet[f'{inv_colindex}3'] = year
                    worksheet[f'{inv_colindex}4'] = float(values['percentage']) / 100
                    inv_colindex = get_column_letter(column_index_from_string(inv_colindex) + 1)

                worksheet[f'{inv_colindex}3'] = len(inv_plan.items()) + 1
                worksheet[f'{inv_colindex}4'] = float(self.fund_data['leftInvestmentPercentage']) / 100

                inv_colindex = 'B'
                for year,values in loan_plan.items():
                    worksheet[f'{inv_colindex}7'] = year
                    worksheet[f'{inv_colindex}8'] = float(values['percentage']) / 100
                    worksheet[f'{inv_colindex}9'] = f"=1-{inv_colindex}8"
                    inv_colindex = get_column_letter(column_index_from_string(inv_colindex) + 1)

            else:
                # 表2
                for i in range(3,10):
                    worksheet[f'C{i}'] = int(self.fund_data[f'days{i-2}']) if self.fund_data[f'days{i-2}'] != '' else "缺失"
                worksheet[f'C10'] = float(self.fund_data[f'days8']) / 100 if self.fund_data[f'days8'] != '' else "缺失"
                worksheet[f'C11'] = float(self.fund_data[f'days9']) / 100 if self.fund_data[f'days9'] != '' else "缺失"
                # 表3
                worksheet['C20'] = float(self.fund_data['loadrate']) / 100 if self.fund_data['loadrate'] != '' else "缺失"
                if self.fund_data['repayMethod'] == 'ave_capital':
                    worksheet['B23'] = int(self.fund_data['payTime']) if self.fund_data['payTime'] != '' else "缺失"
                else:
                    worksheet['B24'] = int(self.fund_data['payTime']) if self.fund_data['payTime'] != '' else "缺失"
                # 表5(缺少折旧方式)
                worksheet['C32'] = int(self.fund_data['dep_year2']) if self.fund_data['dep_year2'] != '' else "缺失"
                worksheet['C33'] = int(self.fund_data['dep_year1']) if self.fund_data['dep_year1'] != '' else "缺失"
                worksheet['C35'] = int(self.fund_data['dep_year3']) if self.fund_data['dep_year3'] != '' else "缺失"
                worksheet['D32'] = float(self.fund_data['res_rate2']) / 100 if self.fund_data['res_rate2'] != '' else "缺失"
                worksheet['D33'] = float(self.fund_data['res_rate1']) / 100 if self.fund_data['res_rate1'] != '' else "缺失"
                # 表6(保留与其他费用那里更新的可能，也可以归为无形资产和其他资产两个分类)
                worksheet['C46'] = int(self.fund_data['amo_year1']) if self.fund_data['amo_year1'] != '' else "缺失"
                worksheet['C47'] = int(self.fund_data['amo_year2']) if self.fund_data['amo_year2'] != '' else "缺失"
                # 表7(考虑前端只输入产量的单位，然后label变为 (单位/万元）)
                worksheet['D52'] = float(self.fund_data['production1']) if self.fund_data['production1'] != '' else "缺失"
                worksheet['E52'] = self.fund_data['dan1'] if self.fund_data['dan1'] != '' else "缺失"
                worksheet['D53'] = float(self.fund_data['production2']) if self.fund_data['production2'] != '' else "缺失"
                worksheet['E53'] = self.fund_data['dan2'] if self.fund_data['dan2'] != '' else "缺失"
                worksheet['D55'] = float(self.fund_data['production3']) if self.fund_data['production3'] != '' else "缺失"
                worksheet['E55'] = self.fund_data['dan3'] if self.fund_data['dan3'] != '' else "缺失"
                worksheet['D56'] = float(self.fund_data['production4']) if self.fund_data['production4'] != '' else "缺失"
                worksheet['E56'] = self.fund_data['dan4'] if self.fund_data['dan4'] != '' else "缺失"
                worksheet['D58'] = int(self.fund_data['production5']) if self.fund_data['production5'] != '' else "缺失"
                worksheet['D59'] = float(self.fund_data['production6'])if self.fund_data['production6'] != '' else "缺失"
                worksheet['D60'] = float(self.fund_data['production7']) / 100 if self.fund_data['production7'] != '' else "缺失"
                worksheet['D61'] = float(self.fund_data['production8']) / 100 if self.fund_data['production8'] != '' else "缺失"
                # 表8
                worksheet['C66'] = float(self.fund_data['production9']) if self.fund_data['production9'] != '' else "缺失"
                worksheet['C67'] = float(self.fund_data['production10']) if self.fund_data['production10'] != '' else "缺失"
                worksheet['C68'] = float(self.fund_data['production11']) / 100 if self.fund_data['production11'] != '' else "缺失"
                worksheet['C69'] = float(self.fund_data['production12']) / 100 if self.fund_data[
                                                                               'production12'] != '' else "缺失"
                worksheet['C70'] = float(self.fund_data['production13']) / 100 if self.fund_data[
                                                                               'production13'] != '' else "缺失"

                # 表10
                worksheet['C74'] = float(self.fund_data['production14']) / 100 if self.fund_data[
                                                                               'production14'] != '' else "缺失"
                worksheet['C75'] = float(self.fund_data['production15']) / 100 if self.fund_data[
                                                                               'production15'] != '' else "缺失"
                worksheet['C76'] = float(self.fund_data['production16']) / 100 if self.fund_data[
                                                                               'production16'] != '' else "缺失"

                # 尝试通过名称访问工作表
        try:
            worksheet = workbook['附表1建设投资估算表']
            # print('found')
        except KeyError:
            # 如果工作表不存在，则创建一个名为'表1'的新工作表
            worksheet = workbook.create_sheet('附表1建设投资估算表')

        # 建立表1
        # 表格第一行
        worksheet['A1'] = '附表1 : 建设投资估算表'
        worksheet['H1'] = '单位：万元'
        worksheet['A2'] = '序号'
        worksheet['B2'] = '工程或费用名称'
        worksheet['C2'] = '建筑工程'
        worksheet['D2'] = '设备工程'
        worksheet['E2'] = '安装工程'
        worksheet['F2'] = '其他费用'
        worksheet['G2'] = '合计'
        worksheet['H2'] = '占总值比'
        # 表格第一列
        worksheet['A3'] = '1'
        worksheet['A4'] = '2'
        worksheet['A5'] = ''
        worksheet['A6'] = '3'
        worksheet['A7'] = '3.1'
        worksheet['A8'] = '3.2'
        worksheet['A9'] = '4'

        worksheet['B3'] = '第一部分：工程费用'
        worksheet['B4'] = '第二部分：其他费用'
        worksheet['B5'] = '第一、二部分费用合计'
        worksheet['B6'] = '预备费用'
        worksheet['B7'] = '基本预备费'
        worksheet['B8'] = '涨价预备费'
        worksheet['B9'] = '建设投资合计（1+2+3）'

        # 遍历工作表中的所有行并设定所有对齐
        for row in worksheet:
            # 遍历每一行中的所有单元格
            for cell in row:
                # 清除单元格的值
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 合并标题
        mergeexcel('A1','G1',worksheet)

        # 插入工程费用栏目
        insert_index = 3  # 初始的工程费用行数为3
        total_index = 8 + int(self.fund_data['projectcostCounter']) + int(self.fund_data['othercostCounter']) + 1
        self.invenstment['total_index'] = f'$G${total_index}'

        for i in range(int(self.fund_data['projectcostCounter'])):
            insert_index += 1
            worksheet.insert_rows(insert_index)
            # 新建项目赋值
            worksheet[f'A{insert_index}'] = f'1.{i+1}'
            worksheet[f'B{insert_index}'] = f"='基础数据2-建设投资估算 '!B{insert_index+9}"
            worksheet[f'C{insert_index}'] = f"='基础数据2-建设投资估算 '!C{insert_index+9}"
            worksheet[f'D{insert_index}'] = f"='基础数据2-建设投资估算 '!D{insert_index+9}"
            worksheet[f'E{insert_index}'] = f"='基础数据2-建设投资估算 '!E{insert_index+9}"
            # worksheet[f'B{insert_index}'] = self.fund_data[f'additionalCostName_{i+1}']
            # worksheet[f'C{insert_index}'] = int(float(self.fund_data[f'additionalCostAmount_{i + 1}']) * float(self.fund_data[f'additionalCostPrice_{i + 1}']))
            # worksheet[f'D{insert_index}'] = int(self.fund_data[f'additionalCostEquipment_{i + 1}'])
            # worksheet[f'E{insert_index}'] = f"=D{insert_index}*{self.fund_data['input7']}/100"
            worksheet[f'G{insert_index}'] = f"=C{insert_index}+D{insert_index}+E{insert_index}"
            worksheet[f'H{insert_index}'] = f"=G{insert_index}/G{total_index}"
            worksheet[f'H{insert_index}'].number_format = "0.0%"


        # 更新阶段求和
        if self.fund_data['projectcostCounter'] != "0":
            worksheet['C3'] = f"=SUM(C4:C{insert_index})"
            worksheet['D3'] = f"=SUM(D4:D{insert_index})"
            worksheet['E3'] = f"=SUM(E4:E{insert_index})"
            worksheet['G3'] = f"=SUM(G4:G{insert_index})"
            worksheet['H3'] = f"=G3/G{total_index}"
            worksheet['H3'].number_format = "0.0%"
        #
        # 插入其它费用栏目
        insert_index += 1  # 略过其他费用合计一行
        otherindex = insert_index  # 记录下其他费用合计的行数

        for i in range(int(self.fund_data['othercostCounter'])):
            insert_index += 1
            worksheet.insert_rows(insert_index)
            worksheet[f'A{insert_index}'] = f'2.{i+1}'
            worksheet[f'B{insert_index}'] = f"='基础数据2-建设投资估算 '!B{insert_index+12}"
            worksheet[f'F{insert_index}'] = f"='基础数据2-建设投资估算 '!C{insert_index+12}"
            worksheet[f'G{insert_index}'] = f"=F{insert_index}"
            worksheet[f'H{insert_index}'] = f"=G{insert_index}/G{total_index}"
            worksheet[f'H{insert_index}'].number_format = "0.0%"

        # 更新阶段求和
        if self.fund_data['othercostCounter'] != "0":
            worksheet[f'F{otherindex}'] = f"=SUM(F{otherindex+1}:F{insert_index})"
            worksheet[f'G{otherindex}'] = f"=SUM(G{otherindex + 1}:G{insert_index})"
            worksheet[f'H{otherindex}'] = f"=G{otherindex}/G{total_index}"
            worksheet[f'H{otherindex}'].number_format = "0.0%"

        # 更新一二部分求和

        insert_index += 1
        worksheet[f'G{insert_index}'] = f"=G3+G{otherindex}"

        # 更新预备费用
        preservedindex = insert_index + 1
        self.invenstment['preservedindex'] = f'$G${preservedindex}'
        worksheet[f'F{preservedindex + 1}'] = f"=G10*'基础数据2-建设投资估算 '!C6"   #基本预备费
        worksheet[f'F{preservedindex + 2}'] = f"=G3*'基础数据2-建设投资估算 '!C7"   #涨价预备费
        worksheet[f'F{preservedindex}'] = f"=F{preservedindex + 1}+F{preservedindex + 2}"
        worksheet[f'G{preservedindex + 1}'] = f"=F{preservedindex + 1}"   #基本预备费总值
        worksheet[f'G{preservedindex + 2}'] = f"=F{preservedindex + 2}"   #涨价预备费总值
        worksheet[f'G{preservedindex}'] = f"=G{preservedindex + 1}+G{preservedindex + 2}"
        worksheet[f'H{preservedindex + 1}'] = f"=G{preservedindex + 1}/G{total_index}"
        worksheet[f'H{preservedindex + 2}'] = f"=G{preservedindex + 2}/G{total_index}"
        worksheet[f'H{preservedindex}'] = f"=G{preservedindex}/G{total_index}"
        worksheet[f'H{preservedindex}'].number_format = "0.0%"
        worksheet[f'H{preservedindex + 1}'].number_format = "0.0%"
        worksheet[f'H{preservedindex + 2}'].number_format = "0.0%"

        # 更新合计费用
        worksheet[f'G{total_index}'] = f"=G3+G{otherindex}+G{preservedindex}"
        worksheet[f'H{total_index}'] = f"=G{total_index}/G{total_index}"
        worksheet[f'H{total_index}'].number_format = "0.0%"

        # 修改显示
        cell_range = worksheet['C3':'G16']
        for each_cell in cell_range:
            for each in each_cell:
                each.number_format = "0"

        # 采用makesheet方法建表2-13
        for iloop in range(2,14):
            ind = str(iloop)
            self.makesheet(ind,workbook)
            self.fillsheet(ind,workbook)

        new_excel_path = "D:/Program Data/flask-html/" + workbookname + ".xlsx"


        # 保存工作簿
        workbook.save(filename=new_excel_path)  # 如果文件已存在，它将被覆盖

        # from win32com.client import Dispatch
        # xlApp = Dispatch("Excel.Application")
        # xlApp.Visible = False
        # xlBook = xlApp.Workbooks.Open(new_excel_path)
        # xlBook.Save()
        # xlBook.Close()

        return new_excel_path

    def getData(self,name,type="0"):
        try:
            if type == "int":
                get = int(self.fund_data[name]) if self.fund_data[name] != '' else "缺失"
            elif type == "float":
                get = float(self.fund_data[name]) if self.fund_data[name] != '' else "缺失"
            elif type == "percent":
                get = float(self.fund_data[name]) / 100 if self.fund_data[name] != '' else "缺失"
            else:
                get = self.fund_data[name] if self.fund_data[name] != '' else "缺失"
            return get
        except KeyError:
            return "缺失"

    def makexspreadsheet(self, sheetindex, xspreadsheet,actual_len=-1):
        # xspreadsheet为字典形式，应包含除rows外其它所有（大部分merges）
        if actual_len == -1:
            actual_len = xspreadsheet["rows"]["len"]
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            colIndex = self.fund_data['input1'] + self.sheetAddCols[sheetindex]
            if sheetindex in ['11', '12']:
                colIndex += 1
        else:
            colIndex = self.fund_data['input1'] - self.fund_data['input2'] + self.sheetAddCols[sheetindex]
        colIndexLetter = number_to_letter(colIndex)
        col_index = colIndex - 1
        xspreadsheet["rows"]["0"]["cells"][str(col_index)] = {"text": "单位：万元", "style": 1}

        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            jlen = self.fund_data['input2']-1 if sheetindex not in ['11','12'] else self.fund_data['input2']
            xspreadsheet["rows"]["1"]["cells"][str(self.sheetAddCols[sheetindex])] = {"text": "建设期", "merge": [0, jlen], "style": 1}
            xspreadsheet["merges"].append(f"{number_to_letter(self.sheetAddCols[sheetindex] + 1)}2:{number_to_letter(self.sheetAddCols[sheetindex] + 1+jlen)}2")

        dalen = self.fund_data['input1'] - self.fund_data['input3']
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            start_col = self.fund_data['input3']+self.sheetAddCols[sheetindex]-1
            if sheetindex in ['11', '12']:
                start_col += 1
        else:
            start_col = self.fund_data['input3']+self.sheetAddCols[sheetindex]-self.fund_data['input2']-1
        if dalen>0:
            xspreadsheet["rows"]["1"]["cells"][str(start_col)] = {"text": "达产期", "merge": [0, dalen], "style": 1}
            xspreadsheet["merges"].append(f"{number_to_letter(start_col+1)}2:{colIndexLetter}2")


        toulen = self.fund_data['input3'] - self.fund_data['input2']-1
        if sheetindex not in ['2', '5', '6', '7', '8', '10','14']:
            start_col = self.fund_data['input2']+self.sheetAddCols[sheetindex]
            if sheetindex in ['11', '12']:
                start_col += 1
        else:
            start_col = self.sheetAddCols[sheetindex]
        if toulen > 1:
            xspreadsheet["rows"]["1"]["cells"][str(start_col)] = {"text": "投产期", "merge": [0, toulen-1], "style": 1}
            xspreadsheet["merges"].append(f"{number_to_letter(start_col+1)}2:{number_to_letter(start_col+toulen)}2")
        timenum_e = self.fund_data['input1'] + 1
        if sheetindex in ['2','5','6','7','8','10','14']:
            timenum_s = self.fund_data['input2']+1
        elif sheetindex in ['11','12']:
            timenum_s = 0
        else:
            timenum_s = 1
        temp = 0
        for loop in range(timenum_s,timenum_e):
            xspreadsheet["rows"]["2"]["cells"][str(temp+self.sheetAddCols[sheetindex])]={"text": str(loop), "style": 1}
            temp += 1
            for rowloop in range(3,actual_len):
                xspreadsheet["rows"][str(rowloop)]["cells"][str(loop-1-self.fund_data['input2']+self.sheetAddCols[sheetindex])]={"style": 1}
        xspreadsheet["cols"]["len"] = colIndex
        xspreadsheet["rows"]["0"]["cells"]["0"]["merge"] = [0,col_index-1]
        xspreadsheet["merges"].append(f"A1:{number_to_letter(col_index)}1")
        return xspreadsheet


    def toCalExcel(self):
        # 根据数据建立cal_sheetdata的表格，并将 4
        workbook = Workbook()
        # 保存新工作簿
        uid = "0000"
        current_time = datetime.datetime.now()
        create_time = current_time.strftime("%Y_%m_%d_%H_%M_%S")
        try:
            projectname = self.fund_data['projectname']
        except KeyError:
            projectname = ""
        workbookname = "cal_" + projectname + "_" + uid + "_" + create_time
        workbook.save(f'{workbookname}.xlsx')
        worksheet = workbook.active
        index = 1

        # 将fund_data中的数据写入基础数据中

        # 基础数据1的写入
        for loop in ['projectname','projectplace','projectind','input1','input2']:
            worksheet[f'A{index}'] = index
            worksheet[f'B{index}'] = loop
            try:
                worksheet[f'C{index}'] = self.fund_data[loop] if self.fund_data[loop] != '' else "缺失"
            except KeyError:
                worksheet[f'C{index}'] = "缺失"
            index += 1

        worksheet[f'A{index}'] = index
        worksheet[f'B{index}'] = "productionloads"
        if self.fund_data['input2'] != '' and self.fund_data['input1'] != '' and self.fund_data['input1'] - \
                self.fund_data['input2'] > 0:
            for loop in range(self.fund_data['input1']-self.fund_data['input2']):
                try:
                    worksheet[f'{get_column_letter(loop+3)}{index}'] = float(self.fund_data[f'productionLoad{loop+self.fund_data["input2"]+1}']) / 100 if self.fund_data[f'productionLoad{loop+self.fund_data["input2"]+1}'] != '' else "缺失"
                except KeyError:
                    worksheet[f'{get_column_letter(loop+3)}{index}'] = "缺失"
        index += 1

        percent = ['benchmarkyield','install_rate','trans_rate','pre_rate','preup_rate']
        for loop in ['benchmarkyield','benchmark_s_paybackperiod','benchmark_d_paybackperiod',
                     'install_rate','trans_rate','pre_rate','preup_rate']:
            worksheet[f'A{index}'] = index
            worksheet[f'B{index}'] = loop
            try:
                if loop in percent:
                    worksheet[f'C{index}'] = float(self.fund_data[loop]) / 100 if self.fund_data[loop] != '' else "缺失"
                else:
                    worksheet[f'C{index}'] = self.fund_data[loop] if self.fund_data[loop] != '' else "缺失"
            except KeyError:
                worksheet[f'C{index}'] = "缺失"
            index += 1

        # 添加工程项目数据
        index += 7  # 避开建设投资年度投资占比和建设投资资金筹措计划数据

        worksheet[f'E{index}'] = f'工程费用基础数据'
        for i in range(int(self.fund_data['projectcostCounter'])):
            index += 1
            worksheet.insert_rows(index)
            # 新建项目赋值
            worksheet[f'E{index}'] = f'{i + 1}'
            worksheet[f'F{index}'] = self.fund_data[f'additionalCostName_{i + 1}']
            worksheet[f'G{index}'] = f'=J{index}*K{index} / 10000'
            worksheet[f'H{index}'] = float(self.fund_data[f'additionalCostEquipment_{i + 1}'])
            worksheet[f'I{index}'] = f"=H{index}*C10*(1-C11)"
            worksheet[f'J{index}'] = self.fund_data[f'additionalCostAmount_{i + 1}']
            worksheet[f'K{index}'] = self.fund_data[f'additionalCostPrice_{i + 1}']
        index += 1  # 切换到其他费用基础数据的位置
        worksheet[f'E{index}'] = f'其它费用基础数据'
        for i in range(int(self.fund_data['othercostCounter'])):
            index += 1
            worksheet.insert_rows(index)
            # 新建项目赋值
            worksheet[f'E{index}'] = f'{i + 1}'
            # print(self.fund_data)
            worksheet[f'F{index}'] = self.fund_data[f'additionalotherCostName_{i + 1}']
            worksheet[f'G{index}'] = float(self.fund_data[f'additionalOtherCost_{i + 1}'])
        index += 1
        index = index - 7 - 2 - int(self.fund_data['projectcostCounter']) - int(self.fund_data['othercostCounter'])

        worksheet[f'A{index}'] = f'建设投资年度投资占比'
        inv_plan = ast.literal_eval(self.fund_data['investmentPlan'])
        loan_plan = ast.literal_eval(self.fund_data['LoanPlan'])
        inv_colindex = 'B'
        index += 1
        for year, values in inv_plan.items():
            worksheet[f'{inv_colindex}{index}'] = year
            worksheet[f'{inv_colindex}{index+1}'] = float(values['percentage']) / 100
            inv_colindex = get_column_letter(column_index_from_string(inv_colindex) + 1)

        worksheet[f'{inv_colindex}{index}'] = len(inv_plan.items()) + 1
        worksheet[f'{inv_colindex}{index+1}'] = float(self.fund_data['leftInvestmentPercentage']) / 100
        index += 2
        worksheet[f'A{index}'] = f'建设投资资金筹措计划（万元）'
        inv_colindex = 'B'
        index += 1
        for year, values in loan_plan.items():
            worksheet[f'{inv_colindex}{index}'] = year
            worksheet[f'{inv_colindex}{index+1}'] = float(values['percentage']) / 100
            worksheet[f'{inv_colindex}{index+2}'] = f"=1-{inv_colindex}{index+1}"
            inv_colindex = get_column_letter(column_index_from_string(inv_colindex) + 1)

        index += 3

        # 表2
        worksheet[f'A{index}'] = f'附表2流动资金估算'
        index += 1
        percent += ['days8','days9']
        for loop in ['days1','days2','days3','days4','days5','days6','days7','days8','days9']:
            worksheet[f'A{index}'] = index
            worksheet[f'B{index}'] = loop
            try:
                if loop in percent:
                    worksheet[f'C{index}'] = float(self.fund_data[loop]) / 100 if self.fund_data[loop] != '' else "缺失"
                else:
                    worksheet[f'C{index}'] = self.fund_data[loop] if self.fund_data[loop] != '' else "缺失"
            except KeyError:
                worksheet[f'C{index}'] = "缺失"
            index += 1
        index += 1

        worksheet[f'A{index}'] = f'附表3：借款还本付息表基础数据'
        index += 1
        # 表3
        worksheet[f'B{index}'] = '长期借款利率'
        try:
            worksheet[f'C{index}'] = float(self.fund_data['loadrate']) / 100 if self.fund_data['loadrate'] != '' else "缺失"
        except KeyError:
            worksheet[f'C{index}'] = "缺失"
        index += 1
        worksheet[f'B{index}'] = '还款方式'
        try:
            if self.fund_data['repayMethod'] == 'ave_capital':
                worksheet[f'C{index}'] = '等额本金'
            else:
                worksheet[f'C{index}'] = '等额本息'
        except KeyError:
            worksheet[f'C{index}'] = "缺失"
        index += 1
        worksheet[f'B{index}'] = "还款年限"
        try:
            worksheet[f'C{index}'] = int(self.fund_data['payTime']) if self.fund_data['payTime'] != '' else "缺失"
        except KeyError:
            worksheet[f'C{index}'] = "缺失"
        # 表5(缺少折旧方式)
        diction = { 'depreciationMethod':'str','dep_year2':'int','dep_year1':'int','dep_year3':'int','res_rate2':'percent','res_rate1':'percent',
         'amo_year1':'int','amo_year2':'int','production1':'float','production2':'float','production3':'float',
         'production4': 'float','production5':'int','production6':'float','production7':'percent','production8':'percent',
         'production9': 'float','production10':'float','production11':'percent','production12':'percent',
         'production13': 'percent','production14':'percent','production15':'percent','production16':'percent'
         }
        index += 1
        for name,type in diction.items():
            worksheet[f'B{index}'] = name
            worksheet[f'C{index}'] = self.getData(name,type)
            index += 1

        # 此时index是60，默认空两行开始表1，则为62，如果项目添加的非常多
        index += max(2,22+int(self.fund_data['projectcostCounter']) + int(self.fund_data['othercostCounter'])-59)
        startrows = {}
        startrows[self.sheetIndexs['1']]=index
        # 表格第一行
        worksheet[f'A{index}'] = '附表1 : 建设投资估算表'
        worksheet[f'H{index}'] = '单位：万元'
        index += 1
        worksheet[f'A{index}'] = '序号'
        worksheet[f'B{index}'] = '工程或费用名称'
        worksheet[f'C{index}'] = '建筑工程'
        worksheet[f'D{index}'] = '设备工程'
        worksheet[f'E{index}'] = '安装工程'
        worksheet[f'F{index}'] = '其他费用'
        worksheet[f'G{index}'] = '合计'
        worksheet[f'H{index}'] = '占总值比'
        index += 1
        # 表格第一列
        temp = 0
        for loop in [('1','第一部分：工程费用'),('2','第二部分：其他费用'),('','第一、二部分费用合计'),
                     ('3','预备费用'),('3.1','基本预备费'),('3.2','涨价预备费'),
                     ('4','建设投资合计（1+2+3）')]:
            worksheet[f'A{index+temp}'] = loop[0]
            worksheet[f'B{index + temp}'] = loop[1]
            temp+=1



        # 遍历工作表中的所有行并设定所有对齐
        for row in worksheet:
            # 遍历每一行中的所有单元格
            for cell in row:
                # 清除单元格的值
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 合并标题
        mergeexcel(f'A{index - 2}', f'G{index - 2}', worksheet)

        # 插入工程费用栏目
        total_index = 6 + int(self.fund_data['projectcostCounter']) + int(self.fund_data['othercostCounter']) + index
        self.invenstment['total_index'] = f'$G${total_index}'
        constrindex = index
        index += 1

        for i in range(int(self.fund_data['projectcostCounter'])):
            worksheet.insert_rows(index+i)
            # 新建项目赋值
            worksheet[f'A{index+i}'] = f'1.{i + 1}'
            worksheet[f'B{index+i}'] = f"=F{i + 22}"
            worksheet[f'C{index+i}'] = f"=G{i + 22}"
            worksheet[f'D{index+i}'] = f"=H{i + 22}"
            worksheet[f'E{index+i}'] = f"=I{i + 22}"
            worksheet[f'G{index+i}'] = f"=C{index+i}+D{index+i}+E{index+i}"
            worksheet[f'H{index+i}'] = f"=G{index+i}/G{total_index}"
            worksheet[f'H{index+i}'].number_format = "0.0%"

        # 更新阶段求和
        if self.fund_data['projectcostCounter'] != "0":
            worksheet[f'C{index - 1}'] = f"=SUM(C{index}:C{index-1+int(self.fund_data['projectcostCounter'])})"
            worksheet[f'D{index - 1}'] = f"=SUM(D{index}:D{index-1+int(self.fund_data['projectcostCounter'])})"
            worksheet[f'E{index - 1}'] = f"=SUM(E{index}:E{index-1+int(self.fund_data['projectcostCounter'])})"
            worksheet[f'G{index - 1}'] = f"=SUM(G{index}:G{index-1+int(self.fund_data['projectcostCounter'])})"
            worksheet[f'H{index - 1}'] = f"=G{index - 1}/G{total_index}"
            worksheet[f'H{index - 1}'].number_format = "0.0%"

        # 插入其它费用栏目
        index += int(self.fund_data['projectcostCounter'])  # 略过其他费用合计一行
        otherindex = index  # 记录下其他费用合计的行数

        for i in range(int(self.fund_data['othercostCounter'])):
            index += 1
            worksheet.insert_rows(index)
            worksheet[f'A{index}'] = f'2.{i + 1}'
            worksheet[f'B{index}'] = f"=F{i + 1 + 22 + int(self.fund_data['projectcostCounter'])}"
            worksheet[f'F{index}'] = f"=G{i + 1 + 22 + int(self.fund_data['projectcostCounter'])}"
            worksheet[f'G{index}'] = f"=F{index}"
            worksheet[f'H{index}'] = f"=G{index}/G{total_index}"
            worksheet[f'H{index}'].number_format = "0.0%"

        # 更新阶段求和
        if self.fund_data['othercostCounter'] != "0":
            worksheet[f'F{otherindex}'] = f"=SUM(F{otherindex + 1}:F{index})"
            worksheet[f'G{otherindex}'] = f"=SUM(G{otherindex + 1}:G{index})"
            worksheet[f'H{otherindex}'] = f"=G{otherindex}/G{total_index}"
            worksheet[f'H{otherindex}'].number_format = "0.0%"

        # 更新一二部分求和

        index += 1
        worksheet[f'G{index}'] = f"=G{constrindex}+G{otherindex}"

        # 更新预备费用
        preservedindex = index + 1
        self.invenstment['preservedindex'] = f'$G${preservedindex}'
        worksheet[f'F{preservedindex + 1}'] = f"=G{index}* C12"  # 基本预备费
        worksheet[f'F{preservedindex + 2}'] = f"=G{constrindex}*C13"  # 涨价预备费
        worksheet[f'F{preservedindex}'] = f"=F{preservedindex + 1}+F{preservedindex + 2}"
        worksheet[f'G{preservedindex + 1}'] = f"=F{preservedindex + 1}"  # 基本预备费总值
        worksheet[f'G{preservedindex + 2}'] = f"=F{preservedindex + 2}"  # 涨价预备费总值
        worksheet[f'G{preservedindex}'] = f"=G{preservedindex + 1}+G{preservedindex + 2}"
        worksheet[f'H{preservedindex + 1}'] = f"=G{preservedindex + 1}/G{total_index}"
        worksheet[f'H{preservedindex + 2}'] = f"=G{preservedindex + 2}/G{total_index}"
        worksheet[f'H{preservedindex}'] = f"=G{preservedindex}/G{total_index}"
        worksheet[f'H{preservedindex}'].number_format = "0.0%"
        worksheet[f'H{preservedindex + 1}'].number_format = "0.0%"
        worksheet[f'H{preservedindex + 2}'].number_format = "0.0%"

        # 更新合计费用
        worksheet[f'G{total_index}'] = f"=G{constrindex}+G{otherindex}+G{preservedindex}"
        worksheet[f'H{total_index}'] = f"=G{total_index}/G{total_index}"
        worksheet[f'H{total_index}'].number_format = "0.0%"

        index = total_index + 2

        # # 修改显示
        # cell_range = worksheet['C3':'G16']
        # for each_cell in cell_range:
        #     for each in each_cell:
        #         each.number_format = "0"

        # 采用makesheet方法建表2-14
        for iloop in range(2, 15):
            ind = str(iloop)
            startrows[self.sheetIndexs[ind]] = index
            index = self.makesheet(ind, workbook,start_row=index) + 15
        for iloop in range(2, 15):
            ind = str(iloop)
            self.fillsheet_cal(ind, workbook,startrows)
        startrows.pop('保留')
        startrows[self.sheetIndexs['15']] = index-1
        # 敏感性分析的公式
        worksheet[f'B{index}']= "变动系数"
        worksheet[f'C{index}'] = "NPV"
        worksheet[f'D{index}'] = "变动率"
        Count = 0
        for para in [1,1.2,1.1,0.9,0.8,1,1.2,1.1,0.9,0.8,1,1.2,1.1,0.9,0.8]:
            Count += 1
            index+= 1
            worksheet[f'B{index}'] = para
            worksheet[f'D{index}'] = para - 1
            worksheet[f'C{index}'] = f"=NPV($C$7,F{index}:{get_column_letter(self.fund_data['input1'] + 6)}{index})"
            for letter_index in range(self.fund_data['input1']+1):
                letter = get_column_letter(6+letter_index)
                if Count<=5:
                    if letter_index < self.fund_data['input2']:
                        formu = f"={get_column_letter(4+letter_index)}{startrows[self.sheetIndexs['12']]+15}*B{index}"
                    else:
                        formu = f"={get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+15}"
                elif Count<=10:
                    if letter_index <= self.fund_data['input2']:
                        formu = f"={get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+15}"
                    else:
                        formu = f"={get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+4}- {get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+8}-{get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+12}*D{index}"
                else:
                    if letter_index < self.fund_data['input2']:
                        formu = f"={get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+15}"
                    else:
                        formu = f"={get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+15}+{get_column_letter(4 + letter_index)}{startrows[self.sheetIndexs['12']]+5}*D{index}"
                worksheet[f'{letter}{index}'] = formu
        index += 15
        startrows[self.sheetIndexs['16']] = index-3
        worksheet[f'D{index}'] = f"=D{index+1}+D{index+6}+D{index+7}"
        index += 1
        worksheet[f'D{index}'] = f"=D{index+1}+D{index+2}+D{index+3}+D{index+4}"
        index += 1
        worksheet[f'D{index}'] = f"=C64"
        index += 1
        worksheet[f'D{index}'] = f"=D64+E64"
        index += 1
        worksheet[f'D{index}'] = f"=F71"
        index += 1
        worksheet[f'D{index}'] = f"=F67"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['4']]+5}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['4']]+6}"
        index += 1
        worksheet[f'D{index}'] = f"=D{index+1}+D{index+2}+D{index+3}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['4']]+8}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['4']]+12}+C{startrows[self.sheetIndexs['4']]+13}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['4']]+14}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['8']]+4}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['8']]+5}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['8']]+6}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['8']]+7}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['8']]+8}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['7']]+12}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['7']]+13}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['7']]+14}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['7']]+15}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['10']]+7}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['10']]+8}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['10']]+9}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['10']]+11}"
        index += 1
        worksheet[f'D{index}'] = f"=C{startrows[self.sheetIndexs['10']]+12}"
        index += 2
        worksheet[f'D{index}'] = f"=D{index-6}/D{index-19}/(C4-C5)"
        index += 1
        worksheet[f'D{index}'] = f"=D{index-5}/D{index-20}/(C4-C5)"

        index += 3
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['11']]+24}"
        index += 1
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['11']]+25}"
        index += 1
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['11']]+22}"
        index += 1
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['11']]+23}"

        index += 2
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['11']]+24}"
        index += 1
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['11']]+25}"
        index += 1
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['11']]+22}"
        index += 1
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['11']]+23}"

        index += 3
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['12']]+24}"
        index += 1
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['12']]+25}"
        index += 1
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['12']]+22}"
        index += 1
        worksheet[f'D{index}'] = f"=E{startrows[self.sheetIndexs['12']]+23}"

        index += 2
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['12']]+24}"
        index += 1
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['12']]+25}"
        index += 1
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['12']]+22}"
        index += 1
        worksheet[f'D{index}'] = f"=F{startrows[self.sheetIndexs['12']]+23}"

        da_letter = get_column_letter(
            self.fund_data['input3'] + self.sheetAddCols['7'] - self.fund_data['input2'])

        index += 1
        worksheet[f'D{index}'] = f"={da_letter}{startrows[self.sheetIndexs['7']]+16}"
        index += 1
        worksheet[f'D{index}'] = f"=C35"
        da_letter = get_column_letter(
            self.fund_data['input3'] + self.sheetAddCols['3'])
        index += 1
        worksheet[f'D{index}'] = f"={da_letter}{startrows[self.sheetIndexs['3']]+15}"
        index += 1
        worksheet[f'D{index}'] = f"={da_letter}{startrows[self.sheetIndexs['3']]+16}"

        # 保存以供读写
        workbook.save(f'{workbookname}.xlsx')

        xspreadsheetmap = {}
        xspreadsheetmap[1] = '{"name":"Sheet","freeze":"A1","styles":[{"align":"center"},{"align":"center","border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}}],"merges":["A1:G1"],"rows":{"0":{"cells":{"0":{"text":"附表1:建设投资估算表","merge":[0,6],"style":1},"1":{"style":0},"2":{"style":0},"3":{"style":0},"4":{"style":0},"5":{"style":0},"6":{"style":0},"7":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","style":1},"1":{"text":"工程或费用名称","style":1},"2":{"text":"建筑工程","style":1},"3":{"text":"设备工程","style":1},"4":{"text":"安装工程","style":1},"5":{"text":"其他费用","style":1},"6":{"text":"合计","style":1},"7":{"text":"占总值比","style":1}}},"2":{"cells":{"0":{"text":"1","style":1},"1":{"text":"第一部分：工程费用","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1}}},'
        # 需要继续添加第一部分的具体工程还有第二部分具体工程以及剩下来的固定内容
        for iloop in range(int(self.fund_data['projectcostCounter'])):
            xspreadsheetmap[1] += f'"{iloop+3}":{"{"}"cells":{{"0":{{"text": "1.{iloop+1}","style": 1}},"1":{{"style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xspreadsheetmap[1] += f'"{ 3 + int(self.fund_data["projectcostCounter"])}":{{"cells":{{"0":{{"text": "2","style": 1}},"1":{{"text": "第二部分：其他费用","style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        for iloop in range(int(self.fund_data['othercostCounter'])):
            xspreadsheetmap[1] += f'"{iloop + 4+int(self.fund_data["projectcostCounter"])}":{{"cells":{{"0":{{"text": "2.{iloop + 1}","style": 1}},"1":{{"style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xindex = 4+int(self.fund_data['projectcostCounter'])+int(self.fund_data['othercostCounter'])
        xspreadsheetmap[1] += f'"{xindex}":{{"cells":{{"0":{{"style": 1}},"1":{{"text": "第一、二部分费用合计","style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xindex += 1
        xspreadsheetmap[
            1] += f'"{xindex}":{{"cells":{{"0":{{"text": "3","style": 1}},"1":{{"text": "预备费用","style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xindex += 1
        xspreadsheetmap[
            1] += f'"{xindex}":{{"cells":{{"0":{{"text": "3.1","style": 1}},"1":{{"text": "基本预备费","style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xindex += 1
        xspreadsheetmap[
            1] += f'"{xindex}":{{"cells":{{"0":{{"text": "3.2","style": 1}},"1":{{"text": "涨价预备费","style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xindex += 1
        xspreadsheetmap[
            1] += f'"{xindex}":{{"cells":{{"0":{{"text": "4","style": 1}},"1":{{"text": "建设投资合计（1+2+3）","style": 1}},"2":{{"style": 1}},"3":{{"style": 1}},"4":{{"style": 1}},"5":{{"style": 1}},"6":{{"style": 1}},"7":{{"style": 1}}}}{"}"},'
        xspreadsheetmap[
            1] += f'"len":{xindex+1}{"}"},"cols":{{"1":{{"width":194}},"len":8}},"validations":[],"autofilter":{{}}{"}"}'

        xspreadsheetmap[1] = json.loads(xspreadsheetmap[1])
        xspreadsheetmap[2] = {"name":"Sheet","freeze":"A1","styles":[{"align":"center"},{"align":"center","border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}}],"merges":["B2:B3","D2:D3","A2:A3","C2:C3"],"rows":{"0":{"cells":{"0":{"text":"附表2：流动资金估算","merge":[],"style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"最低周转天数","merge":[1,0],"style":1},"3":{"text":"最低周转次数","merge":[1,0],"style":1},"4":{"text":"投产期","style":1}}},"2":{"cells":{"0":{"style":0},"1":{"style":1},"2":{"style":0},"3":{"style":1},"4":{"style":1}}},"3":{"cells":{"0":{"text":"1","style":1},"1":{"text":"流动资产","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"4":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"应收帐款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"5":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"存货","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"6":{"cells":{"0":{"text":"1.2.1","style":1},"1":{"text":"原材料","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"7":{"cells":{"0":{"text":"1.2.2","style":1},"1":{"text":"燃料动力","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"8":{"cells":{"0":{"text":"1.2.3","style":1},"1":{"text":"在产品","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"9":{"cells":{"0":{"text":"1.2.4","style":1},"1":{"text":"产成品","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"10":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"现金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"11":{"cells":{"0":{"text":"2","style":1},"1":{"text":"流动负债","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"12":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"应付帐款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"13":{"cells":{"0":{"text":"3","style":1},"1":{"text":"流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"14":{"cells":{"0":{"text":"4","style":1},"1":{"text":"流动资金本年增加额","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"15":{"cells":{"0":{"text":"5","style":1},"1":{"text":"流动资金借款（本年）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"16":{"cells":{"0":{"text":"6","style":1},"1":{"text":"流动资金利息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"len":16},"cols":{"0":{"width":75},"1":{"width":150},"len":9},"validations":[],"autofilter":{}}
        xspreadsheetmap[2] = self.makexspreadsheet("2",xspreadsheetmap[2])
        xspreadsheetmap[3] = {"name":"Sheet","freeze":"A1","styles":[{"align":"center"},{"align":"center","border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}}],"merges":["A2:A3","B2:B3","D2:D3","C2:C3"],"rows":{"0":{"cells":{"0":{"text":"附表3：借款还本付息表","merge":[],"style":1},"1":{"style":0},"2":{"style":0},"3":{"style":0},"4":{"style":0},"5":{"style":0},"6":{"style":0},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"利率","merge":[1,0],"style":1},"3":{"text":"合计","merge":[1,0],"style":1},"4":{"text":"建设期","style":1},"5":{"style":0},"6":{"text":"投产期","style":1},"7":{"style":0},"8":{"text":"达产期","style":1},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"2":{"cells":{"0":{"style":0},"1":{"style":1},"2":{"style":0},"3":{"style":1},"4":{"text":"1","style":1},"5":{"text":"2","style":1},"6":{"text":"3","style":1},"7":{"text":"4","style":1},"8":{"text":"5","style":1},"9":{"text":"6","style":1},"10":{"text":"7","style":1},"11":{"text":"8","style":1}}},"3":{"cells":{"0":{"text":"1","style":1},"1":{"text":"借款及还本付息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"4":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"年初借款本息和","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"5":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"当期借款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"6":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"当期应计利息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"7":{"cells":{"0":{"text":"1.4","style":1},"1":{"text":"当期还本付息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"8":{"cells":{"0":{"text":"1.4.1","style":1},"1":{"text":"其中：还本","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"9":{"cells":{"0":{"text":"1.4.2","style":1},"1":{"text":"付息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"10":{"cells":{"0":{"text":"1.5","style":1},"1":{"text":"期末借款余额","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"11":{"cells":{"0":{"text":"2","style":1},"1":{"text":"偿还借款本金资金来源","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"12":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"利润","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"13":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"折旧","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"14":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"摊销","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"15":{"cells":{"0":{"text":"3","style":1},"1":{"text":"利息备付率","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"16":{"cells":{"0":{"text":"4","style":1},"1":{"text":"偿债备付率","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"len":17},"cols":{"0":{"width":72},"1":{"width":154},"len":12},"validations":[],"autofilter":{}}
        xspreadsheetmap[3] = self.makexspreadsheet("3", xspreadsheetmap[3])
        xspreadsheetmap[4] = {"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["B2:B3","A2:A3","C2:C3"],"rows":{"0":{"cells":{"0":{"text":"附表4：投资使用计划与资金筹措表","merge":[],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"style":2},"9":{"style":2},"10":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"建设期","merge":[0,1],"style":1},"4":{"style":2},"5":{"text":"投产期","merge":[0,1],"style":1},"6":{"style":2},"7":{"text":"达产期","merge":[0,3],"style":1},"8":{"style":2},"9":{"style":2},"10":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"text":"1","style":1},"4":{"text":"2","style":1},"5":{"text":"3","style":1},"6":{"text":"4","style":1},"7":{"text":"5","style":1},"8":{"text":"6","style":1},"9":{"text":"7","style":1},"10":{"text":"8","style":1}}},"3":{"cells":{"0":{"text":"1","style":1},"1":{"text":"资金总额","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"4":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"建设投资","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"5":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"建设期利息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"6":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"7":{"cells":{"0":{"text":"2","style":1},"1":{"text":"资金筹措","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"8":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"项目资本金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"9":{"cells":{"0":{"text":"2.1.1","style":1},"1":{"text":"用于建设投资","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"10":{"cells":{"0":{"text":"2.1.2","style":1},"1":{"text":"用于流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"11":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"债务资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"12":{"cells":{"0":{"text":"2.2.1","style":1},"1":{"text":"用于建设投资","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"13":{"cells":{"0":{"text":"2.2.2","style":1},"1":{"text":"用于建设期利息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"14":{"cells":{"0":{"text":"2.2.3","style":1},"1":{"text":"用于流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"len":15},"cols":{"len":11},"validations":[],"autofilter":{}}
        xspreadsheetmap[4] = self.makexspreadsheet("4", xspreadsheetmap[4])
        xspreadsheetmap[5] = {"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"},{"align":"left"}],"merges":["D2:D3","C2:C3","E2:E3","B2:B3","A2:A3","B23:D23","B24:D24"],"rows":{"0":{"cells":{"0":{"text":"附表5：固定资产折旧估算表","merge":[],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"style":2},"9":{"style":2},"10":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"原值","merge":[1,0],"style":1},"3":{"text":"折旧年限","merge":[1,0],"style":1},"4":{"text":"残值率","merge":[1,0],"style":1},"5":{"text":"投产期","merge":[0,1],"style":1},"6":{"style":2},"7":{"text":"达产期","merge":[0,3],"style":1},"8":{"style":2},"9":{"style":2},"10":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"style":1},"4":{"style":2},"5":{"text":"3","style":1},"6":{"text":"4","style":1},"7":{"text":"5","style":1},"8":{"text":"6","style":1},"9":{"text":"7","style":1},"10":{"text":"8","style":1}}},"3":{"cells":{"0":{"text":"1","style":1},"1":{"text":"建筑工程","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"4":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"原值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"5":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"折旧费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"6":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"7":{"cells":{"0":{"style":1},"1":{"style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"8":{"cells":{"0":{"text":"2","style":1},"1":{"text":"设备工程","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"9":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"原值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"10":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"折旧费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"11":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"12":{"cells":{"0":{"style":1},"1":{"style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"13":{"cells":{"0":{"text":"3","style":1},"1":{"text":"土地征用费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"14":{"cells":{"0":{"text":"3.1","style":1},"1":{"text":"原值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"15":{"cells":{"0":{"text":"3.2","style":1},"1":{"text":"折旧费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"16":{"cells":{"0":{"text":"3.3","style":1},"1":{"text":"净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"17":{"cells":{"0":{"style":1},"1":{"style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"18":{"cells":{"0":{"text":"4","style":1},"1":{"text":"固定资产合计","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"19":{"cells":{"0":{"text":"4.1","style":1},"1":{"text":"原值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"20":{"cells":{"0":{"text":"4.2","style":1},"1":{"text":"折旧费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"21":{"cells":{"0":{"text":"4.3","style":1},"1":{"text":"净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"22":{"cells":{"0":{"style":2},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"style":2},"9":{"style":2},"10":{"style":2}}},"23":{"cells":{"0":{"text":"备注","style":2},"1":{"text":"建筑工程包含预备费和建设期利息","style":2,"merge":[0,2]},"10":{"style":2}}},"24":{"cells":{"0":{"style":2},"1":{"text":"设备工程包含安装工程费","style":2,"merge":[0,2]},"10":{"style":2}}},"len":25},"cols":{"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[5] = self.makexspreadsheet("5", xspreadsheetmap[5],21)
        xspreadsheetmap[6] ={"name":"Sheet","freeze":"A1","styles":[{"align":"center"},{"align":"center","border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}}],"merges":["B2:B3","D2:D3","A2:A3","C2:C3"],"rows":{"0":{"cells":{"0":{"text":"附表6：无形资产及其他资产摊销表","merge":[0,8],"style":1},"1":{"style":0},"2":{"style":0},"3":{"style":0},"4":{"style":0},"5":{"style":0},"6":{"style":0},"7":{"style":0},"8":{"style":0},"9":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"原值","merge":[1,0],"style":1},"3":{"text":"摊销年限","merge":[1,0],"style":1},"4":{"text":"投产期","merge":[0,1],"style":1},"5":{"style":0},"6":{"text":"达产期","merge":[0,3],"style":1},"7":{"style":0},"8":{"style":0},"9":{"style":0}}},"2":{"cells":{"0":{"style":0},"1":{"style":1},"2":{"style":0},"3":{"style":1},"4":{"text":"3","style":1},"5":{"text":"4","style":1},"6":{"text":"5","style":1},"7":{"text":"6","style":1},"8":{"text":"7","style":1},"9":{"text":"8","style":1}}},"3":{"cells":{"0":{"text":"1","style":1},"1":{"text":"无形资产","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"4":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"摊销费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"5":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"6":{"cells":{"0":{"text":"2","style":1},"1":{"text":"其他资产","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"7":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"摊销费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"8":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"9":{"cells":{"0":{"text":"3","style":1},"1":{"text":"无形资产及其他资产合计","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"10":{"cells":{"0":{"text":"3.1","style":1},"1":{"text":"摊销合计","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"11":{"cells":{"0":{"text":"3.2","style":1},"1":{"text":"净值合计","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"12":{"cells":{}},"len":12},"cols":{"0":{"width":60},"1":{"width":157},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[6] = self.makexspreadsheet("6", xspreadsheetmap[6])
        xspreadsheetmap[7] ={"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["C2:C3","A2:A3","B2:B3"],"rows":{"0":{"cells":{"0":{"text":"附表7：总成本费用估算表","merge":[0,7],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"投产期","merge":[0,1],"style":1},"4":{"style":2},"5":{"text":"达产期","merge":[0,3],"style":1},"6":{"style":2},"7":{"style":2},"8":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"text":"3","style":1},"4":{"text":"4","style":1},"5":{"text":"5","style":1},"6":{"text":"6","style":1},"7":{"text":"7","style":1},"8":{"text":"8","style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"外购原材料","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"5":{"cells":{"0":{"text":"2","style":1},"1":{"text":"外购燃料动力","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"6":{"cells":{"0":{"text":"3","style":1},"1":{"text":"工资及福利费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"7":{"cells":{"0":{"text":"4","style":1},"1":{"text":"修理费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"8":{"cells":{"0":{"text":"5","style":1},"1":{"text":"折旧费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"9":{"cells":{"0":{"text":"6","style":1},"1":{"text":"摊销费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"10":{"cells":{"0":{"text":"7","style":1},"1":{"text":"利息支出","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"11":{"cells":{"0":{"text":"8","style":1},"1":{"text":"其他费用","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"12":{"cells":{"0":{"text":"9","style":1},"1":{"text":"总成本费用","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"13":{"cells":{"0":{"text":"9.1","style":1},"1":{"text":"固定成本","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"14":{"cells":{"0":{"text":"9.2","style":1},"1":{"text":"可变成本","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"15":{"cells":{"0":{"text":"9.3","style":1},"1":{"text":"经营成本","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"16":{"cells":{"0":{"text":"10","style":1},"1":{"text":"盈亏平衡点（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"17":{"cells":{}},"len":17},"cols":{"0":{"width":70},"1":{"width":112},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[7] = self.makexspreadsheet("7", xspreadsheetmap[7])
        xspreadsheetmap[8] ={"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["A2:A3","C2:C3","B2:B3"],"rows":{"0":{"cells":{"0":{"text":"附表8：营业收入、营业税金及附加表","merge":[0,7],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"投产期","merge":[0,1],"style":1},"4":{"style":2},"5":{"text":"达产期","merge":[0,3],"style":1},"6":{"style":2},"7":{"style":2},"8":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"text":"3","style":1},"4":{"text":"4","style":1},"5":{"text":"5","style":1},"6":{"text":"6","style":1},"7":{"text":"7","style":1},"8":{"text":"8","style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"营业收入","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"5":{"cells":{"0":{"text":"2","style":1},"1":{"text":"营业税金及附加","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"6":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"增值税（17%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"7":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"城市维护建设税（4%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"8":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"教育费附加（2%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"len":9},"cols":{"0":{"width":68},"1":{"width":143},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[8] = self.makexspreadsheet("8", xspreadsheetmap[8])
        xspreadsheetmap[9] = {"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["A2:A3","B2:B3","C2:C3"],"rows":{"0":{"cells":{"0":{"text":"附表9：资金来源与运用表","merge":[0,7],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"建设期","merge":[0,1],"style":1},"4":{"style":2},"5":{"text":"投产期","merge":[0,1],"style":1},"6":{"style":2},"7":{"text":"达产期","merge":[0,3],"style":1},"8":{"style":2},"9":{"style":2},"10":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"text":"1","style":1},"4":{"text":"2","style":1},"5":{"text":"3","style":1},"6":{"text":"4","style":1},"7":{"text":"5","style":1},"8":{"text":"6","style":1},"9":{"text":"7","style":1},"10":{"text":"8","style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"资金来源","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"5":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"利润总额","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"6":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"折旧费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"7":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"摊销费","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"8":{"cells":{"0":{"text":"1.4","style":1},"1":{"text":"长期借款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"9":{"cells":{"0":{"text":"1.5","style":1},"1":{"text":"流动资金借款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"10":{"cells":{"0":{"text":"1.6","style":1},"1":{"text":"自有资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"11":{"cells":{"0":{"text":"1.7","style":1},"1":{"text":"回收固定资产余值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"12":{"cells":{"0":{"text":"1.8","style":1},"1":{"text":"回收流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"13":{"cells":{"0":{"text":"2","style":1},"1":{"text":"资金应用","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"14":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"建设投资","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"15":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"建设期利息","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"16":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"17":{"cells":{"0":{"text":"2.4","style":1},"1":{"text":"所得税","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"18":{"cells":{"0":{"text":"2.6","style":1},"1":{"text":"长期借款本金偿还","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"19":{"cells":{"0":{"text":"3","style":1},"1":{"text":"盈余资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"20":{"cells":{"0":{"text":"4","style":1},"1":{"text":"累计盈余资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1}}},"len":21},"cols":{"0":{"width":66},"1":{"width":130},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[9] = self.makexspreadsheet("9", xspreadsheetmap[9])
        xspreadsheetmap[10] ={"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["B2:B3","C2:C3","A2:A3"],"rows":{"0":{"cells":{"0":{"text":"附表10：损益表","merge":[0,7],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"投产期","merge":[0,1],"style":1},"4":{"style":2},"5":{"text":"达产期","merge":[0,3],"style":1},"6":{"style":2},"7":{"style":2},"8":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"text":"3","style":1},"4":{"text":"4","style":1},"5":{"text":"5","style":1},"6":{"text":"6","style":1},"7":{"text":"7","style":1},"8":{"text":"8","style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"营业收入","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"5":{"cells":{"0":{"text":"2","style":1},"1":{"text":"营业税金及附加","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"6":{"cells":{"0":{"text":"3","style":1},"1":{"text":"总成本费用","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"7":{"cells":{"0":{"text":"4","style":1},"1":{"text":"利润总额（1-2-3）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"8":{"cells":{"0":{"text":"5","style":1},"1":{"text":"所得税","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"9":{"cells":{"0":{"text":"6","style":1},"1":{"text":"税后利润（4-5）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"10":{"cells":{"0":{"text":"7","style":1},"1":{"text":"可供分配利润","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"11":{"cells":{"0":{"text":"7.1","style":1},"1":{"text":"盈余公积金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"12":{"cells":{"0":{"text":"7.2","style":1},"1":{"text":"盈余公益金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"13":{"cells":{"0":{"text":"7.3","style":1},"1":{"text":"未分配利润(7-7.1-7.2)","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"14":{"cells":{"0":{"text":"8","style":1},"1":{"text":"累计未分配利润","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1}}},"len":15},"cols":{"0":{"width":68},"1":{"width":138},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[10] = self.makexspreadsheet("10", xspreadsheetmap[10])
        xspreadsheetmap[11] ={"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["B2:B3","C2:C3","A2:A3"],"rows":{"0":{"cells":{"0":{"text":"附表11-1：现金流量表（全部投资)","merge":[0,10],"style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"建设期","merge":[0,2],"style":1}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"style":2},"3":{"text":"0","style":1},"4":{"text":"1","style":1},"5":{"text":"2","style":1},"6":{"text":"3","style":1},"7":{"text":"4","style":1},"8":{"text":"5","style":1},"9":{"text":"6","style":1},"10":{"text":"7","style":1},"11":{"text":"8","style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"现金流入","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"5":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"营业收入","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"6":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"回收固定资产余值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"7":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"回收流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"8":{"cells":{"0":{"text":"2","style":1},"1":{"text":"现金流出","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"9":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"建设投资","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"10":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"11":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"经营成本","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"12":{"cells":{"0":{"text":"2.4","style":1},"1":{"text":"营业税金及附加","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"13":{"cells":{"0":{"text":"2.5","style":1},"1":{"text":"所得税","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"14":{"cells":{"0":{"text":"3","style":1},"1":{"text":"净现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"15":{"cells":{"0":{"text":"4","style":1},"1":{"text":"累计现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"16":{"cells":{"0":{"text":"5","style":1},"1":{"text":"所得税前净现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"17":{"cells":{"0":{"text":"6","style":1},"1":{"text":"所得税前累计净现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"18":{"cells":{}},"19":{"cells":{}},"20":{"cells":{}},"21":{"cells":{"3":{"text":"指标计算","style":1},"4":{"text":"税前","style":1},"5":{"text":"税后","style":1},"6":{"style":1,"text":"备注"}}},"22":{"cells":{"3":{"text":"静态回收期（年）","style":1},"4":{"style":1},"5":{"style":1},"6":{"text":"（从建设期算起）","style":1}}},"23":{"cells":{"3":{"text":"动态回收期（年）","style":1},"4":{"style":1},"5":{"style":1},"6":{"text":"（从建设期算起）","style":1}}},"24":{"cells":{"3":{"text":"内部收益率","style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1}}},"25":{"cells":{"3":{"text":"净现值（万元）","style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1}}},"26":{"cells":{}},"27":{"cells":{}},"28":{"cells":{}},"29":{"cells":{}},"30":{"cells":{}},"31":{"cells":{}},"len":26},"cols":{"0":{"width":60},"1":{"width":149},"3":{"width":120},"6":{"width":120},"len":12},"validations":[],"autofilter":{}}
        xspreadsheetmap[11] = self.makexspreadsheet("11", xspreadsheetmap[11],18)
        xspreadsheetmap[12] = {"name":"Sheet","freeze":"A1","styles":[{"align":"center"},{"align":"center","border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}}],"merges":["B2:B3","D2:F2","A2:A3","C2:C3"],"rows":{"0":{"cells":{"0":{"text":"附表11-2：现金流量表（自有资金)","merge":[0,10],"style":1},"1":{"style":0},"2":{"style":0},"3":{"style":0},"4":{"style":0},"5":{"style":0},"6":{"style":0},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"合计","merge":[1,0],"style":1},"3":{"text":"建设期","merge":[0,2],"style":1},"4":{"style":0},"5":{"style":0},"6":{"text":"投产期","merge":[0,1],"style":1},"7":{"style":0},"8":{"text":"达产期","merge":[0,3],"style":1},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"2":{"cells":{"0":{"style":0},"1":{"style":1},"2":{"style":0},"3":{"text":"0","style":1},"4":{"text":"1","style":1},"5":{"text":"2","style":1},"6":{"text":"3","style":1},"7":{"text":"4","style":1},"8":{"text":"5","style":1},"9":{"text":"6","style":1},"10":{"text":"7","style":1},"11":{"text":"8","style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"现金流入","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"5":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"营业收入","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"6":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"回收固定资产余值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"7":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"回收流动资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"8":{"cells":{"0":{"text":"2","style":1},"1":{"text":"现金流出","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"9":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"项目资本金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"10":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"借款本金偿还","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"11":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"借款利息支付","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"12":{"cells":{"0":{"text":"2.4","style":1},"1":{"text":"经营成本","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"13":{"cells":{"0":{"text":"2.5","style":1},"1":{"text":"营业税金及附加","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"14":{"cells":{"0":{"text":"2.6","style":1},"1":{"text":"所得税","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"15":{"cells":{"0":{"text":"3","style":1},"1":{"text":"净现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"16":{"cells":{"0":{"text":"4","style":1},"1":{"text":"累计现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"17":{"cells":{"0":{"text":"所得税前","style":1},"1":{"style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"18":{"cells":{"0":{"text":"7","style":1},"1":{"text":"净现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"19":{"cells":{"0":{"text":"8","style":1},"1":{"text":"累计现金流量","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1},"10":{"style":1},"11":{"style":1}}},"20":{"cells":{"0":{"style":0},"1":{"style":0},"2":{"style":0},"3":{"style":0},"4":{"style":0},"5":{"style":0},"6":{"style":0},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"21":{"cells":{"0":{"style":0},"1":{"style":0},"2":{"style":0},"3":{"text":"指标计算","style":1},"4":{"text":"税前","style":1},"5":{"text":"税后","style":1},"6":{"style":1},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"22":{"cells":{"0":{"style":0},"1":{"style":0},"2":{"style":0},"3":{"text":"静态回收期（年）","style":1},"4":{"style":1},"5":{"style":1},"6":{"text":"（从建设期算起）","style":1},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"23":{"cells":{"0":{"style":0},"1":{"style":0},"2":{"style":0},"3":{"text":"动态回收期（年）","style":1},"4":{"style":1},"5":{"style":1},"6":{"text":"（从建设期算起）","style":1},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"24":{"cells":{"0":{"style":0},"1":{"style":0},"2":{"style":0},"3":{"text":"内部收益率","style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"25":{"cells":{"0":{"style":0},"1":{"style":0},"2":{"style":0},"3":{"text":"净现值（万元）","style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":0},"8":{"style":0},"9":{"style":0},"10":{"style":0},"11":{"style":0}}},"26":{"cells":{}},"27":{"cells":{}},"28":{"cells":{}},"29":{"cells":{}},"30":{"cells":{}},"31":{"cells":{}},"32":{"cells":{}},"33":{"cells":{}},"len":26},"cols":{"0":{"width":69},"1":{"width":125},"3":{"width":120},"6":{"width":120},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[12] = self.makexspreadsheet("12", xspreadsheetmap[12],20)
        xspreadsheetmap[13] = {"name":"Sheet","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"},{"align":"center"}],"merges":["A2:A3","B2:B3","A24:A26"],"rows":{"0":{"cells":{"0":{"text":"附表12：资产负债表","merge":[0,8],"style":1},"1":{"style":2},"2":{"style":2},"3":{"style":2},"4":{"style":2},"5":{"style":2},"6":{"style":2},"7":{"style":2},"8":{"style":2},"9":{"text":"单位：万元","style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1},"2":{"text":"建设期","merge":[0,1],"style":1},"3":{"style":2},"4":{"text":"投产期","merge":[0,1],"style":1},"5":{"style":2},"6":{"text":"达产期","merge":[0,3],"style":1},"7":{"style":2},"8":{"style":2},"9":{"style":2}}},"2":{"cells":{"0":{"style":2},"1":{"style":1},"2":{"text":"1","style":1},"3":{"text":"2","style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"生产负荷（%）","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"资产","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"5":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"流动资产总额","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"6":{"cells":{"0":{"text":"1.1.1","style":1},"1":{"text":"应收帐款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"7":{"cells":{"0":{"text":"1.1.2","style":1},"1":{"text":"存货","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"8":{"cells":{"0":{"text":"1.1.3","style":1},"1":{"text":"现金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"9":{"cells":{"0":{"text":"1.1.4","style":1},"1":{"text":"累计盈余资金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"10":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"固定资产净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"11":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"无形及递延资产净值","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"12":{"cells":{"0":{"text":"1.4","style":1},"1":{"text":"在建工程","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"13":{"cells":{"0":{"text":"2","style":1},"1":{"text":"负债及所有者权益","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"14":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"流动负债总额","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"15":{"cells":{"0":{"text":"2.1.1","style":1},"1":{"text":"应付帐款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"16":{"cells":{"0":{"text":"2.1.2","style":1},"1":{"text":"流动资金借款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"17":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"长期借款","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"18":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"所有者权益","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"19":{"cells":{"0":{"text":"2.3.1","style":1},"1":{"text":"累计资本金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"20":{"cells":{"0":{"text":"2.3.2","style":1},"1":{"text":"累计盈余公积金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"21":{"cells":{"0":{"text":"2.3.3","style":1},"1":{"text":"累计盈余公益金","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"22":{"cells":{"0":{"text":"2.3.4","style":1},"1":{"text":"累计未分配利润","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"23":{"cells":{"0":{"text":"计算指标","merge":[2,0],"style":1},"1":{"text":"资产负债率","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"24":{"cells":{"0":{"style":2},"1":{"text":"流动比率","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"25":{"cells":{"0":{"style":2},"1":{"text":"速动比率","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1},"5":{"style":1},"6":{"style":1},"7":{"style":1},"8":{"style":1},"9":{"style":1}}},"len":26},"cols":{"1":{"width":160},"len":26},"validations":[],"autofilter":{}}
        xspreadsheetmap[13] = self.makexspreadsheet("13", xspreadsheetmap[13])
        xspreadsheetmap[13]["rows"]["23"]["cells"]["0"]["merge"] = [2,0]
        xspreadsheetmap[14] = {"name":"Sheet","freeze":"A1","styles":[{"align":"center"},{"align":"center","border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}}],"merges":["B2:B3","A2:A3"],"rows":{"0":{"cells":{"0":{"text":"盈亏平衡分析数据","merge":[],"style":1}}},"1":{"cells":{"0":{"text":"序号","merge":[1,0],"style":1},"1":{"text":"项目","merge":[1,0],"style":1}}},"2":{"cells":{"0":{"style":0},"1":{"style":1}}},"3":{"cells":{"0":{"style":1},"1":{"text":"年设计生产能力","style":1}}},"4":{"cells":{"0":{"text":"1","style":1},"1":{"text":"固定成本","style":1}}},"5":{"cells":{"0":{"text":"2","style":1},"1":{"text":"单位产品变动成本","style":1}}},"6":{"cells":{"0":{"text":"3","style":1},"1":{"text":"单位产品售价","style":1}}},"7":{"cells":{"0":{"text":"4","style":1},"1":{"text":"单位产品销售税金及附加","style":1}}},"8":{"cells":{"0":{"style":1},"1":{"style":1,"text":""}}},"9":{"cells":{"0":{"style":1,"text":"5"},"1":{"text":"盈亏平衡产量","style":1}}},"10":{"cells":{"0":{"text":"6","style":1},"1":{"text":"盈亏平衡销售价格","style":1}}},"11":{"cells":{"0":{"text":"7","style":1},"1":{"text":"盈亏平衡生产能力利用率","style":1}}},"12":{"cells":{"0":{"text":"6","style":1},"1":{"text":"流动资金利息","style":1}}},"len":12},"cols":{"0":{"width":75},"1":{"width":173},"len":6},"validations":[],"autofilter":{}}
        xspreadsheetmap[14] = self.makexspreadsheet("14", xspreadsheetmap[14])
        xspreadsheetmap[15] = {"name":"sheet6","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"}],"merges":["A1:D1","A3:A7","A8:A12","A13:A17"],"rows":{"0":{"cells":{"0":{"text":"单因素敏感性分析数据","style":1,"merge":[0,3]},"1":{"text":"undefined"},"2":{"text":"undefined"},"3":{"text":"undefined"}}},"1":{"cells":{"0":{"text":"因素","style":1},"1":{"text":"变动系数","style":1},"2":{"text":"NPV","style":1},"3":{"text":"变动率","style":1}}},"2":{"cells":{"0":{"text":"建设投资","style":1,"merge":[4,0]},"1":{"text":"1.0","style":1},"2":{"style":1},"3":{"style":1,"text":"0"}}},"3":{"cells":{"1":{"text":"1.2","style":1},"2":{"style":1},"3":{"style":1,"text":"20"}}},"4":{"cells":{"1":{"text":"1.1","style":1},"2":{"style":1},"3":{"style":1,"text":"10"}}},"5":{"cells":{"1":{"text":"0.9","style":1},"2":{"style":1},"3":{"style":1,"text":"-10"}}},"6":{"cells":{"1":{"text":"0.8","style":1},"2":{"style":1},"3":{"style":1,"text":"-20"}}},"7":{"cells":{"0":{"text":"经营成本","style":1,"merge":[4,0]},"1":{"text":"1.0","style":1},"2":{"style":1},"3":{"style":1,"text":"0"}}},"8":{"cells":{"1":{"text":"1.2","style":1},"2":{"style":1},"3":{"style":1,"text":"20"}}},"9":{"cells":{"1":{"text":"1.1","style":1},"2":{"style":1},"3":{"style":1,"text":"10"}}},"10":{"cells":{"1":{"text":"0.9","style":1},"2":{"style":1},"3":{"style":1,"text":"-10"}}},"11":{"cells":{"1":{"text":"0.8","style":1},"2":{"style":1},"3":{"style":1,"text":"-20"}}},"12":{"cells":{"0":{"text":"销售收入","style":1,"merge":[4,0]},"1":{"text":"1.0","style":1},"2":{"style":1},"3":{"style":1,"text":"0"}}},"13":{"cells":{"1":{"text":"1.2","style":1},"2":{"style":1},"3":{"style":1,"text":"20"}}},"14":{"cells":{"1":{"text":"1.1","style":1},"2":{"style":1},"3":{"style":1,"text":"10"}}},"15":{"cells":{"1":{"text":"0.9","style":1},"2":{"style":1},"3":{"style":1,"text":"-10"}}},"16":{"cells":{"1":{"text":"0.8","style":1},"2":{"style":1},"3":{"style":1,"text":"-20"}}},"len":17},"cols":{"len":4},"validations":[],"autofilter":{}}
        xspreadsheetmap[16] = {"name":"sheet6","freeze":"A1","styles":[{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]}},{"border":{"bottom":["thin","#000"],"top":["thin","#000"],"left":["thin","#000"],"right":["thin","#000"]},"align":"center"}],"merges":["A1:E1"],"rows":{"0":{"cells":{"0":{"text":"财务评价汇总表","style":1,"merge":[0,4]}}},"1":{"cells":{"0":{"text":"序号","style":1},"1":{"text":"指标名称","style":1},"2":{"text":"单位","style":1},"3":{"text":"指标值","style":1},"4":{"text":"备注","style":1}}},"2":{"cells":{"0":{"text":"一","style":1},"1":{"text":"经济数据","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"3":{"cells":{"0":{"text":"1","style":1},"1":{"text":"总投资","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"4":{"cells":{"0":{"text":"1.1","style":1},"1":{"text":"建设投资","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"5":{"cells":{"0":{"text":"1.1.1","style":1},"1":{"text":"建筑工程","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"6":{"cells":{"0":{"text":"1.1.2","style":1},"1":{"text":"设备及安装工程","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"7":{"cells":{"0":{"text":"1.1.3","style":1},"1":{"text":"预备费","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"8":{"cells":{"0":{"text":"1.1.4","style":1},"1":{"text":"其他费用","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"9":{"cells":{"0":{"text":"1.2","style":1},"1":{"text":"建设期利息","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"10":{"cells":{"0":{"text":"1.3","style":1},"1":{"text":"流动资金","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"11":{"cells":{"0":{"text":"2","style":1},"1":{"text":"项目资金来源","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"12":{"cells":{"0":{"text":"2.1","style":1},"1":{"text":"资本金","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"13":{"cells":{"0":{"text":"2.2","style":1},"1":{"text":"建设投资借款","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"14":{"cells":{"0":{"text":"2.3","style":1},"1":{"text":"流动资金借款","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"15":{"cells":{"0":{"text":"3","style":1},"1":{"text":"营业收入","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"16":{"cells":{"0":{"text":"4","style":1},"1":{"text":"营业税金及附加","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"17":{"cells":{"0":{"text":"4.1","style":1},"1":{"text":"增值税","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"18":{"cells":{"0":{"text":"4.2","style":1},"1":{"text":"城市维护建设税","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"19":{"cells":{"0":{"text":"4.3","style":1},"1":{"text":"教育费附加","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"20":{"cells":{"0":{"text":"5","style":1},"1":{"text":"总成本费用","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"21":{"cells":{"0":{"text":"5.1","style":1},"1":{"text":"其中：固定成本","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"22":{"cells":{"0":{"text":"5.2","style":1},"1":{"text":"可变成本","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"23":{"cells":{"0":{"text":"5.3","style":1},"1":{"text":"经营成本","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"24":{"cells":{"0":{"text":"6","style":1},"1":{"text":"利润总额","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"25":{"cells":{"0":{"text":"7","style":1},"1":{"text":"所得税","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"26":{"cells":{"0":{"text":"8","style":1},"1":{"text":"税后利润","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"27":{"cells":{"0":{"text":"9","style":1},"1":{"text":"提取盈余公积金","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"28":{"cells":{"0":{"text":"10","style":1},"1":{"text":"提取盈余公益金","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"text":"合计","style":1}}},"29":{"cells":{"0":{"text":"二","style":1},"1":{"text":"财务评价指标","style":1},"2":{"text":"单位","style":1},"3":{"style":1},"4":{"style":1}}},"30":{"cells":{"0":{"text":"1","style":1},"1":{"text":"总投资收益率","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"text":"正常年","style":1}}},"31":{"cells":{"0":{"text":"2","style":1},"1":{"text":"资本金净利润率","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"text":"正常年","style":1}}},"32":{"cells":{"0":{"text":"3","style":1},"1":{"text":"全部投资财务指标","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"33":{"cells":{"0":{"text":"3.1","style":1},"1":{"text":"所得税前","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"34":{"cells":{"0":{"text":"3.1.1","style":1},"1":{"text":"内部收益率","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"style":1}}},"35":{"cells":{"0":{"text":"3.1.2","style":1},"1":{"text":"净现值","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"36":{"cells":{"0":{"text":"3.1.3","style":1},"1":{"text":"静态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"37":{"cells":{"0":{"text":"3.1.4","style":1},"1":{"text":"动态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"38":{"cells":{"0":{"text":"3.2","style":1},"1":{"text":"所得税后","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"39":{"cells":{"0":{"text":"3.2.1","style":1},"1":{"text":"内部收益率","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"style":1}}},"40":{"cells":{"0":{"text":"3.2.2","style":1},"1":{"text":"净现值","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"41":{"cells":{"0":{"text":"3.2.3","style":1},"1":{"text":"静态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"42":{"cells":{"0":{"text":"3.2.4","style":1},"1":{"text":"动态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"43":{"cells":{"0":{"text":"4","style":1},"1":{"text":"资本金投资财务指标","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"44":{"cells":{"0":{"text":"4.1","style":1},"1":{"text":"所得税前","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"45":{"cells":{"0":{"text":"4.1.1","style":1},"1":{"text":"内部收益率","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"style":1}}},"46":{"cells":{"0":{"text":"4.1.2","style":1},"1":{"text":"净现值","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"47":{"cells":{"0":{"text":"4.1.3","style":1},"1":{"text":"静态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"48":{"cells":{"0":{"text":"4.1.4","style":1},"1":{"text":"动态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"49":{"cells":{"0":{"text":"4.2","style":1},"1":{"text":"所得税后","style":1},"2":{"style":1},"3":{"style":1},"4":{"style":1}}},"50":{"cells":{"0":{"text":"4.2.1","style":1},"1":{"text":"内部收益率","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"style":1}}},"51":{"cells":{"0":{"text":"4.2.2","style":1},"1":{"text":"净现值","style":1},"2":{"text":"万元","style":1},"3":{"style":1},"4":{"style":1}}},"52":{"cells":{"0":{"text":"4.2.3","style":1},"1":{"text":"静态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"53":{"cells":{"0":{"text":"4.2.4","style":1},"1":{"text":"动态投资回收期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"54":{"cells":{"0":{"text":"5","style":1},"1":{"text":"盈亏平衡点(生产能力利用率)","style":1},"2":{"text":"%","style":1},"3":{"style":1},"4":{"text":"正常年","style":1}}},"55":{"cells":{"0":{"text":"6","style":1},"1":{"text":"借款偿还期","style":1},"2":{"text":"年","style":1},"3":{"style":1},"4":{"style":1}}},"56":{"cells":{"0":{"text":"7","style":1},"1":{"text":"利息备付率","style":1},"2":{"style":1},"3":{"style":1},"4":{"text":"正常年","style":1}}},"57":{"cells":{"0":{"text":"8","style":1},"1":{"text":"偿债备付率","style":1},"2":{"style":1},"3":{"style":1},"4":{"text":"正常年","style":1}}},"len":58},"cols":{"1":{"width":180},"len":5},"validations":[],"autofilter":{}}


        return f'{workbookname}.xlsx',startrows,xspreadsheetmap


if __name__ == '__main__':
    # renderexcel("D:/Program Data/flask-html/表1.xlsx")


    data = {'projectname1': '示例8_2', 'projectplace': '示例', 'projectind': '示例', 'input1': '8', 'input2': '2', 'productionLoad3': '60',
            'productionLoad4': '80', 'productionLoad5': '100', 'productionLoad6': '100', 'productionLoad7': '100',
            'productionLoad8': '100', 'benchmarkyield': '10', 'benchmark_s_paybackperiod': '4.5', 'benchmark_d_paybackperiod': '5',
            'install_rate': '10', 'trans_rate': '5', 'pre_rate': '8', 'preup_rate': '9', 'additionalCostName_1': '主要项目',
            'additionalCostAmount_1': '9000', 'additionalCostPrice_1': '400', 'additionalCostEquipment_1': '800', 'additionalCostName_2': '辅助项目',
            'additionalCostAmount_2': '2000', 'additionalCostPrice_2': '400', 'additionalCostEquipment_2': '280', 'additionalCostName_3': '公共工程',
            'additionalCostAmount_3': '1000', 'additionalCostPrice_3': '400', 'additionalCostEquipment_3': '444', 'additionalCostName_4': '环境工程',
            'additionalCostAmount_4': '1000', 'additionalCostPrice_4': '400', 'additionalCostEquipment_4': '33', 'additionalOtherCost_1': '4555',
            'additionalOtherCost_2': '900', 'additionalOtherCost_3': '2222', 'additionalotherCostName_1': '土地征用费', 'additionalotherCostName_2': '开办费', 'additionalotherCostName_3': '专利技术使用费', 'days1': '60', 'days2': '30', 'days3': '30', 'days4': '20', 'days5': '30',
            'days6': '30', 'days7': '60', 'days8': '4', 'days9': '70', 'loadrate': '5', 'repayMethod': 'ave_capital', 'payTime': '4', 'investAmount_1': '5011',
            'investPercentage_1': '46', 'loanAmount_1': '2756', 'loanPercentage_1': '55', 'dep_year1': '10', 'res_rate1': '10', 'dep_year2': '10',
            'res_rate2': '10', 'dep_year3': '10', 'res_rate3': '0', 'depreciationMethod': 'residualValue', 'amo_year1': '4', 'amo_year2': '4',
            'production1': '100', 'dan1': '吨', 'production2': '60', 'dan2': '万元', 'production3': '100', 'dan3': '吨', 'production4': '3', 'dan4': '万元',
            'production5': '100', 'production6': '10', 'dan6': '万元', 'production7': '20', 'production8': '5', 'production9': '150', 'production10': '140',
            'production11': '17', 'production12': '4', 'production13': '2', 'production14': '25', 'production15': '10', 'production16': '10',
            'projectcostCounter': '4', 'othercostCounter': '3', 'constInvenstCounter': '1',
            'additionalCosts': '{"主要项目":"9000","辅助项目":"2000","公共工程":"1000","环境工程":"1000","土地征用费":"4555","开办费":"900","专利技术使用费":"2222"}',
            'investmentPlan': '{"1":{"amount":"5011","percentage":"46"}}', 'LoanPlan': '{"1":{"amount":"2756","percentage":"55"},"2":{"amount":"2647","percentage":"45"}}',
            'leftInvestment': '5883', 'leftInvestmentPercentage': '54.0'}

    data_test = {'projectname': '示例_8_3', 'projectplace': '示例', 'projectind': '示例', 'input1': '8', 'input2': '3',
            'productionLoad4': '80', 'productionLoad5': '100', 'productionLoad6': '100', 'productionLoad7': '100',
            'productionLoad8': '100', 'benchmarkyield': '10', 'benchmark_s_paybackperiod': '4.5',
            'benchmark_d_paybackperiod': '5',
            'install_rate': '10', 'trans_rate': '5', 'pre_rate': '8', 'preup_rate': '9', 'additionalCostName_1': '主要项目',
            'additionalCostAmount_1': '9000', 'additionalCostPrice_1': '400', 'additionalCostEquipment_1': '800',
            'additionalCostName_2': '辅助项目',
            'additionalCostAmount_2': '2000', 'additionalCostPrice_2': '400', 'additionalCostEquipment_2': '280',
            'additionalCostName_3': '公共工程',
            'additionalCostAmount_3': '1000', 'additionalCostPrice_3': '400', 'additionalCostEquipment_3': '444',
            'additionalCostName_4': '环境工程',
            'additionalCostAmount_4': '1000', 'additionalCostPrice_4': '400', 'additionalCostEquipment_4': '33',
            'additionalOtherCost_1': '4555',
            'additionalOtherCost_2': '900', 'additionalOtherCost_3': '2222', 'additionalotherCostName_1': '土地征用费',
            'additionalotherCostName_2': '开办费', 'additionalotherCostName_3': '专利技术使用费', 'days1': '60', 'days2': '30',
            'days3': '30', 'days4': '20', 'days5': '30',
            'days6': '30', 'days7': '60', 'days8': '4', 'days9': '70', 'loadrate': '5', 'repayMethod': 'ave_capital',
            'payTime': '4', 'investAmount_1': '5011',
            'investPercentage_1': '46', 'loanAmount_1': '2756', 'loanPercentage_1': '55', 'dep_year1': '10',
            'res_rate1': '10', 'dep_year2': '10',
            'res_rate2': '10', 'dep_year3': '10', 'res_rate3': '0', 'depreciationMethod': 'residualValue',
            'amo_year1': '4', 'amo_year2': '4',
            'production1': '100', 'dan1': '吨', 'production2': '60', 'dan2': '万元', 'production3': '100', 'dan3': '吨',
            'production4': '3', 'dan4': '万元',
            'production5': '100', 'production6': '10', 'dan6': '万元', 'production7': '20', 'production8': '5',
            'production9': '150', 'production10': '140',
            'production11': '17', 'production12': '4', 'production13': '2', 'production14': '25', 'production15': '10',
            'production16': '10',
            'projectcostCounter': '4', 'othercostCounter': '3', 'constInvenstCounter': '1',
            'additionalCosts': '{"主要项目":"9000","辅助项目":"2000","公共工程":"1000","环境工程":"1000","土地征用费":"4555","开办费":"900","专利技术使用费":"2222"}',
            'investmentPlan': '{"1":{"amount":"5011","percentage":"46"}}',
            'LoanPlan': '{"1":{"amount":"2756","percentage":"55"},"2":{"amount":"2647","percentage":"45"}}',
            'leftInvestment': '5883', 'leftInvestmentPercentage': '54.0'}

    data_s = {'projectname': '', 'projectplace': '', 'projectind': '', 'input1': '', 'input2': '', 'benchmarkyield': '', 'benchmark_s_paybackperiod': '', 'benchmark_d_paybackperiod': '', 'install_rate': '', 'trans_rate': '', 'pre_rate': '', 'preup_rate': '', 'days1': '', 'days2': '', 'days3': '', 'days4': '', 'days5': '', 'days6': '', 'days7': '', 'days8': '', 'days9': '', 'loadrate': '', 'payTime': '', 'loanAmount_1': '', 'loanPercentage_1': '', 'dep_year1': '', 'res_rate1': '', 'dep_year2': '', 'res_rate2': '', 'dep_year3': '', 'res_rate3': '', 'amo_year1': '', 'amo_year2': '', 'production1': '', 'dan1': '', 'production2': '', 'dan2': '', 'production3': '', 'dan3': '', 'production4': '', 'dan4': '', 'production5': '', 'production6': '', 'dan6': '', 'production7': '', 'production8': '', 'production9': '', 'production10': '', 'production11': '', 'production12': '', 'production13': '', 'production14': '', 'production15': '', 'production16': '', 'depreciationMethod': 'undefined', 'projectcostCounter': '0', 'othercostCounter': '0', 'constInvenstCounter': '0', 'additionalCosts': '{}', 'investmentPlan': '{}', 'LoanPlan': '{"1":{"amount":"","percentage":""}}', 'leftInvestment': 'NaN', 'leftInvestmentPercentage': 'NaN'}

    data_book = {'projectname': '教材检测', 'projectplace': '示例', 'projectind': '示例', 'input1': '6', 'input2': '2',
                 'productionLoad3': '60', 'productionLoad4': '80', 'productionLoad5': '100', 'productionLoad6': '100',
                 'benchmarkyield': '10', 'benchmark_s_paybackperiod': '5', 'benchmark_d_paybackperiod': '5.5',
                 'install_rate': '10', 'trans_rate': '0', 'pre_rate': '10', 'preup_rate': '10',
                 'additionalCostName_1': '主要项目', 'additionalCostAmount_1': '100', 'additionalCostPrice_1': '100000',
                 'additionalCostEquipment_1': '200', 'additionalCostName_2': '辅助项目', 'additionalCostAmount_2': '100',
                 'additionalCostPrice_2': '100000', 'additionalCostEquipment_2': '200',
                 'additionalotherCostName_1': '土地征用费', 'additionalOtherCostType_1': 'tudi',
                 'additionalOtherCost_1': '2000', 'additionalotherCostName_2': '开办费',
                 'additionalOtherCostType_2': 'other', 'additionalOtherCost_2': '800', 'days1': '60', 'days2': '30',
                 'days3': '30', 'days4': '20', 'days5': '30', 'days6': '30', 'days7': '60', 'days8': '4', 'days9': '70',
                 'loadrate': '5', 'repayMethod': 'ave_capital', 'payTime': '4', 'investAmount_1': '3004',
                 'investPercentage_1': '50', 'loanAmount_1': '1202', 'loanPercentage_1': '40', 'dep_year1': '10',
                 'res_rate1': '10', 'dep_year2': '10', 'res_rate2': '10', 'dep_year3': '10', 'res_rate3': '0',
                 'depreciationMethod': 'residualValue', 'amo_year1': '4', 'amo_year2': '4', 'production1': '100',
                 'dan1': '', 'production2': '60', 'dan2': '', 'production3': '100', 'dan3': '', 'production4': '3',
                 'dan4': '', 'production5': '100', 'production6': '10', 'dan6': '', 'production7': '20',
                 'production8': '5', 'production9': '100', 'production10': '140', 'production11': '17',
                 'production12': '4', 'production13': '2', 'production14': '25', 'production15': '10',
                 'production16': '10', 'projectcostCounter': '2', 'othercostCounter': '2', 'constInvenstCounter': '1',
                 'additionalCosts': {"主要项目": "100", "辅助项目": "100", "土地征用费": "2000", "开办费": "800"},
                 'investmentPlan': '{"1":{"amount":"3004","percentage":"50"}}',
                 'LoanPlan': '{"1":{"amount":"1202","percentage":"40"},"2":{"amount":"1202","percentage":"40"}}',
                 'leftInvestment': '3004', 'leftInvestmentPercentage': '50.0'}

    data_secure = {'projectname': '示例', 'projectplace': '示例', 'projectind': '示例', 'input1': '6', 'input2': '2', 'productionLoad3': '60', 'productionLoad4': '80', 'productionLoad5': '100', 'productionLoad6': '100', 'benchmarkyield': '1', 'benchmark_s_paybackperiod': '4', 'benchmark_d_paybackperiod': '4', 'install_rate': '10', 'trans_rate': '10', 'pre_rate': '10', 'preup_rate': '10', 'additionalCostName_1': '主要项目', 'additionalCostAmount_1': '1000', 'additionalCostPrice_1': '100000', 'additionalCostEquipment_1': '10000', 'additionalCostName_2': '辅助项目', 'additionalCostAmount_2': '1000', 'additionalCostPrice_2': '100000', 'additionalCostEquipment_2': '10000', 'additionalotherCostName_1': '土地征用费', 'additionalOtherCost_1': '100000', 'additionalOtherCostType_1': '', 'days1': '20', 'days2': '20', 'days3': '20', 'days4': '20', 'days5': '20', 'days6': '20', 'days7': '20', 'days8': '20', 'days9': '20', 'loadrate': '5', 'repayMethod': 'ave_capital', 'payTime': '4', 'loanAmount_1': '100000', 'loanPercentage_1': '62.4', 'dep_year1': '10', 'res_rate1': '10', 'dep_year2': '10', 'res_rate2': '10', 'dep_year3': '10', 'depreciationMethod': 'SYD', 'amo_year1': '4', 'amo_year2': '4', 'production1': '100', 'production2': '60', 'production3': '100', 'production4': '3', 'production5': '100', 'production6': '10', 'production7': '20', 'production8': '5', 'production9': '100', 'production10': '140', 'production11': '17', 'production12': '4', 'production13': '2', 'production14': '25', 'production15': '10', 'production16': '10', 'projectcostCounter': '2', 'othercostCounter': '1', 'constInvenstCounter': '0', 'additionalCosts': '{"主要项目":"1000","辅助项目":"1000","土地征用费":"100000"}', 'investmentPlan': '{}', 'LoanPlan': '{"1":{"amount":"100000","percentage":"62.4"}}', 'leftInvestment': '160160', 'leftInvestmentPercentage': '100.0'}

    data = {'projectname': '11', 'projectplace': '11', 'projectind': '11', 'input1': '8', 'input2': '3',
            'productionLoad4': '70', 'productionLoad5': '80', 'productionLoad6': '90',
            'productionLoad7': '100', 'productionLoad8': '100', 'benchmarkyield': '10',
            'benchmark_s_paybackperiod': '6', 'benchmark_d_paybackperiod': '6', 'install_rate': '10', 'trans_rate': '0',
            'pre_rate': '10', 'preup_rate': '10', 'additionalCostName_1': '11', 'additionalCostAmount_1': '1000',
            'additionalCostPrice_1': '10000', 'additionalCostEquipment_1': '1000', 'additionalCostName_2': '22',
            'additionalCostAmount_2': '1000', 'additionalCostPrice_2': '10000', 'additionalCostEquipment_2': '1000',
            'additionalotherCostName_1': '33', 'additionalOtherCost_1': '1000', 'additionalOtherCostType_1': 'tudi',
            'additionalotherCostName_2': '44', 'additionalOtherCost_2': '800',
            'additionalOtherCostType_2': 'intangible', 'additionalotherCostName_3': '55',
            'additionalOtherCost_3': '900', 'additionalOtherCostType_3': 'other', 'days1': '30', 'days2': '20',
            'days3': '20', 'days4': '20', 'days5': '30', 'days6': '30', 'days7': '20', 'days8': '5', 'days9': '70',
            'loadrate': '4', 'repayMethod': 'ave_cap_int', 'payTime': '5', 'investAmount_1': '4005',
            'investPercentage_1': '50', 'loanAmount_1': '2403', 'loanPercentage_1': '60', 'dep_year1': '10',
            'res_rate1': '10', 'dep_year2': '10', 'res_rate2': '10', 'dep_year3': '10', 'depreciationMethod': 'DDB',
            'amo_year1': '5', 'amo_year2': '5', 'production1': '100', 'production2': '60', 'production3': '100',
            'production4': '3', 'production5': '100', 'production6': '10', 'production7': '10', 'production8': '4',
            'production9': '100', 'production10': '140', 'production11': '17', 'production12': '4', 'production13': '2',
            'production14': '25', 'production15': '10', 'production16': '10', 'projectcostCounter': '2',
            'othercostCounter': '3', 'constInvenstCounter': '1',
            'additionalCosts': '{"11":"1000","22":"1000","33":"1000","44":"800","55":"900"}',
            'investmentPlan': '{"1":{"amount":"4005","percentage":"50"}}',
            'LoanPlan': '{"1":{"amount":"2403","percentage":"60"},"2":{"amount":"2403","percentage":"60"}}',
            'leftInvestment': '4005', 'leftInvestmentPercentage': '50.0'}

    f = Fdata(data)
    latest_file, sheetmap, xspreadsheetmap = f.toCalExcel()
    s = renderexcel(latest_file, sheetmap, xspreadsheetmap)
    print(s)

    # from win32com.client import Dispatch
    #
    # xlApp = Dispatch("Excel.Application")
    # xlApp.Visible = False
    # xlBook = xlApp.Workbooks.Open("D:/Program Data/flask-html/test.xlsx")
    # xlBook.Save()
    # xlBook.Close()
