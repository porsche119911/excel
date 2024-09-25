import logging
from xlcalculator import ModelCompiler, Model, Evaluator
import openpyxl
from openpyxl.utils import get_column_letter
import re


def remove_absolute_reference_in_excel(input_file):
    """
    读取Excel文件，遍历所有单元格，移除公式中的$符号
    """
    workbook = openpyxl.load_workbook(input_file, keep_vba=False)  # 打开Excel文件

    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith('='):  # 如果单元格包含公式
                    # 移除公式中的$符号
                    cell.value = cell.value.replace('$', '')
                    cell.value = cell.value.replace(' ', '')

    # 创建一个中间文件保存移除$符号后的结果
    modified_file = "modified_" + input_file
    workbook.save(modified_file)

    return modified_file

def convert_to_excel_value(value):
    """
    将xlcalculator的结果转换为Excel可接受的值
    """
    if hasattr(value, 'value'):  # 如果值有 'value' 属性
        return value.value
    elif isinstance(value, (int, float, str, bool)):  # 基本类型直接返回
        return value
    else:  # 其他类型转为字符串
        return str(value)

def calculate_and_save_excel(input_file, output_file):
    """
    计算给定Excel文件中的公式，并将结果保存到新的Excel文件。
    """
    logging.basicConfig(level=logging.INFO)

    # 先移除公式中的$符号
    modified_input_file = remove_absolute_reference_in_excel(input_file)

    # 使用修改后的文件进行解析和计算
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(modified_input_file)
    evaluator = Evaluator(model)

    new_workbook = openpyxl.Workbook()
    created_sheets = set()

    for cell_address, cell in model.cells.items():
        sheet_name, address = cell_address.split('!')

        if sheet_name not in created_sheets:
            if sheet_name in new_workbook.sheetnames:
                new_sheet = new_workbook[sheet_name]
            else:
                new_sheet = new_workbook.create_sheet(sheet_name)
            created_sheets.add(sheet_name)
        else:
            new_sheet = new_workbook[sheet_name]

        # 正确解析单元格的列和行
        column_letter = ''.join(re.findall(r'[A-Za-z]', address))
        row = int(''.join(re.findall(r'[0-9]', address)))
        # column_letter, row = address[0], int(address[1:])
        column = openpyxl.utils.column_index_from_string(column_letter)

        if cell.formula:
            try:
                result = evaluator.evaluate(cell_address)
                converted_result = convert_to_excel_value(result)
                new_sheet.cell(row=row, column=column, value=converted_result)
                print(f"计算结果 {cell_address}: {converted_result}")
            except Exception as e:
                print(f"计算 {cell_address} 时出错: {str(e)}")
                new_sheet.cell(row=row, column=column, value=f"ERROR: {str(e)}")
        else:
            new_sheet.cell(row=row, column=column, value=convert_to_excel_value(cell.value))

    if "Sheet" in new_workbook.sheetnames:
        new_workbook.remove(new_workbook["Sheet"])

    new_workbook.save(output_file)
    print(f"结果已保存到 {output_file}")

import pycel as formulas

def calculate_excel(input_file, output_file):
    # 加载 Excel 文件
    wb = formulas.ExcelModel().loads(input_file).finish()

    # 计算所有公式
    evaluator = formulas.Parser()
    wb.calculate()

    # 保存结果
    wb.write(output_file)

# 使用示例
calculate_excel('input.xlsx', 'output_calculated.xlsx')



# 使用示例
# calculate_and_save_excel("sheetdata.xlsx", "calculated_result.xlsx")